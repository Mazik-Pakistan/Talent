"""Parse recruiter roster spreadsheets (.xlsx / .csv) into row dicts."""

from __future__ import annotations

import csv
import io
import re
from datetime import date, datetime
from typing import Any

from fastapi import HTTPException, UploadFile

from app.schemas.offer import DEFAULT_BENEFITS

BULK_INVITE_MAX_ROWS = 200
BULK_INVITE_MAX_BYTES = 5 * 1024 * 1024

# Canonical column → accepted header aliases (lowercased).
COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "email": ("email", "email address", "candidate email", "email *"),
    "full_name": ("full_name", "name", "candidate name", "full name", "full name *"),
    "job_title": (
        "job_title",
        "designation",
        "title",
        "job title",
        "position",
        "job title *",
        "designation *",
    ),
    "department": ("department", "dept", "department *"),
    "reporting_manager": (
        "reporting_manager",
        "manager",
        "reporting manager",
        "manager name",
        "reporting manager *",
    ),
    "start_date": (
        "start_date",
        "start date",
        "joining date",
        "join date",
        "start date *",
        "start date (yyyy-mm-dd)",
        "start date * (yyyy-mm-dd)",
    ),
    "monthly_salary": (
        "monthly_salary",
        "salary",
        "monthly salary",
        "gross salary",
        "ctc",
        "monthly salary *",
    ),
    "office_location": ("office_location", "office", "location", "office location"),
    "employment_type": ("employment_type", "employment type", "type"),
    "is_remote": ("is_remote", "remote", "remote employee", "work mode", "work_mode"),
    "currency": ("currency", "curr"),
    "offer_expiry_days": ("offer_expiry_days", "offer expiry", "offer expiry days"),
    "message_to_candidate": (
        "message_to_candidate",
        "message",
        "candidate message",
        "note",
        "message to candidate",
    ),
    "terms": ("terms", "offer terms", "offer letter terms"),
    "benefits": ("benefits", "benefit list"),
}

REQUIRED_COLUMNS = (
    "email",
    "full_name",
    "job_title",
    "department",
    "reporting_manager",
    "start_date",
    "monthly_salary",
)

# Default salary breakdown components (same as manual invite form).
DEFAULT_PAY_COMPONENTS = (
    "Basic",
    "Housing",
    "Transport",
    "Medical allowance",
    "Other",
)

# Human-friendly template columns (core fields — pay + benefits added separately).
TEMPLATE_CORE_HEADERS = [
    "Email *",
    "Full name *",
    "Job title *",
    "Department *",
    "Reporting manager *",
    "Start date * (YYYY-MM-DD)",
    "Monthly salary *",
]

TEMPLATE_OFFER_META_HEADERS = [
    "Office location",
    "Employment type",
    "Remote (Yes/No)",
    "Currency",
    "Offer expiry days",
    "Message to candidate",
    "Offer terms",
]

TEMPLATE_SAMPLE_CORE = [
    "jane.doe@example.com",
    "Jane Doe",
    "Software Engineer",
    "Engineering",
    "Alex Manager",
    "2026-09-01",
    "150000",
]

TEMPLATE_SAMPLE_PAY = {
    "Basic": 90000,
    "Housing": 40000,
    "Transport": 20000,
    "Medical allowance": 0,
    "Other": 0,
}

TEMPLATE_SAMPLE_META = [
    "Karachi",
    "Full-time",
    "No",
    "PKR",
    "14",
    "Welcome to the team!",
    "",  # offer terms — blank uses standard company terms
]

# Sample row: most benefits Yes, a couple No — shows recruiters how to tick.
TEMPLATE_SAMPLE_BENEFIT_FLAGS = {
    "Medical insurance": "Yes",
    "Provident fund": "Yes",
    "Annual leave": "Yes",
    "Hybrid / remote flexibility": "Yes",
    "Company laptop": "Yes",
    "Learning & training budget": "No",
    "Fuel / conveyance allowance": "No",
    "Performance bonus eligibility": "Yes",
}

_YES_VALUES = {
    "yes",
    "y",
    "true",
    "1",
    "x",
    "✓",
    "✔",
    "☑",
    "tick",
    "checked",
    "include",
    "on",
}
_NO_VALUES = {
    "no",
    "n",
    "false",
    "0",
    "",
    "✗",
    "✘",
    "☐",
    "cross",
    "unchecked",
    "off",
    "-",
}


def _slug_benefit(label: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", (label or "").lower()).strip("-")


def benefit_column_header(label: str) -> str:
    """Excel header for a benefit Yes/No column."""
    return f"Benefit: {label}"


def pay_column_header(label: str) -> str:
    """Excel header for a salary-breakdown amount column."""
    return f"Pay: {label}"


def _normalize_header(value: Any) -> str:
    return str(value or "").strip().lower()


def resolve_pay_columns(header: list[str]) -> list[tuple[str, int]]:
    """Map Pay: Basic / Basic amount columns → (component label, index)."""
    known = {_normalize_header(c): c for c in DEFAULT_PAY_COMPONENTS}
    found: list[tuple[str, int]] = []
    seen: set[str] = set()

    for idx, raw in enumerate(header):
        h = (raw or "").strip()
        if not h:
            continue
        label = None
        if h.startswith("pay:"):
            label = h.split(":", 1)[1].strip()
        elif h.startswith("salary:"):
            label = h.split(":", 1)[1].strip()
        elif h.endswith(" amount"):
            label = h[: -len(" amount")].strip()
        elif h in known:
            # Bare "basic" / "housing" — only if not a core field alias
            core_hit = any(h in aliases for aliases in COLUMN_ALIASES.values())
            if not core_hit:
                label = known[h]

        if not label:
            continue
        canonical = known.get(_normalize_header(label)) or " ".join(label.split()).title()
        # Prefer exact DEFAULT casing
        for default in DEFAULT_PAY_COMPONENTS:
            if default.lower() == canonical.lower():
                canonical = default
                break
        if canonical.lower() in seen:
            continue
        seen.add(canonical.lower())
        found.append((canonical, idx))

    return found


def _is_truthy_flag(raw: Any) -> bool:
    text = _cell_str(raw).lower()
    if text in _YES_VALUES:
        return True
    if text in _NO_VALUES:
        return False
    # Boolean / numeric from Excel
    if isinstance(raw, bool):
        return raw
    if isinstance(raw, (int, float)) and not isinstance(raw, bool):
        return raw != 0
    return False


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalize_start_date(raw: str) -> str:
    text = (raw or "").strip()
    if not text:
        return ""
    if " " in text:
        text = text.split(" ", 1)[0]
    if "T" in text:
        text = text.split("T", 1)[0]
    for sep in ("/", "-"):
        parts = text.split(sep)
        if len(parts) == 3 and len(parts[0]) <= 2 and len(parts[2]) == 4:
            try:
                day, month, year = int(parts[0]), int(parts[1]), int(parts[2])
                if day <= 31 and month <= 12:
                    return f"{year:04d}-{month:02d}-{day:02d}"
            except ValueError:
                pass
    return text


def _parse_salary(raw: str) -> float | None:
    text = (raw or "").strip().replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _pick_candidates_sheet(workbook):
    """Prefer the Candidates / Bulk invites sheet over Instructions."""
    preferred = ("candidates", "bulk invites", "roster", "invites")
    by_name = {str(ws.title or "").strip().lower(): ws for ws in workbook.worksheets}
    for name in preferred:
        if name in by_name:
            return by_name[name]
    # Fall back to first sheet whose header row looks like a roster.
    for ws in workbook.worksheets:
        first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not first:
            continue
        headers = [_normalize_header(h) for h in first]
        if any(h in {"email", "email *", "email address"} for h in headers):
            return ws
    return workbook.active


async def read_upload_rows(file: UploadFile) -> tuple[list[str], list[tuple[int, list[Any]]]]:
    """Return (header, [(sheet_row_number, cells), ...])."""
    filename = (file.filename or "").lower()
    is_csv = filename.endswith(".csv")
    is_xlsx = filename.endswith(".xlsx") or filename.endswith(".xlsm")
    if not (is_csv or is_xlsx):
        raise HTTPException(status_code=400, detail="Please upload a .xlsx or .csv spreadsheet.")

    raw = await file.read()
    if len(raw) > BULK_INVITE_MAX_BYTES:
        raise HTTPException(status_code=400, detail="File is too large (5MB limit).")

    try:
        if is_csv:
            text = raw.decode("utf-8-sig", errors="replace")
            reader = csv.reader(io.StringIO(text))
            rows = list(reader)
            if not rows:
                raise HTTPException(status_code=400, detail="The CSV file is empty.")
            header = [_normalize_header(h) for h in rows[0]]
            data = [(i, list(row)) for i, row in enumerate(rows[1:], start=2)]
            return header, data

        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        sheet = _pick_candidates_sheet(workbook)
        rows_iter = sheet.iter_rows(values_only=True)
        first = next(rows_iter, None)
        if first is None:
            raise HTTPException(status_code=400, detail="The spreadsheet is empty.")
        header = [_normalize_header(h) for h in first]
        data = [(i, list(row)) for i, row in enumerate(rows_iter, start=2)]
        return header, data
    except HTTPException:
        raise
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=400, detail=f"Could not read the spreadsheet: {exc}") from exc


def resolve_column_indexes(header: list[str]) -> dict[str, int | None]:
    indexes: dict[str, int | None] = {}
    for canonical, aliases in COLUMN_ALIASES.items():
        indexes[canonical] = None
        for alias in aliases:
            if alias in header:
                indexes[canonical] = header.index(alias)
                break
    return indexes


def resolve_benefit_columns(header: list[str]) -> list[tuple[str, int]]:
    """Map benefit Yes/No columns → (display label, column index).

    Accepts:
      - Benefit: Medical insurance
      - Medical insurance? / Medical insurance (Yes/No)
      - Exact DEFAULT_BENEFITS labels
    """
    known = { _normalize_header(b): b for b in DEFAULT_BENEFITS }
    known_slugs = { _slug_benefit(b): b for b in DEFAULT_BENEFITS }
    found: list[tuple[str, int]] = []
    seen_labels: set[str] = set()

    for idx, raw in enumerate(header):
        h = (raw or "").strip()
        if not h:
            continue
        label = None
        if h.startswith("benefit:"):
            label = h.split(":", 1)[1].strip()
        elif h.endswith("?"):
            label = h[:-1].strip()
        elif h.endswith("(yes/no)"):
            label = h[: -len("(yes/no)")].strip()
        elif h in known:
            label = known[h]
        elif _slug_benefit(h) in known_slugs and h not in {
            a for aliases in COLUMN_ALIASES.values() for a in aliases
        }:
            # Only treat bare known benefit names as benefit cols when not a core field alias.
            label = known_slugs[_slug_benefit(h)]

        if not label:
            continue
        # Prefer canonical casing from DEFAULT_BENEFITS when matched.
        canonical = known.get(_normalize_header(label)) or known_slugs.get(_slug_benefit(label)) or label
        if canonical in seen_labels:
            continue
        # Skip if this header was already claimed as a core field.
        core_hit = False
        for aliases in COLUMN_ALIASES.values():
            if h in aliases:
                core_hit = True
                break
        if core_hit:
            continue
        seen_labels.add(canonical)
        found.append((canonical, idx))

    return found


def missing_required_headers(indexes: dict[str, int | None]) -> list[str]:
    missing = []
    for col in REQUIRED_COLUMNS:
        if indexes.get(col) is None:
            if col == "full_name":
                missing.append("Full name")
            elif col == "job_title":
                missing.append("Job title")
            elif col == "reporting_manager":
                missing.append("Reporting manager")
            elif col == "monthly_salary":
                missing.append("Monthly salary")
            elif col == "start_date":
                missing.append("Start date")
            else:
                missing.append(col.replace("_", " ").title())
    return missing


def row_to_candidate(
    row: list[Any],
    indexes: dict[str, int | None],
    sheet_row: int,
    benefit_columns: list[tuple[str, int]] | None = None,
    pay_columns: list[tuple[str, int]] | None = None,
) -> dict:
    def get(key: str) -> str:
        idx = indexes.get(key)
        if idx is None or idx >= len(row):
            return ""
        return _cell_str(row[idx])

    email = get("email").lower()
    full_name = get("full_name")
    job_title = get("job_title")
    department = get("department")
    reporting_manager = get("reporting_manager")
    start_date = _normalize_start_date(get("start_date"))
    salary_raw = get("monthly_salary")
    monthly_salary = _parse_salary(salary_raw)

    missing_fields: list[str] = []
    for label, value in (
        ("email", email),
        ("full_name", full_name),
        ("job_title", job_title),
        ("department", department),
        ("reporting_manager", reporting_manager),
        ("start_date", start_date),
    ):
        if not value or (label != "email" and len(value) < 2):
            missing_fields.append(label)
    if monthly_salary is None or monthly_salary < 0:
        missing_fields.append("monthly_salary")
    if email and "@" not in email:
        missing_fields.append("email (invalid)")

    offer_exp_raw = get("offer_expiry_days")

    # Salary breakdown from Pay: columns
    salary_breakdown: list[dict] = []
    breakdown_total = 0.0
    for label, idx in pay_columns or []:
        raw = row[idx] if idx < len(row) else None
        amount = _parse_salary(_cell_str(raw))
        if amount is None or amount <= 0:
            continue
        salary_breakdown.append({"label": label, "amount": amount})
        breakdown_total += amount

    if (
        salary_breakdown
        and monthly_salary is not None
        and breakdown_total - monthly_salary > 0.01
    ):
        missing_fields.append("salary_breakdown (total exceeds monthly salary)")

    benefit_cols = benefit_columns or []
    if benefit_cols:
        benefits = []
        benefit_flags: dict[str, bool] = {}
        for label, idx in benefit_cols:
            checked = _is_truthy_flag(row[idx] if idx < len(row) else None)
            benefit_flags[label] = checked
            if checked:
                benefits.append(label)
        benefits_mode = "checkboxes"
    else:
        benefits_raw = get("benefits")
        benefits = [b.strip() for b in benefits_raw.split(",") if b.strip()] if benefits_raw else []
        benefit_flags = {}
        benefits_mode = "list" if benefits else "default"

    terms = get("terms") or None

    return {
        "row": sheet_row,
        "email": email or None,
        "full_name": full_name or None,
        "job_title": job_title or None,
        "department": department or None,
        "reporting_manager": reporting_manager or None,
        "start_date": start_date or None,
        "monthly_salary": monthly_salary,
        "salary_breakdown": salary_breakdown,
        "breakdown_total": breakdown_total if salary_breakdown else None,
        "office_location": get("office_location") or None,
        "employment_type": get("employment_type") or "Full-time",
        "is_remote": _is_truthy_flag(get("is_remote")),
        "currency": (get("currency") or "PKR").upper(),
        "expires_in_days": 2,
        "offer_expiry_days": int(offer_exp_raw) if offer_exp_raw.isdigit() else 14,
        "message_to_candidate": get("message_to_candidate") or None,
        "terms": terms,
        "benefits": benefits,
        "benefit_flags": benefit_flags,
        "benefits_mode": benefits_mode,
        "missing_fields": missing_fields,
        "valid": not missing_fields,
    }


def build_xlsx_template_bytes() -> bytes:
    """Friendly roster template: offer fields + pay breakdown + benefit Yes/No."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()

    info = wb.active
    info.title = "How to fill"
    info["A1"] = "Bulk invite — same fields as the single invite form"
    info["A1"].font = Font(bold=True, size=14, color="0B1F3A")
    tips = [
        "",
        "1. Open the Candidates sheet and copy the green sample row.",
        "2. Fill required columns (marked *): email, name, role, manager, start date, monthly salary.",
        "3. Salary breakdown — enter amounts in Pay: Basic / Housing / Transport / …",
        "   The sum of Pay columns must not exceed Monthly salary *.",
        "4. Benefits — pick Yes or No from each Benefit: dropdown (like a tick box).",
        "5. Optional: office, employment type, currency, message, offer terms.",
        "6. Upload on Invite → Bulk Excel. History / rehire checks run before you send.",
        "",
        "Colour guide",
        "• Navy headers = candidate & role",
        "• Blue headers = salary breakdown (Pay: …)",
        "• Teal headers = benefits (Yes/No)",
        "• Grey headers = optional offer extras",
        "",
        "Required fields",
        "Email, Full name, Job title, Department, Reporting manager, Start date, Monthly salary",
    ]
    for i, line in enumerate(tips, start=2):
        info[f"A{i}"] = line
        if line in {"Colour guide", "Required fields"}:
            info[f"A{i}"].font = Font(bold=True, color="0B1F3A")
    info.column_dimensions["A"].width = 96

    ws = wb.create_sheet("Candidates", 0)
    pay_headers = [pay_column_header(c) for c in DEFAULT_PAY_COMPONENTS]
    benefit_headers = [benefit_column_header(b) for b in DEFAULT_BENEFITS]
    headers = (
        list(TEMPLATE_CORE_HEADERS)
        + pay_headers
        + list(TEMPLATE_OFFER_META_HEADERS)
        + benefit_headers
    )
    ws.append(headers)

    sample = (
        list(TEMPLATE_SAMPLE_CORE)
        + [TEMPLATE_SAMPLE_PAY.get(c, 0) for c in DEFAULT_PAY_COMPONENTS]
        + list(TEMPLATE_SAMPLE_META)
        + [TEMPLATE_SAMPLE_BENEFIT_FLAGS.get(b, "Yes") for b in DEFAULT_BENEFITS]
    )
    ws.append(sample)

    empty_row = (
        [""] * len(TEMPLATE_CORE_HEADERS)
        + [""] * len(DEFAULT_PAY_COMPONENTS)
        + [""] * len(TEMPLATE_OFFER_META_HEADERS)
        + ["No"] * len(DEFAULT_BENEFITS)
    )
    for _ in range(3):
        ws.append(list(empty_row))

    header_font = Font(bold=True, color="FFFFFF", size=10)
    navy = PatternFill("solid", fgColor="0B1F3A")
    required = PatternFill("solid", fgColor="1E3A5F")
    pay_fill = PatternFill("solid", fgColor="1D4ED8")
    meta_fill = PatternFill("solid", fgColor="475569")
    benefit_fill = PatternFill("solid", fgColor="0F766E")
    sample_fill = PatternFill("solid", fgColor="ECFDF5")
    thin = Border(
        left=Side(style="thin", color="D0D7DE"),
        right=Side(style="thin", color="D0D7DE"),
        top=Side(style="thin", color="D0D7DE"),
        bottom=Side(style="thin", color="D0D7DE"),
    )

    core_n = len(TEMPLATE_CORE_HEADERS)
    pay_n = len(DEFAULT_PAY_COMPONENTS)
    meta_n = len(TEMPLATE_OFFER_META_HEADERS)
    pay_start = core_n + 1
    meta_start = core_n + pay_n + 1
    benefit_start = core_n + pay_n + meta_n + 1

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(1, col_idx)
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = thin
        if col_idx <= core_n:
            cell.fill = required if "*" in header else navy
        elif col_idx < meta_start:
            cell.fill = pay_fill
        elif col_idx < benefit_start:
            cell.fill = meta_fill
        else:
            cell.fill = benefit_fill

    for col_idx in range(1, len(headers) + 1):
        for row_idx in range(2, 6):
            cell = ws.cell(row_idx, col_idx)
            cell.border = thin
            cell.alignment = Alignment(vertical="center")
            if row_idx == 2:
                cell.fill = sample_fill

    for col_idx in range(1, len(headers) + 1):
        letter = get_column_letter(col_idx)
        header = headers[col_idx - 1]
        if col_idx == 1:
            ws.column_dimensions[letter].width = 28
        elif header.startswith("Pay:"):
            ws.column_dimensions[letter].width = 14
        elif header.startswith("Benefit:"):
            ws.column_dimensions[letter].width = 15
        elif "Message" in header or "terms" in header.lower():
            ws.column_dimensions[letter].width = 26
        elif "Start date" in header:
            ws.column_dimensions[letter].width = 20
        else:
            ws.column_dimensions[letter].width = 16

    ws.row_dimensions[1].height = 40
    ws.freeze_panes = "A2"

    # Benefit Yes/No dropdowns
    if benefit_start <= len(headers):
        start_letter = get_column_letter(benefit_start)
        end_letter = get_column_letter(len(headers))
        dv = DataValidation(
            type="list",
            formula1='"Yes,No"',
            allow_blank=True,
            showDropDown=False,
            showErrorMessage=True,
            errorTitle="Benefits",
            error="Choose Yes or No for each benefit.",
            promptTitle="Include benefit?",
            prompt="Yes = include on offer · No = leave off",
        )
        dv.add(f"{start_letter}2:{end_letter}201")
        ws.add_data_validation(dv)

    # Employment type + currency (relative to meta block)
    emp_col = get_column_letter(meta_start + 1)  # Employment type
    cur_col = get_column_letter(meta_start + 2)  # Currency
    emp_dv = DataValidation(
        type="list", formula1='"Full-time,Part-time,Contract,Internship"', allow_blank=True
    )
    emp_dv.add(f"{emp_col}2:{emp_col}201")
    ws.add_data_validation(emp_dv)
    cur_dv = DataValidation(type="list", formula1='"PKR,USD,EUR,GBP"', allow_blank=True)
    cur_dv.add(f"{cur_col}2:{cur_col}201")
    ws.add_data_validation(cur_dv)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
