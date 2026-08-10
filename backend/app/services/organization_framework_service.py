"""Organization Framework Service — multi-tenant career structure management.

Manages departments, role ladders, career roadmaps, and promotion rules scoped
to an organization. Every collection is keyed by ``organization_id`` so data
never leaks across tenants.

Excel import/export follows the UI setup order:
Departments → Career Roles (hierarchy via Next Role) → Career Roadmaps → Promotion Rules.
"""

from __future__ import annotations

import asyncio
import re
from datetime import UTC, datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from app.core.database import database


# ─── Helpers ─────────────────────────────────────────────────────────────────

def _now() -> datetime:
    return datetime.now(UTC)


def _oid_filter(organization_id: str) -> dict:
    return {"organization_id": organization_id}


def _doc_id(organization_id: str, name: str) -> str:
    return f"{organization_id}:{name}"


# ╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝ //
#   Departments
# ╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝ //

async def list_departments(organization_id: str) -> list[dict]:
    docs = await database.org_framework_departments.find(
        _oid_filter(organization_id), {"_id": 0}
    ).sort("name", 1).to_list(1000)
    return docs


async def get_department(organization_id: str, name: str) -> dict | None:
    return await database.org_framework_departments.find_one(
        {"organization_id": organization_id, "name": name},
        {"_id": 0},
    )


async def create_department(organization_id: str, data: dict) -> dict:
    name = (data.get("name") or "").strip()
    if not name:
        raise ValueError("Department name is required.")
    existing = await get_department(organization_id, name)
    if existing:
        raise ValueError(f'Department "{name}" already exists.')
    now = _now()
    doc = {
        "organization_id": organization_id,
        "name": name,
        "description": (data.get("description") or "").strip(),
        "created_at": now,
        "updated_at": now,
    }
    await database.org_framework_departments.insert_one(doc)
    return {k: v for k, v in doc.items() if k != "_id"}


async def update_department(organization_id: str, name: str, data: dict) -> dict:
    dept = await get_department(organization_id, name)
    if not dept:
        raise ValueError(f'Department "{name}" not found.')
    update: dict[str, Any] = {"updated_at": _now()}
    if "name" in data and data["name"].strip() and data["name"].strip() != name:
        new_name = data["name"].strip()
        dup = await get_department(organization_id, new_name)
        if dup:
            raise ValueError(f'Department "{new_name}" already exists.')
        update["name"] = new_name
        # Cascade rename in roles
        await database.org_framework_roles.update_many(
            {"organization_id": organization_id, "department": name},
            {"$set": {"department": new_name, "updated_at": _now()}},
        )
    if "description" in data:
        update["description"] = data["description"].strip()
    await database.org_framework_departments.update_one(
        {"organization_id": organization_id, "name": name},
        {"$set": update},
    )
    return await get_department(organization_id, update.get("name", name))


async def delete_department(organization_id: str, name: str) -> bool:
    # Cascade-delete roles in the department and their dependents so no
    # orphaned framework records remain.
    roles = await database.org_framework_roles.find(
        {"organization_id": organization_id, "department": name}, {"_id": 0, "name": 1}
    ).to_list(10000)
    role_names = [r["name"] for r in roles]
    if role_names:
        await database.org_framework_skills.delete_many(
            {"organization_id": organization_id, "role_name": {"$in": role_names}},
        )
        await database.org_framework_certifications.delete_many(
            {"organization_id": organization_id, "role_name": {"$in": role_names}},
        )
        await database.org_framework_roadmaps.delete_many(
            {"organization_id": organization_id, "role_name": {"$in": role_names}},
        )
        await database.org_framework_promotion_rules.delete_many(
            {"organization_id": organization_id, "role_name": {"$in": role_names}},
        )
    await database.org_framework_roles.delete_many(
        {"organization_id": organization_id, "department": name},
    )
    result = await database.org_framework_departments.delete_one(
        {"organization_id": organization_id, "name": name},
    )
    return result.deleted_count > 0


# ╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝ //
#   Roles
# ╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝ //

async def list_roles(organization_id: str, department: str | None = None) -> list[dict]:
    q: dict = _oid_filter(organization_id)
    if department:
        q["department"] = department
    docs = await database.org_framework_roles.find(q, {"_id": 0}).sort(
        [("department", 1), ("level_number", 1)]
    ).to_list(5000)
    return docs


async def get_role(organization_id: str, role_id: str) -> dict | None:
    return await database.org_framework_roles.find_one(
        {"organization_id": organization_id, "role_id": role_id},
        {"_id": 0},
    )


async def get_role_by_name(organization_id: str, name: str, department: str | None = None) -> dict | None:
    q: dict = {"organization_id": organization_id, "name": name}
    if department:
        q["department"] = department
    return await database.org_framework_roles.find_one(q, {"_id": 0})


async def _assert_no_role_cycle(
    organization_id: str,
    department: str,
    role_name: str,
    next_role: str | None,
    *,
    old_name: str | None = None,
) -> None:
    """Walk Promotes-to links with the proposed edge; reject loops."""
    if not next_role:
        return
    roles = await list_roles(organization_id, department)
    links: dict[str, str | None] = {}
    for r in roles:
        links[r["name"]] = (r.get("next_role") or None)
    # Drop old name when renaming, then apply the proposed edge.
    if old_name and old_name != role_name and old_name in links:
        del links[old_name]
        for k, v in list(links.items()):
            if v == old_name:
                links[k] = role_name
    links[role_name] = next_role
    cursor: str | None = next_role
    guard: set[str] = set()
    while cursor:
        if cursor == role_name:
            raise ValueError(
                f'Promotes to "{next_role}" would create a loop back to "{role_name}".'
            )
        if cursor in guard:
            break
        guard.add(cursor)
        cursor = links.get(cursor)


async def recompute_department_level_numbers(organization_id: str, department: str) -> None:
    """Set L1 = entry roles (nothing promotes into them), then +1 along Promotes to."""
    roles = await list_roles(organization_id, department)
    if not roles:
        return
    by_name = {r["name"]: r for r in roles}
    parents: dict[str, list[str]] = {}
    for r in roles:
        nxt = (r.get("next_role") or "").strip()
        if nxt and nxt in by_name:
            parents.setdefault(nxt, []).append(r["name"])

    levels: dict[str, int] = {}

    def depth(name: str, stack: set[str]) -> int:
        if name in levels:
            return levels[name]
        if name in stack:
            return 1
        stack.add(name)
        preds = parents.get(name) or []
        value = 1 if not preds else 1 + max(depth(p, stack) for p in preds)
        stack.discard(name)
        levels[name] = value
        return value

    for r in roles:
        depth(r["name"], set())

    now = _now()
    for r in roles:
        new_level = levels.get(r["name"], 1)
        if int(r.get("level_number") or 0) != new_level:
            await database.org_framework_roles.update_one(
                {"organization_id": organization_id, "role_id": r["role_id"]},
                {"$set": {"level_number": new_level, "updated_at": now}},
            )


async def create_role(organization_id: str, data: dict) -> dict:
    name = (data.get("name") or "").strip()
    department = (data.get("department") or "").strip()
    if not name or not department:
        raise ValueError("Role name and department are required.")
    dept = await get_department(organization_id, department)
    if not dept:
        raise ValueError(f'Department "{department}" does not exist in the organization framework.')
    existing = await get_role_by_name(organization_id, name, department)
    if existing:
        raise ValueError(f'Role "{name}" already exists in {department}.')
    next_role = (data.get("next_role") or "").strip() or None
    if next_role and next_role == name:
        raise ValueError("A role cannot be its own next role.")
    if next_role:
        target = await get_role_by_name(organization_id, next_role, department)
        if not target:
            raise ValueError(
                f'Next role "{next_role}" must already exist in department "{department}".'
            )
        await _assert_no_role_cycle(organization_id, department, name, next_role)
    # Level is derived from the ladder after insert; seed with 1.
    level_number = 1
    now = _now()
    role_id = f"{organization_id}:{department}:{name}"
    doc = {
        "organization_id": organization_id,
        "role_id": role_id,
        "name": name,
        "department": department,
        "next_role": next_role,
        "level_number": level_number,
        "description": (data.get("description") or "").strip(),
        "created_at": now,
        "updated_at": now,
    }
    await database.org_framework_roles.insert_one(doc)
    await recompute_department_level_numbers(organization_id, department)
    refreshed = await get_role(organization_id, role_id)
    return refreshed or {k: v for k, v in doc.items() if k != "_id"}


async def update_role(organization_id: str, role_id: str, data: dict) -> dict:
    role = await get_role(organization_id, role_id)
    if not role:
        raise ValueError(f"Role not found.")
    update: dict[str, Any] = {"updated_at": _now()}
    for field in ("name", "department", "next_role", "level_number", "description"):
        if field in data:
            val = data[field]
            if isinstance(val, str):
                val = val.strip()
            if field in ("name", "department") and not val:
                raise ValueError(f"{field} cannot be empty.")
            update[field] = val

    new_name = update.get("name", role["name"])
    new_department = update.get("department", role["department"])
    if new_name != role["name"]:
        dup = await get_role_by_name(organization_id, new_name, new_department)
        if dup and dup.get("role_id") != role_id:
            raise ValueError(f'Role "{new_name}" already exists in this department.')

    if new_department != role["department"]:
        dept = await get_department(organization_id, new_department)
        if not dept:
            raise ValueError(f'Department "{new_department}" does not exist in the organization framework.')
        dup = await get_role_by_name(organization_id, new_name, new_department)
        if dup and dup.get("role_id") != role_id:
            raise ValueError(f'Role "{new_name}" already exists in department "{new_department}".')

    # Level numbers are always recomputed from Promotes-to links.
    update.pop("level_number", None)

    next_role = update.get("next_role", role.get("next_role"))
    if isinstance(next_role, str):
        next_role = next_role.strip() or None
    if next_role and next_role == new_name:
        raise ValueError("A role cannot be its own next role.")
    if next_role:
        target = await get_role_by_name(organization_id, next_role, new_department)
        if not target:
            raise ValueError(
                f'Next role "{next_role}" must already exist in department "{new_department}".'
            )
        await _assert_no_role_cycle(
            organization_id,
            new_department,
            new_name,
            next_role,
            old_name=role["name"],
        )
    update["next_role"] = next_role or None

    old_department = role["department"]

    await database.org_framework_roles.update_one(
        {"organization_id": organization_id, "role_id": role_id},
        {"$set": update},
    )

    # Cascade a rename to dependents that reference the role by name.
    if new_name != role["name"]:
        now = _now()
        for coll, field in (
            (database.org_framework_skills, "role_name"),
            (database.org_framework_certifications, "role_name"),
            (database.org_framework_roadmaps, "role_name"),
            (database.org_framework_promotion_rules, "role_name"),
        ):
            await coll.update_many(
                {"organization_id": organization_id, field: role["name"]},
                {"$set": {field: new_name, "updated_at": now}},
            )
        # Fix other roles whose next_role pointed at the renamed role.
        await database.org_framework_roles.update_many(
            {"organization_id": organization_id, "next_role": role["name"]},
            {"$set": {"next_role": new_name, "updated_at": now}},
        )
    # Cascade a department change to the role's dependents.
    if update.get("department") and update["department"] != role["department"]:
        # Roles are keyed by (department, name); a department move re-keys the role.
        await database.org_framework_roles.update_one(
            {"organization_id": organization_id, "role_id": role_id},
            {"$set": {"role_id": f"{organization_id}:{update['department']}:{new_name}"}},
        )
        role_id = f"{organization_id}:{update['department']}:{new_name}"

    await recompute_department_level_numbers(organization_id, new_department)
    if old_department != new_department:
        await recompute_department_level_numbers(organization_id, old_department)
    return await get_role(organization_id, role_id)


async def delete_role(organization_id: str, role_id: str) -> bool:
    role = await get_role(organization_id, role_id)
    if not role:
        return False
    department = role.get("department")
    # Clear Promotes-to links that pointed at this role so the ladder reconnects cleanly.
    await database.org_framework_roles.update_many(
        {"organization_id": organization_id, "next_role": role["name"]},
        {"$set": {"next_role": None, "updated_at": _now()}},
    )
    # Remove dependents that reference the role so no orphans remain.
    await database.org_framework_skills.delete_many(
        {"organization_id": organization_id, "role_name": role["name"]},
    )
    await database.org_framework_certifications.delete_many(
        {"organization_id": organization_id, "role_name": role["name"]},
    )
    await database.org_framework_roadmaps.delete_many(
        {"organization_id": organization_id, "role_name": role["name"]},
    )
    await database.org_framework_promotion_rules.delete_many(
        {"organization_id": organization_id, "role_name": role["name"]},
    )
    result = await database.org_framework_roles.delete_one(
        {"organization_id": organization_id, "role_id": role_id},
    )
    if result.deleted_count > 0 and department:
        await recompute_department_level_numbers(organization_id, department)
    return result.deleted_count > 0


# ╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝ //
#   Skills (per role)
# ╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝ //

PROFICIENCY_LEVELS = ("Beginner", "Intermediate", "Advanced", "Expert")


async def list_skills(organization_id: str, role_id: str | None = None, role_name: str | None = None) -> list[dict]:
    q: dict = _oid_filter(organization_id)
    if role_id:
        q["role_id"] = role_id
    elif role_name:
        q["role_name"] = role_name
    docs = await database.org_framework_skills.find(q, {"_id": 0}).to_list(10000)

    try:
        emp_filter = {"organization_id": organization_id, "status": "active"} if organization_id else {"status": "active"}
        emp_ids = await database.employees.find(
            emp_filter, {"_id": 0, "employee_id": 1}
        ).to_list(length=50000)
        id_set = [e["employee_id"] for e in emp_ids if e.get("employee_id")]
        if id_set:
            pipeline = [
                {"$match": {"employee_id": {"$in": id_set}}},
                {"$group": {
                    "_id": "$skill_name",
                    "proficiency": {"$first": "$proficiency"},
                    "employee_count": {"$sum": 1},
                }},
                {"$sort": {"_id": 1}},
            ]
            results = await database.employee_skills.aggregate(pipeline).to_list(length=10000)
            existing_names = {d["skill_name"] for d in docs}
            for r in results:
                name = r.get("_id") or ""
                if name and name not in existing_names:
                    docs.append({
                        "skill_id": f"EMP-{organization_id}:{name}",
                        "role_name": "Employee Skills",
                        "skill_name": name,
                        "proficiency": r.get("proficiency") or "Intermediate",
                        "weight": 20,
                        "source": "employee_skills",
                        "employee_count": r.get("employee_count", 0),
                    })
                    existing_names.add(name)
    except Exception:
        pass

    return docs


def _validate_proficiency(value) -> str:
    proficiency = (value or "Intermediate").strip().title()
    if proficiency not in PROFICIENCY_LEVELS:
        raise ValueError(
            f'Invalid proficiency "{value}". Must be one of: {", ".join(PROFICIENCY_LEVELS)}.'
        )
    return proficiency


def _validate_weight(value) -> int:
    try:
        weight = int(float(str(value)))
    except (TypeError, ValueError):
        raise ValueError("Weight must be a whole number between 1 and 100.")
    if weight < 1 or weight > 100:
        raise ValueError("Weight must be a whole number between 1 and 100.")
    return weight


async def create_skill(organization_id: str, data: dict) -> dict:
    role_name = (data.get("role_name") or "").strip()
    skill_name = (data.get("skill_name") or "").strip()
    if not role_name or not skill_name:
        raise ValueError("Role and skill name are required.")
    role = await get_role_by_name(organization_id, role_name)
    if not role:
        raise ValueError(f'Role "{role_name}" does not exist in the organization framework.')
    proficiency = _validate_proficiency(data.get("proficiency"))
    weight = _validate_weight(data.get("weight") if data.get("weight") is not None else 20)
    now = _now()
    skill_id = f"{organization_id}:{role_name}:{skill_name}"
    existing = await database.org_framework_skills.find_one(
        {"organization_id": organization_id, "skill_id": skill_id},
        {"_id": 0},
    )
    if existing:
        raise ValueError(f'Skill "{skill_name}" already exists for role "{role_name}".')
    doc = {
        "organization_id": organization_id,
        "skill_id": skill_id,
        "role_name": role_name,
        "skill_name": skill_name,
        "proficiency": proficiency,
        "weight": weight,
        "created_at": now,
        "updated_at": now,
    }
    await database.org_framework_skills.insert_one(doc)
    return {k: v for k, v in doc.items() if k != "_id"}


async def update_skill(organization_id: str, skill_id: str, data: dict) -> dict:
    existing = await database.org_framework_skills.find_one(
        {"organization_id": organization_id, "skill_id": skill_id},
        {"_id": 0},
    )
    if not existing:
        raise ValueError("Skill not found.")
    update: dict[str, Any] = {"updated_at": _now()}
    if "proficiency" in data:
        update["proficiency"] = _validate_proficiency(data["proficiency"])
    if "weight" in data:
        update["weight"] = _validate_weight(data["weight"])
    if "skill_name" in data and (data["skill_name"] or "").strip():
        new_name = data["skill_name"].strip()
        if new_name != existing["skill_name"]:
            new_id = f"{organization_id}:{existing['role_name']}:{new_name}"
            dup = await database.org_framework_skills.find_one(
                {"organization_id": organization_id, "skill_id": new_id},
                {"_id": 0},
            )
            if dup:
                raise ValueError(
                    f'Skill "{new_name}" already exists for role "{existing["role_name"]}".'
                )
            update["skill_id"] = new_id
            update["skill_name"] = new_name
    if not update:
        return existing
    await database.org_framework_skills.update_one(
        {"organization_id": organization_id, "skill_id": skill_id},
        {"$set": update},
    )
    doc = await database.org_framework_skills.find_one(
        {"organization_id": organization_id, "skill_id": update.get("skill_id", skill_id)},
        {"_id": 0},
    )
    return doc


async def delete_skill(organization_id: str, skill_id: str) -> bool:
    result = await database.org_framework_skills.delete_one(
        {"organization_id": organization_id, "skill_id": skill_id},
    )
    return result.deleted_count > 0


# ╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝ //
#   Certifications (per role)
# ╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝ //

async def list_certifications(organization_id: str, role_name: str | None = None) -> list[dict]:
    q: dict = _oid_filter(organization_id)
    if role_name:
        q["role_name"] = role_name
    docs = await database.org_framework_certifications.find(q, {"_id": 0}).to_list(5000)

    try:
        emp_filter = {"organization_id": organization_id, "status": "active"} if organization_id else {"status": "active"}
        emp_ids = await database.employees.find(
            emp_filter, {"_id": 0, "employee_id": 1}
        ).to_list(length=50000)
        id_set = [e["employee_id"] for e in emp_ids if e.get("employee_id")]
        if id_set:
            pipeline = [
                {"$match": {"employee_id": {"$in": id_set}, "verification_status": "verified"}},
                {"$group": {
                    "_id": "$course_title",
                    "employee_count": {"$sum": 1},
                }},
                {"$sort": {"_id": 1}},
            ]
            results = await database.learning_certificates.aggregate(pipeline).to_list(length=10000)
            existing_names = {d["certification_name"] for d in docs}
            for r in results:
                name = r.get("_id") or ""
                if name and name not in existing_names:
                    docs.append({
                        "cert_id": f"EMP-{organization_id}:{name}",
                        "role_name": "Employee Certifications",
                        "certification_name": name,
                        "mandatory": False,
                        "source": "learning_certificates",
                        "employee_count": r.get("employee_count", 0),
                    })
                    existing_names.add(name)
    except Exception:
        pass

    return docs


async def create_certification(organization_id: str, data: dict) -> dict:
    role_name = (data.get("role_name") or "").strip()
    cert_name = (data.get("certification_name") or "").strip()
    if not role_name or not cert_name:
        raise ValueError("Role and certification name are required.")
    role = await get_role_by_name(organization_id, role_name)
    if not role:
        raise ValueError(f'Role "{role_name}" does not exist in the organization framework.')
    expiration_months = data.get("expiration_months")
    if expiration_months not in (None, ""):
        try:
            expiration_months = int(float(str(expiration_months)))
        except (TypeError, ValueError):
            raise ValueError("Expiration must be a positive number of months.")
        if expiration_months < 1:
            raise ValueError("Expiration must be a positive number of months.")
    now = _now()
    cert_id = f"{organization_id}:{role_name}:{cert_name}"
    existing = await database.org_framework_certifications.find_one(
        {"organization_id": organization_id, "cert_id": cert_id},
        {"_id": 0},
    )
    if existing:
        raise ValueError(f'Certification "{cert_name}" already exists for role "{role_name}".')
    doc = {
        "organization_id": organization_id,
        "cert_id": cert_id,
        "role_name": role_name,
        "certification_name": cert_name,
        "mandatory": bool(data.get("mandatory", True)),
        "expiration_months": expiration_months,
        "created_at": now,
        "updated_at": now,
    }
    await database.org_framework_certifications.insert_one(doc)
    return {k: v for k, v in doc.items() if k != "_id"}


async def update_certification(organization_id: str, cert_id: str, data: dict) -> dict:
    existing = await database.org_framework_certifications.find_one(
        {"organization_id": organization_id, "cert_id": cert_id},
        {"_id": 0},
    )
    if not existing:
        raise ValueError("Certification not found.")
    update: dict[str, Any] = {"updated_at": _now()}
    if "mandatory" in data:
        update["mandatory"] = bool(data["mandatory"])
    if "expiration_months" in data:
        value = data["expiration_months"]
        if value in (None, ""):
            update["expiration_months"] = None
        else:
            try:
                months = int(float(str(value)))
            except (TypeError, ValueError):
                raise ValueError("Expiration must be a positive number of months.")
            if months < 1:
                raise ValueError("Expiration must be a positive number of months.")
            update["expiration_months"] = months
    if "certification_name" in data and (data["certification_name"] or "").strip():
        new_name = data["certification_name"].strip()
        if new_name != existing["certification_name"]:
            new_id = f"{organization_id}:{existing['role_name']}:{new_name}"
            dup = await database.org_framework_certifications.find_one(
                {"organization_id": organization_id, "cert_id": new_id},
                {"_id": 0},
            )
            if dup:
                raise ValueError(
                    f'Certification "{new_name}" already exists for role "{existing["role_name"]}".'
                )
            update["cert_id"] = new_id
            update["certification_name"] = new_name
    if not update:
        return existing
    result = await database.org_framework_certifications.update_one(
        {"organization_id": organization_id, "cert_id": cert_id},
        {"$set": update},
    )
    if result.matched_count == 0:
        raise ValueError("Certification not found.")
    doc = await database.org_framework_certifications.find_one(
        {"organization_id": organization_id, "cert_id": update.get("cert_id", cert_id)},
        {"_id": 0},
    )
    return doc


async def delete_certification(organization_id: str, cert_id: str) -> bool:
    result = await database.org_framework_certifications.delete_one(
        {"organization_id": organization_id, "cert_id": cert_id},
    )
    return result.deleted_count > 0


# ╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝ //
#   Course Catalog (organization-wide, reusable)
# ╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝ //

async def list_courses(organization_id: str) -> list[dict]:
    oid_filter = _oid_filter(organization_id)
    framework_docs = await database.org_framework_courses.find(
        oid_filter, {"_id": 0}
    ).sort("name", 1).to_list(10000)

    learning_query = {
        "$or": [{"organization_id": organization_id}, {"organization_id": {"$exists": False}}]
    }
    learning_docs = await database.learning_courses.find(
        learning_query
    ).sort("title", 1).to_list(10000)

    seen = {c["course_id"] for c in framework_docs}
    for ld in learning_docs:
        link_id = ld.get("org_framework_course_id") or ""
        if link_id and link_id in seen:
            continue
        synthetic_id = link_id or f"LC-{str(ld.get('_id', ''))}"
        framework_docs.append({
            "organization_id": organization_id,
            "course_id": synthetic_id,
            "name": ld.get("title") or "",
            "provider": ld.get("provider") or "",
            "category": ld.get("category") or ld.get("designation") or "",
            "duration_hours": round((ld.get("duration_minutes") or 0) / 60, 1) if ld.get("duration_minutes") else None,
            "difficulty": ld.get("difficulty") or "Beginner",
            "url": ld.get("url"),
            "description": ld.get("description") or "",
            "source": "learning_courses",
        })
        seen.add(synthetic_id)
    return framework_docs


async def get_course(organization_id: str, course_id: str) -> dict | None:
    return await database.org_framework_courses.find_one(
        {"organization_id": organization_id, "course_id": course_id},
        {"_id": 0},
    )


async def create_course(organization_id: str, data: dict) -> dict:
    name = (data.get("name") or "").strip()
    if not name:
        raise ValueError("Course name is required.")
    now = _now()
    course_id = data.get("course_id") or f"ORG-{now.strftime('%Y%m%d%H%M%S')}-{name[:20].upper().replace(' ', '-')}"
    doc = {
        "organization_id": organization_id,
        "course_id": course_id,
        "name": name,
        "provider": (data.get("provider") or "").strip(),
        "category": (data.get("category") or "").strip(),
        "duration_hours": data.get("duration_hours"),
        "difficulty": (data.get("difficulty") or "Beginner").strip(),
        "url": (data.get("url") or "").strip() or None,
        "description": (data.get("description") or "").strip(),
        "created_at": now,
        "updated_at": now,
    }
    await database.org_framework_courses.insert_one(doc)
    try:
        from app.services.course_sync_service import sync_to_learning
        await sync_to_learning(organization_id, doc)
    except Exception:
        pass
    return {k: v for k, v in doc.items() if k != "_id"}


async def update_course(organization_id: str, course_id: str, data: dict) -> dict:
    update: dict[str, Any] = {"updated_at": _now()}
    for field in ("name", "provider", "category", "duration_hours", "difficulty", "url", "description"):
        if field in data:
            val = data[field]
            if isinstance(val, str):
                val = val.strip() or None
            update[field] = val
    result = await database.org_framework_courses.update_one(
        {"organization_id": organization_id, "course_id": course_id},
        {"$set": update},
    )
    if result.matched_count == 0:
        raise ValueError("Course not found.")
    try:
        from app.services.course_sync_service import sync_to_learning
        updated = await get_course(organization_id, course_id)
        await sync_to_learning(organization_id, updated)
    except Exception:
        pass
    return await get_course(organization_id, course_id)


async def delete_course(organization_id: str, course_id: str) -> bool:
    # Remove roadmap entries referencing the course so no orphans remain.
    await database.org_framework_roadmaps.delete_many(
        {"organization_id": organization_id, "course_id": course_id},
    )
    try:
        from app.services.course_sync_service import sync_delete_from_learning
        await sync_delete_from_learning(organization_id, course_id)
    except Exception:
        pass
    result = await database.org_framework_courses.delete_one(
        {"organization_id": organization_id, "course_id": course_id},
    )
    return result.deleted_count > 0


# ╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝ //
#   Learning Roadmap (per role)
# ╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝ //

async def list_roadmaps(organization_id: str, role_name: str | None = None) -> list[dict]:
    q: dict = _oid_filter(organization_id)
    if role_name:
        q["role_name"] = role_name
    docs = await database.org_framework_roadmaps.find(q, {"_id": 0}).sort(
        [("role_name", 1), ("order", 1)]
    ).to_list(10000)

    try:
        level_query = {"$or": [{"organization_id": organization_id}, {"organization_id": {"$exists": False}}]}
        if role_name:
            level_query = {"$and": [level_query, {"role_title": {"$regex": f"^{re.escape(role_name.strip())}$", "$options": "i"}}]}
        levels = await database.career_levels.find(
            level_query, {"_id": 0, "role_title": 1, "learning_path": 1}
        ).to_list(length=10000)

        seen = {(d.get("role_name"), d.get("course_id")) for d in docs}
        for level in levels:
            role = level.get("role_title") or ""
            if not role:
                continue
            for course in level.get("learning_path") or []:
                title = course.get("course_title") or ""
                if not title:
                    continue
                row_key = (role, title)
                if row_key in seen:
                    continue
                seen.add(row_key)
                docs.append({
                    "role_name": role,
                    "course_id": course.get("course_uid") or title,
                    "course_name": title,
                    "order": course.get("order") or 1,
                    "source": "career_levels",
                })
    except Exception:
        pass

    return docs


async def create_roadmap(organization_id: str, data: dict) -> dict:
    role_name = (data.get("role_name") or "").strip()
    course_id = (data.get("course_id") or "").strip()
    if not role_name or not course_id:
        raise ValueError("Role and course are required.")
    role = await get_role_by_name(organization_id, role_name)
    if not role:
        raise ValueError(f'Role "{role_name}" does not exist in the organization framework.')
    course = await get_course(organization_id, course_id)
    if course:
        course_name = (data.get("course_name") or "").strip() or course["name"]
    else:
        # Courses may come from the external learning catalog (Microsoft Learn,
        # Coursera, managed providers). In that case the display name is sent by
        # the client so the roadmap entry is readable even though the course is
        # not stored in the organization framework catalog.
        course_name = (data.get("course_name") or "").strip()
        if not course_name:
            raise ValueError(
                f'Course "{course_id}" is not in the organization framework catalog '
                "and no course name was provided."
            )
    existing = await database.org_framework_roadmaps.find_one(
        {"organization_id": organization_id, "role_name": role_name, "course_id": course_id},
        {"_id": 0},
    )
    if existing:
        raise ValueError(f'Course "{course_id}" is already on the roadmap for role "{role_name}".')
    now = _now()
    roadmap_rows = await database.org_framework_roadmaps.find(
        {"organization_id": organization_id, "role_name": role_name}
    ).to_list(10000)
    order = data.get("order") or (len(roadmap_rows) + 1)
    roadmap_id = f"{organization_id}:{role_name}:{course_id}:{now.strftime('%Y%m%d%H%M%S%f')}"
    catalog_type = (data.get("catalog_type") or data.get("type") or "").strip() or None
    skills = data.get("skills")
    if not isinstance(skills, list):
        skills = []
    skills = [str(s).strip() for s in skills if str(s).strip()][:12]
    competency = (data.get("competency") or "").strip() or None
    if competency and competency not in skills:
        skills = [competency, *skills][:12]
    category = (data.get("category") or "").strip() or None
    certifications = data.get("certifications")
    if not isinstance(certifications, list):
        certifications = []
    certifications = [str(c).strip() for c in certifications if str(c).strip()][:8]
    if catalog_type and "cert" in catalog_type.lower() and course_name and course_name not in certifications:
        certifications = [course_name, *certifications][:8]
    doc = {
        "organization_id": organization_id,
        "roadmap_id": roadmap_id,
        "role_name": role_name,
        "course_id": course_id,
        "course_name": course_name,
        "course_url": (
            (data.get("course_url") or data.get("url") or "").strip()
            or ((course or {}).get("url") if course else None)
            or ((course or {}).get("course_url") if course else None)
            or None
        ),
        "catalog_type": catalog_type,
        "category": category,
        "competency": competency,
        "skills": skills,
        "certifications": certifications,
        "mandatory": bool(data.get("mandatory", True)),
        "order": order,
        "prerequisite_course_id": data.get("prerequisite_course_id"),
        "created_at": now,
        "updated_at": now,
    }
    # Prefer live catalog URL when the org framework course has none.
    if not doc.get("course_url") and course_id:
        try:
            from app.services import catalog_service

            item = await catalog_service.get_course_by_uid(course_id)
            url = ((item or {}).get("url") or "").strip()
            if url.startswith("http://") or url.startswith("https://"):
                doc["course_url"] = url
        except Exception:
            pass
    await database.org_framework_roadmaps.insert_one(doc)
    return {k: v for k, v in doc.items() if k != "_id"}


async def reorder_roadmap(organization_id: str, role_name: str, ordered_ids: list[str]) -> list[dict]:
    """Set the recommended order for a role's roadmap entries.

    ``ordered_ids`` must contain exactly the roadmap entries that belong to the
    role (in the new order); any entry missing from the list is appended after
    them in its previous relative order.
    """
    role_name = (role_name or "").strip()
    if not role_name:
        raise ValueError("Role name is required.")
    rows = await database.org_framework_roadmaps.find(
        {"organization_id": organization_id, "role_name": role_name}
    ).to_list(10000)
    by_id = {r["roadmap_id"]: r for r in rows}
    unknown = [rid for rid in ordered_ids if rid not in by_id]
    if unknown:
        raise ValueError(
            f"Roadmap entries {unknown} do not belong to role '{role_name}'."
        )
    now = _now()
    ordered = [rid for rid in ordered_ids if rid in by_id]
    remaining = sorted(
        (r for rid, r in by_id.items() if rid not in ordered),
        key=lambda r: (r.get("order") is None, r.get("order") or 0),
    )
    position = 1
    for rid in ordered + [r["roadmap_id"] for r in remaining]:
        await database.org_framework_roadmaps.update_one(
            {"organization_id": organization_id, "roadmap_id": rid},
            {"$set": {"order": position, "updated_at": now}},
        )
        position += 1
    return await list_roadmaps(organization_id, role_name)


async def update_roadmap(organization_id: str, roadmap_id: str, data: dict) -> dict:
    update: dict[str, Any] = {"updated_at": _now()}
    for field in ("mandatory", "order", "prerequisite_course_id"):
        if field in data:
            update[field] = data[field]
    result = await database.org_framework_roadmaps.update_one(
        {"organization_id": organization_id, "roadmap_id": roadmap_id},
        {"$set": update},
    )
    if result.matched_count == 0:
        raise ValueError("Roadmap entry not found.")
    doc = await database.org_framework_roadmaps.find_one(
        {"organization_id": organization_id, "roadmap_id": roadmap_id},
        {"_id": 0},
    )
    return doc


async def delete_roadmap(organization_id: str, roadmap_id: str) -> bool:
    result = await database.org_framework_roadmaps.delete_one(
        {"organization_id": organization_id, "roadmap_id": roadmap_id},
    )
    return result.deleted_count > 0


# ╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝ //
#   Promotion Rules (per role)
# ╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝ //

async def list_promotion_rules(organization_id: str) -> list[dict]:
    docs = await database.org_framework_promotion_rules.find(
        _oid_filter(organization_id), {"_id": 0}
    ).sort([("department", 1), ("role_name", 1)]).to_list(5000)

    # Backfill department on legacy rules that only stored role_name.
    roles = await list_roles(organization_id)
    roles_by_name: dict[str, list[dict]] = {}
    for role in roles:
        roles_by_name.setdefault(role.get("name") or "", []).append(role)
    for doc in docs:
        if doc.get("department"):
            continue
        matches = roles_by_name.get(doc.get("role_name") or "") or []
        if len(matches) == 1:
            doc["department"] = matches[0].get("department")

    try:
        level_query = {"$or": [{"organization_id": organization_id}, {"organization_id": {"$exists": False}}]}
        levels = await database.career_levels.find(
            level_query, {"_id": 0, "role_title": 1, "min_experience_years": 1,
                          "manager_approval_required": 1, "required_skills": 1,
                          "required_certifications": 1}
        ).to_list(length=10000)

        seen = {(d.get("department") or "", d.get("role_name") or "") for d in docs}
        for level in levels:
            role = (level.get("role_title") or "").strip()
            if not role or ("", role) in seen:
                continue
            seen.add(("", role))
            skills_req = level.get("required_skills") or []
            certs_req = level.get("required_certifications") or []
            docs.append({
                "role_name": role,
                "department": None,
                "min_experience_months": round((level.get("min_experience_years") or 0) * 12),
                "required_readiness_pct": 80,
                "manager_approval_required": bool(level.get("manager_approval_required", False)),
                "min_skills_completed_pct": 100 if skills_req else 0,
                "min_certs_completed": len(certs_req),
                "source": "career_levels",
            })
    except Exception:
        pass

    return docs


async def get_promotion_rule(
    organization_id: str, role_name: str, department: str | None = None
) -> dict | None:
    q: dict = {"organization_id": organization_id, "role_name": role_name}
    if department:
        q["department"] = department
    return await database.org_framework_promotion_rules.find_one(q, {"_id": 0})


async def upsert_promotion_rule(organization_id: str, data: dict) -> dict:
    role_name = (data.get("role_name") or "").strip()
    if not role_name:
        raise ValueError("Role name is required.")
    department = (data.get("department") or "").strip() or None
    role = await get_role_by_name(organization_id, role_name, department)
    if not role:
        raise ValueError(
            f'Role "{role_name}" does not exist'
            + (f' in department "{department}"' if department else " in the organization framework")
            + "."
        )
    department = (role.get("department") or department or "").strip()
    if not department:
        raise ValueError(f'Role "{role_name}" has no department — fix the role ladder first.')

    def _percent(value, field: str, default: int) -> int:
        if value is None or value == "":
            return default
        try:
            number = int(float(str(value)))
        except (TypeError, ValueError):
            raise ValueError(f"{field} must be a whole number between 0 and 100.")
        if number < 0 or number > 100:
            raise ValueError(f"{field} must be a whole number between 0 and 100.")
        return number

    def _count(value, field: str, default: int) -> int:
        if value is None or value == "":
            return default
        try:
            number = int(float(str(value)))
        except (TypeError, ValueError):
            raise ValueError(f"{field} must be a non-negative whole number.")
        if number < 0:
            raise ValueError(f"{field} must be a non-negative whole number.")
        return number

    now = _now()
    doc = {
        "organization_id": organization_id,
        "department": department,
        "role_name": role_name,
        "min_experience_months": _count(data.get("min_experience_months"), "Minimum experience (months)", 0),
        "required_readiness_pct": _percent(data.get("required_readiness_pct"), "Required readiness %", 80),
        "manager_approval_required": bool(data.get("manager_approval_required", True)),
        "min_skills_completed_pct": _percent(data.get("min_skills_completed_pct"), "Minimum skills completed %", 100),
        "min_certs_completed": _count(data.get("min_certs_completed"), "Minimum certifications completed", 0),
        "updated_at": now,
    }
    # Upsert on dept+role; also clear any legacy org+role-only duplicate.
    await database.org_framework_promotion_rules.delete_many(
        {
            "organization_id": organization_id,
            "role_name": role_name,
            "$or": [
                {"department": {"$exists": False}},
                {"department": None},
                {"department": ""},
            ],
        }
    )
    await database.org_framework_promotion_rules.update_one(
        {
            "organization_id": organization_id,
            "department": department,
            "role_name": role_name,
        },
        {"$set": doc, "$setOnInsert": {"created_at": now}},
        upsert=True,
    )
    return doc


async def delete_promotion_rule(
    organization_id: str, role_name: str, department: str | None = None
) -> bool:
    q: dict = {"organization_id": organization_id, "role_name": role_name}
    if department:
        q["department"] = department
    result = await database.org_framework_promotion_rules.delete_many(q)
    return result.deleted_count > 0


# ╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝ //
#   Dashboard Summary
# ╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝ //

async def get_framework_summary(organization_id: str) -> dict:
    """Aggregate counts across all framework collections."""
    departments = await database.org_framework_departments.count_documents(_oid_filter(organization_id))
    roles = await database.org_framework_roles.count_documents(_oid_filter(organization_id))
    skills_framework = await database.org_framework_skills.count_documents(_oid_filter(organization_id))
    certifications_framework = await database.org_framework_certifications.count_documents(_oid_filter(organization_id))
    courses_framework = await database.org_framework_courses.count_documents(_oid_filter(organization_id))
    learning_query = {
        "$or": [{"organization_id": organization_id}, {"organization_id": {"$exists": False}}]
    }
    courses_learning = await database.learning_courses.count_documents(learning_query)
    courses = courses_framework + courses_learning
    roadmaps = len(await list_roadmaps(organization_id))
    promotion_rules = len(await list_promotion_rules(organization_id))

    # Employees in this org
    try:
        emp_filter = {"organization_id": organization_id, "status": "active"} if organization_id else {"status": "active"}
        employees = await database.employees.count_documents(emp_filter)
    except Exception:
        employees = 0

    # Unique skills & certifications from the Learning module (employee skills + certs)
    try:
        emp_ids = await database.employees.find(
            emp_filter, {"_id": 0, "employee_id": 1}
        ).to_list(length=50000)
        id_set = [e["employee_id"] for e in emp_ids if e.get("employee_id")]
        if id_set:
            skills_pipeline = [
                {"$match": {"employee_id": {"$in": id_set}}},
                {"$group": {"_id": "$skill_name"}},
                {"$count": "total"},
            ]
            result = await database.employee_skills.aggregate(skills_pipeline).to_list(1)
            skills_learning = result[0]["total"] if result else 0

            certs_pipeline = [
                {"$match": {"employee_id": {"$in": id_set}, "verification_status": "verified"}},
                {"$group": {"_id": "$course_title"}},
                {"$count": "total"},
            ]
            result = await database.learning_certificates.aggregate(certs_pipeline).to_list(1)
            certs_learning = result[0]["total"] if result else 0
        else:
            skills_learning = 0
            certs_learning = 0
    except Exception:
        skills_learning = 0
        certs_learning = 0

    skills = skills_framework + skills_learning
    certifications = certifications_framework + certs_learning

    # Recently updated across all collections
    recent: list[dict] = []
    for coll_name in ("org_framework_roles", "org_framework_courses", "org_framework_departments"):
        try:
            cursor = database[coll_name].find(
                _oid_filter(organization_id)
            ).sort("updated_at", -1).limit(5)
            async for doc in cursor:
                recent.append({
                    "type": coll_name.replace("org_framework_", "").rstrip("s"),
                    "name": doc.get("name") or doc.get("role_name") or doc.get("certification_name") or "?",
                    "updated_at": doc.get("updated_at"),
                })
        except Exception:
            pass
    recent.sort(key=lambda r: r.get("updated_at") or _now(), reverse=True)

    return {
        "departments": departments,
        "roles": roles,
        "skills": skills,
        "certifications": certifications,
        "courses": courses,
        "roadmaps": roadmaps,
        "promotion_rules": promotion_rules,
        "employees": employees,
        "recent_updates": recent[:10],
    }


async def get_org_structure_options(organization_id: str) -> dict:
    """Compact, org-scoped option sets for every module's dropdowns.

    Single source of truth for departments, roles, skills, certifications and
    courses so recruiter, candidate and employee UIs (and their AI assistants)
    never need independent or hardcoded lists.
    """
    [departments, roles, skills, certifications, courses] = await asyncio.gather(
        list_departments(organization_id),
        list_roles(organization_id),
        list_skills(organization_id),
        list_certifications(organization_id),
        list_courses(organization_id),
    )
    return {
        "departments": [d["name"] for d in departments if d.get("name")],
        "roles": [
            {"name": r["name"], "department": r["department"], "next_role": r.get("next_role")}
            for r in roles
            if r.get("name")
        ],
        "skills": sorted({s.get("skill_name") for s in skills if s.get("skill_name")}),
        "certifications": sorted(
            {c.get("certification_name") for c in certifications if c.get("certification_name")}
        ),
        "courses": [
            {"course_id": c["course_id"], "name": c["name"], "provider": c.get("provider"), "category": c.get("category")}
            for c in courses
            if c.get("name")
        ],
    }


# ╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝ //
#   Seed from existing records
# ╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝ //

async def seed_framework_from_existing(organization_id: str) -> dict:
    """Bootstrap the org framework from records the org already owns.

    Reads department/job_title values from the org's employees, candidates and
    recruiters (READ-ONLY — those people records are never modified, deleted or
    re-tagged) and writes only into the org_framework_* collections so existing
    people keep their current values everywhere.

    Idempotent: existing framework entries are never duplicated or overwritten,
    so re-running after manual edits only adds what is still missing.
    """
    now = _now()

    async def _distinct_pairs(collection: str) -> set[tuple[str, str]]:
        docs = await database[collection].find(
            {"organization_id": organization_id},
            {"_id": 0, "department": 1, "job_title": 1},
        ).to_list(100000)
        pairs: set[tuple[str, str]] = set()
        for d in docs:
            dept = (d.get("department") or "").strip()
            title = (d.get("job_title") or "").strip()
            if dept and title:
                pairs.add((dept, title))
        return pairs

    pairs: set[tuple[str, str]] = set()
    for coll in ("employees", "candidates", "recruiters"):
        try:
            pairs.update(await _distinct_pairs(coll))
        except Exception:
            continue

    existing_departments = {d["name"] for d in await list_departments(organization_id)}
    existing_roles = {
        (r["department"], r["name"]) for r in await list_roles(organization_id)
    }

    created_departments: list[str] = []
    for dept, _title in pairs:
        if dept in existing_departments:
            continue
        await database.org_framework_departments.insert_one({
            "organization_id": organization_id,
            "name": dept,
            "description": "",
            "created_at": now,
            "updated_at": now,
        })
        existing_departments.add(dept)
        created_departments.append(dept)

    role_docs: list[dict] = []
    created_roles: list[str] = []
    for dept, title in pairs:
        if (dept, title) in existing_roles:
            continue
        role_docs.append({
            "organization_id": organization_id,
            "role_id": f"{organization_id}:{dept}:{title}",
            "name": title,
            "department": dept,
            "next_role": None,
            "level_number": 1,
            "description": "",
            "created_at": now,
            "updated_at": now,
        })
        existing_roles.add((dept, title))
        created_roles.append(title)
    if role_docs:
        await database.org_framework_roles.insert_many(role_docs)

    if created_departments or created_roles:
        await create_version_snapshot(organization_id, "Seeded from existing records")

    return {
        "departments_created": created_departments,
        "roles_created": created_roles,
        "departments_total": len(existing_departments),
        "roles_total": len(existing_roles),
    }


# ╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝ //
#   Dashboard Summary
# ╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝ //
#   Version History
# ╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝╝ //

async def create_version_snapshot(organization_id: str, label: str) -> dict:
    now = _now()
    version_id = f"v{now.strftime('%Y%m%d%H%M%S')}"
    snapshot: dict[str, Any] = {
        "organization_id": organization_id,
        "version_id": version_id,
        "label": label,
        "created_at": now,
        "counts": {
            "departments": await database.org_framework_departments.count_documents(_oid_filter(organization_id)),
            "roles": await database.org_framework_roles.count_documents(_oid_filter(organization_id)),
            "skills": await database.org_framework_skills.count_documents(_oid_filter(organization_id)),
            "certifications": await database.org_framework_certifications.count_documents(_oid_filter(organization_id)),
            "courses": await database.org_framework_courses.count_documents(_oid_filter(organization_id)),
        },
    }
    await database.org_framework_versions.insert_one(snapshot)
    return {k: v for k, v in snapshot.items() if k != "_id"}


async def list_versions(organization_id: str) -> list[dict]:
    docs = await database.org_framework_versions.find(
        _oid_filter(organization_id), {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    return docs

# ------------------------------------------------------------------------------- //
#   Excel Export / Import
# ------------------------------------------------------------------------------- //

def _style_sheet_header(ws, headers) -> None:
    fill = PatternFill("solid", fgColor="1F3D5C")
    font = Font(color="FFFFFF", bold=True, size=11)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = fill
        cell.font = font


def _append_rows(ws, rows) -> None:
    for row in rows:
        ws.append(row)


async def _catalog_index_rows(organization_id: str, roadmaps: list[dict]) -> list[list]:
    """Rows for the Catalog Index helper sheet (managed + already-mapped courses)."""
    rows: list[list] = []
    seen: set[str] = set()
    try:
        from app.services.managed_learning_service import managed_learning_service

        result = await managed_learning_service.list_courses(
            organization_id=organization_id,
            page=1,
            page_size=2000,
            archived=False,
        )
        for course in result.get("courses") or []:
            uid = (course.get("uid") or "").strip()
            if not uid or uid in seen:
                continue
            seen.add(uid)
            rows.append([
                uid,
                course.get("title") or "",
                course.get("provider") or course.get("source") or "",
                course.get("type") or "course",
            ])
    except Exception:
        pass

    for r in roadmaps:
        uid = (r.get("course_id") or "").strip()
        if not uid or uid in seen:
            continue
        seen.add(uid)
        rows.append([
            uid,
            r.get("course_name") or uid,
            "",
            r.get("catalog_type") or "",
        ])

    rows.sort(key=lambda x: ((x[2] or "").lower(), (x[1] or "").lower()))
    return rows


async def build_export_workbook(organization_id: str) -> BytesIO:
    """Build an XLSX workbook matching the Organization Setup flow.

    Sheets (in setup order):
      1. Departments
      2. Career Roles  — hierarchy via Next Role (no manual Career Level)
      3. Career Roadmaps — catalog items + skills/certs per role
      4. Promotion Rules — optional readiness thresholds
      5. Catalog Index — reference list of Course IDs (copy into Career Roadmaps)
    """
    [departments, roles, roadmaps, rules] = await asyncio.gather(
        list_departments(organization_id),
        list_roles(organization_id),
        list_roadmaps(organization_id),
        list_promotion_rules(organization_id),
    )

    roles_sorted = sorted(
        roles,
        key=lambda r: (
            (r.get("department") or "").lower(),
            int(r.get("level_number") or 0),
            (r.get("name") or "").lower(),
        ),
    )
    # Map role name → department (unique titles only; ambiguous → first dept).
    role_dept: dict[str, str] = {}
    for r in roles_sorted:
        key = (r.get("name") or "").strip().lower()
        if key and key not in role_dept:
            role_dept[key] = r.get("department") or ""
    known_roles = set(role_dept)
    # Skip orphan roadmap rows whose role was deleted — keeps re-import valid.
    roadmaps_sorted = sorted(
        (
            r for r in roadmaps
            if (r.get("role_name") or "").strip().lower() in known_roles
        ),
        key=lambda r: (
            (role_dept.get((r.get("role_name") or "").strip().lower(), "") or "").lower(),
            (r.get("role_name") or "").lower(),
            int(r.get("order") or 0),
            (r.get("course_name") or "").lower(),
        ),
    )

    wb = Workbook()
    wb.remove(wb.active)

    # 1) Departments
    ws = wb.create_sheet("Departments")
    headers = ["Department Name", "Description"]
    _style_sheet_header(ws, headers)
    _append_rows(ws, [[d["name"], d.get("description", "")] for d in departments])
    for col in ("A", "B"):
        ws.column_dimensions[col].width = 32

    # 2) Career Roles — fill all roles per department; Next Role builds the ladder
    ws = wb.create_sheet("Career Roles")
    headers = ["Department", "Role Name", "Next Role", "Description"]
    _style_sheet_header(ws, headers)
    _append_rows(
        ws,
        [
            [
                r.get("department") or "",
                r.get("name") or "",
                r.get("next_role") or "",
                r.get("description") or "",
            ]
            for r in roles_sorted
        ],
    )
    for col, width in (("A", 24), ("B", 30), ("C", 30), ("D", 40)):
        ws.column_dimensions[col].width = width

    # 3) Career Roadmaps — one row per course/item for each role
    # Course ID may be blank on import when Course Name (+ optional Provider) resolves
    # via Catalog Index or the live learning catalog.
    ws = wb.create_sheet("Career Roadmaps")
    headers = [
        "Department",
        "Role",
        "Course ID",
        "Course Name",
        "Provider",
        "Catalog Type",
        "Mandatory",
        "Order",
        "Skills",
        "Certifications",
    ]
    _style_sheet_header(ws, headers)
    _append_rows(
        ws,
        [
            [
                role_dept.get((r.get("role_name") or "").strip().lower(), ""),
                r.get("role_name") or "",
                r.get("course_id") or "",
                r.get("course_name") or "",
                r.get("provider") or "",
                r.get("catalog_type") or "",
                "Yes" if r.get("mandatory", True) else "No",
                r.get("order") or "",
                _join_csv_list(r.get("skills")),
                _join_csv_list(r.get("certifications")),
            ]
            for r in roadmaps_sorted
        ],
    )
    for col, width in (
        ("A", 22), ("B", 28), ("C", 28), ("D", 36), ("E", 20),
        ("F", 14), ("G", 12), ("H", 10), ("I", 36), ("J", 36),
    ):
        ws.column_dimensions[col].width = width

    # 4) Promotion Rules
    ws = wb.create_sheet("Promotion Rules")
    headers = [
        "Department",
        "Role",
        "Minimum Experience (Months)",
        "Required Readiness %",
        "Manager Approval Required",
        "Minimum Skills Completed %",
        "Minimum Certifications Completed",
    ]
    _style_sheet_header(ws, headers)
    _append_rows(
        ws,
        [
            [
                r.get("department") or "",
                r["role_name"],
                r.get("min_experience_months", 0),
                r.get("required_readiness_pct", 80),
                "Yes" if r.get("manager_approval_required") else "No",
                r.get("min_skills_completed_pct", 100),
                r.get("min_certs_completed", 0),
            ]
            for r in rules
            if r.get("source") != "career_levels"
        ],
    )
    for col, width in (("A", 24), ("B", 30), ("C", 24), ("D", 20), ("E", 24), ("F", 26), ("G", 28)):
        ws.column_dimensions[col].width = width

    # 5) Catalog Index — lookup sheet (not required on import; used to resolve names → IDs)
    ws = wb.create_sheet("Catalog Index")
    headers = ["Course ID", "Course Name", "Provider", "Catalog Type"]
    _style_sheet_header(ws, headers)
    _append_rows(ws, await _catalog_index_rows(organization_id, roadmaps_sorted))
    for col, width in (("A", 36), ("B", 40), ("C", 22), ("D", 14)):
        ws.column_dimensions[col].width = width

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _parse_bool(value) -> bool:
    if value is None:
        return True
    return str(value).strip().lower() in ("yes", "true", "1", "y")


def _parse_int(value, default):
    if value is None or value == "":
        return default
    try:
        return int(float(str(value).strip()))
    except (ValueError, TypeError):
        return default


def _clean_str(value) -> str:
    return str(value).strip() if value is not None else ""


def _join_csv_list(values) -> str:
    if not values:
        return ""
    if isinstance(values, str):
        return values.strip()
    parts = []
    for v in values:
        text = str(v).strip()
        if text and text not in parts:
            parts.append(text)
    return ", ".join(parts)


def _split_csv_list(value) -> list[str]:
    if value is None or value == "":
        return []
    if isinstance(value, (list, tuple)):
        return [str(v).strip() for v in value if str(v).strip()]
    parts = []
    for chunk in re.split(r"[,;|]", str(value)):
        text = chunk.strip()
        if text and text not in parts:
            parts.append(text)
    return parts[:12]


def _normalize_header(value) -> str:
    return (
        _clean_str(value)
        .strip()
        .lower()
        .replace(" ", "")
        .replace("(", "")
        .replace(")", "")
        .replace("%", "")
    )


# Current template (UI flow). Legacy sheets are still accepted on import.
SHEET_SCHEMA: dict[str, list[str]] = {
    "Departments": ["Department Name", "Description"],
    "Career Roles": ["Department", "Role Name", "Next Role", "Description"],
    "Career Roadmaps": [
        "Department",
        "Role",
        "Course ID",
        "Course Name",
        "Provider",
        "Catalog Type",
        "Mandatory",
        "Order",
        "Skills",
        "Certifications",
    ],
    "Promotion Rules": [
        "Department",
        "Role",
        "Minimum Experience (Months)",
        "Required Readiness %",
        "Manager Approval Required",
        "Minimum Skills Completed %",
        "Minimum Certifications Completed",
    ],
    "Catalog Index": ["Course ID", "Course Name", "Provider", "Catalog Type"],
}

# Older exports used these names / columns; still parsed when present.
_LEGACY_CAREER_ROLES_HEADERS = ["Department", "Current Role", "Next Role", "Career Level", "Description"]
_LEGACY_ROADMAP_HEADERS = ["Role", "Course ID", "Mandatory", "Recommended Order"]
_LEGACY_CAREER_ROADMAPS_HEADERS = [
    "Department",
    "Role",
    "Course ID",
    "Course Name",
    "Catalog Type",
    "Mandatory",
    "Order",
    "Skills",
    "Certifications",
]


def _headers_match(actual, expected) -> bool:
    actual_norm = [_normalize_header(h) for h in actual if h is not None]
    expected_norm = [_normalize_header(h) for h in expected]
    if len(actual_norm) < len(expected_norm):
        return False
    return actual_norm[: len(expected_norm)] == expected_norm


def parse_import_workbook(file_bytes: bytes) -> dict:
    """Parse an uploaded workbook into structured framework data.

    Prefers the current 4-sheet template. Also accepts the legacy 7-sheet
    workbook (Skills / Certifications / Course Catalog / Learning Roadmap).
    """
    wb = load_workbook(BytesIO(file_bytes), data_only=True)

    def sheet_rows(name: str):
        if name not in wb.sheetnames:
            return []
        rows = list(wb[name].iter_rows(values_only=True))
        out = []
        for idx, row in enumerate(rows[1:], start=2):
            if any(c is not None and str(c).strip() for c in row):
                out.append((idx, row))
        return out

    def sheet_header(name: str) -> tuple:
        if name not in wb.sheetnames:
            return ()
        header_row = list(wb[name].iter_rows(min_row=1, max_row=1, values_only=True))
        return header_row[0] if header_row else ()

    sheet_issues: list[dict[str, Any]] = []

    # Departments (required)
    if "Departments" not in wb.sheetnames:
        sheet_issues.append({
            "sheet": "Departments",
            "row": 1,
            "column": "-",
            "reason": "Required sheet 'Departments' is missing from the workbook.",
        })
    elif not _headers_match(sheet_header("Departments"), SHEET_SCHEMA["Departments"]):
        sheet_issues.append({
            "sheet": "Departments",
            "row": 1,
            "column": "Department Name",
            "reason": "Departments headers must be: Department Name, Description.",
        })

    # Career Roles — new or legacy headers
    roles_legacy = False
    if "Career Roles" not in wb.sheetnames:
        sheet_issues.append({
            "sheet": "Career Roles",
            "row": 1,
            "column": "-",
            "reason": "Required sheet 'Career Roles' is missing from the workbook.",
        })
    else:
        role_headers = sheet_header("Career Roles")
        if _headers_match(role_headers, SHEET_SCHEMA["Career Roles"]):
            roles_legacy = False
        elif _headers_match(role_headers, _LEGACY_CAREER_ROLES_HEADERS):
            roles_legacy = True
        else:
            sheet_issues.append({
                "sheet": "Career Roles",
                "row": 1,
                "column": "Role Name",
                "reason": (
                    "Career Roles headers must be: Department, Role Name, Next Role, Description "
                    "(or the legacy Current Role / Career Level layout)."
                ),
            })

    # Roadmaps — Career Roadmaps (new) or Learning Roadmap (+ optional Course Catalog)
    has_new_roadmaps = "Career Roadmaps" in wb.sheetnames
    has_legacy_roadmaps = "Learning Roadmap" in wb.sheetnames
    roadmaps_have_provider = False
    if not has_new_roadmaps and not has_legacy_roadmaps:
        sheet_issues.append({
            "sheet": "Career Roadmaps",
            "row": 1,
            "column": "-",
            "reason": "Required sheet 'Career Roadmaps' is missing (or legacy 'Learning Roadmap').",
        })
    elif has_new_roadmaps:
        rm_headers = sheet_header("Career Roadmaps")
        if _headers_match(rm_headers, SHEET_SCHEMA["Career Roadmaps"]):
            roadmaps_have_provider = True
        elif _headers_match(rm_headers, _LEGACY_CAREER_ROADMAPS_HEADERS):
            roadmaps_have_provider = False
        else:
            sheet_issues.append({
                "sheet": "Career Roadmaps",
                "row": 1,
                "column": "Department",
                "reason": (
                    "Career Roadmaps headers must be: Department, Role, Course ID, Course Name, "
                    "Provider, Catalog Type, Mandatory, Order, Skills, Certifications."
                ),
            })
    elif has_legacy_roadmaps and not _headers_match(
        sheet_header("Learning Roadmap"), _LEGACY_ROADMAP_HEADERS
    ):
        sheet_issues.append({
            "sheet": "Learning Roadmap",
            "row": 1,
            "column": "Role",
            "reason": "Learning Roadmap headers must be: Role, Course ID, Mandatory, Recommended Order.",
        })

    # Catalog Index is optional — if present, headers should match for reliable name→ID lookup.
    if "Catalog Index" in wb.sheetnames and not _headers_match(
        sheet_header("Catalog Index"), SHEET_SCHEMA["Catalog Index"]
    ):
        sheet_issues.append({
            "sheet": "Catalog Index",
            "row": 1,
            "column": "Course ID",
            "reason": "Catalog Index headers must be: Course ID, Course Name, Provider, Catalog Type.",
        })

    # Promotion Rules (required sheet; may be empty)
    if "Promotion Rules" not in wb.sheetnames:
        sheet_issues.append({
            "sheet": "Promotion Rules",
            "row": 1,
            "column": "-",
            "reason": "Required sheet 'Promotion Rules' is missing from the workbook.",
        })
    else:
        promo_headers = sheet_header("Promotion Rules")
        # Accept with or without leading Department (legacy Role-first)
        promo_ok = _headers_match(promo_headers, SHEET_SCHEMA["Promotion Rules"]) or _headers_match(
            promo_headers,
            [
                "Role",
                "Minimum Experience (Months)",
                "Required Readiness %",
                "Manager Approval Required",
                "Minimum Skills Completed %",
                "Minimum Certifications Completed",
            ],
        )
        if not promo_ok:
            sheet_issues.append({
                "sheet": "Promotion Rules",
                "row": 1,
                "column": "Department",
                "reason": (
                    "Promotion Rules headers must start with Department, Role, "
                    "Minimum Experience (Months), …"
                ),
            })

    data: dict[str, Any] = {"_sheet_issues": sheet_issues, "_format": "current"}

    # Departments
    depts = []
    for row_num, row in sheet_rows("Departments"):
        name = _clean_str(row[0])
        if not name:
            continue
        depts.append({
            "_row": row_num,
            "name": name,
            "description": _clean_str(row[1]) if len(row) > 1 else "",
        })
    data["departments"] = depts

    # Career Roles
    roles = []
    for row_num, row in sheet_rows("Career Roles"):
        if len(row) < 2:
            continue
        department = _clean_str(row[0])
        name = _clean_str(row[1])
        if not name or not department:
            continue
        if roles_legacy:
            next_role = _clean_str(row[2]) if len(row) > 2 else ""
            description = _clean_str(row[4]) if len(row) > 4 else ""
        else:
            next_role = _clean_str(row[2]) if len(row) > 2 else ""
            description = _clean_str(row[3]) if len(row) > 3 else ""
        roles.append({
            "_row": row_num,
            "department": department,
            "name": name,
            "next_role": next_role,
            "level_number": 1,  # recomputed after import from Next Role links
            "description": description,
        })
    data["roles"] = roles

    # Legacy-only collections (optional)
    skills = []
    if "Skills" in wb.sheetnames:
        for row_num, row in sheet_rows("Skills"):
            if len(row) < 2:
                continue
            role_name = _clean_str(row[0])
            skill_name = _clean_str(row[1])
            if not role_name or not skill_name:
                continue
            skills.append({
                "_row": row_num,
                "role_name": role_name,
                "skill_name": skill_name,
                "proficiency": _clean_str(row[2]) if len(row) > 2 else "Intermediate",
                "weight": _parse_int(row[3], 20) if len(row) > 3 else 20,
            })
    data["skills"] = skills

    certifications = []
    if "Certifications" in wb.sheetnames:
        for row_num, row in sheet_rows("Certifications"):
            if len(row) < 2:
                continue
            role_name = _clean_str(row[0])
            cert_name = _clean_str(row[1])
            if not role_name or not cert_name:
                continue
            certifications.append({
                "_row": row_num,
                "role_name": role_name,
                "certification_name": cert_name,
                "mandatory": _parse_bool(row[2]) if len(row) > 2 else True,
                "expiration_months": _parse_int(row[3], None) if len(row) > 3 else None,
            })
    data["certifications"] = certifications

    courses = []
    if "Course Catalog" in wb.sheetnames:
        for row_num, row in sheet_rows("Course Catalog"):
            if len(row) < 2 or not row[1]:
                continue
            course_id = _clean_str(row[0]) or f"ORG-{len(courses) + 1}"
            courses.append({
                "_row": row_num,
                "course_id": course_id,
                "name": _clean_str(row[1]),
                "provider": _clean_str(row[2]) if len(row) > 2 else "",
                "category": _clean_str(row[3]) if len(row) > 3 else "",
                "duration_hours": _parse_int(row[4], None) if len(row) > 4 else None,
                "difficulty": _clean_str(row[5]) if len(row) > 5 else "Beginner",
                "url": _clean_str(row[6]) if len(row) > 6 else "",
                "description": _clean_str(row[7]) if len(row) > 7 else "",
            })
    data["courses"] = courses

    # Optional Catalog Index — local Course ID lookup for one-shot Excel setups
    catalog_index = []
    if "Catalog Index" in wb.sheetnames:
        for row_num, row in sheet_rows("Catalog Index"):
            if len(row) < 2:
                continue
            course_id = _clean_str(row[0])
            course_name = _clean_str(row[1]) if len(row) > 1 else ""
            if not course_id and not course_name:
                continue
            catalog_index.append({
                "_row": row_num,
                "course_id": course_id,
                "course_name": course_name or course_id,
                "provider": _clean_str(row[2]) if len(row) > 2 else "",
                "catalog_type": _clean_str(row[3]) if len(row) > 3 else "",
            })
    data["catalog_index"] = catalog_index

    # Career Roadmaps (new) or Learning Roadmap (legacy)
    roadmaps = []
    if has_new_roadmaps:
        data["_format"] = "current"
        for row_num, row in sheet_rows("Career Roadmaps"):
            if len(row) < 2:
                continue
            department = _clean_str(row[0]) if len(row) > 0 else ""
            role_name = _clean_str(row[1]) if len(row) > 1 else ""
            course_id = _clean_str(row[2]) if len(row) > 2 else ""
            course_name = _clean_str(row[3]) if len(row) > 3 else ""
            if not role_name or (not course_id and not course_name):
                continue
            if roadmaps_have_provider:
                provider = _clean_str(row[4]) if len(row) > 4 else ""
                catalog_type = _clean_str(row[5]) if len(row) > 5 else ""
                mandatory = _parse_bool(row[6]) if len(row) > 6 else True
                order = _parse_int(row[7], None) if len(row) > 7 else None
                skills = _split_csv_list(row[8] if len(row) > 8 else "")
                certifications = _split_csv_list(row[9] if len(row) > 9 else "")
            else:
                provider = ""
                catalog_type = _clean_str(row[4]) if len(row) > 4 else ""
                mandatory = _parse_bool(row[5]) if len(row) > 5 else True
                order = _parse_int(row[6], None) if len(row) > 6 else None
                skills = _split_csv_list(row[7] if len(row) > 7 else "")
                certifications = _split_csv_list(row[8] if len(row) > 8 else "")
            roadmaps.append({
                "_row": row_num,
                "department": department or None,
                "role_name": role_name,
                "course_id": course_id,
                "course_name": course_name or course_id,
                "provider": provider or None,
                "catalog_type": catalog_type or None,
                "mandatory": mandatory,
                "order": order,
                "skills": skills,
                "certifications": certifications,
            })
    else:
        data["_format"] = "legacy"
        for row_num, row in sheet_rows("Learning Roadmap"):
            if len(row) < 2:
                continue
            role_name = _clean_str(row[0])
            course_id = _clean_str(row[1])
            if not role_name or not course_id:
                continue
            course_name = next(
                (c["name"] for c in courses if c["course_id"] == course_id),
                course_id,
            )
            roadmaps.append({
                "_row": row_num,
                "department": None,
                "role_name": role_name,
                "course_id": course_id,
                "course_name": course_name,
                "provider": None,
                "catalog_type": None,
                "mandatory": _parse_bool(row[2]) if len(row) > 2 else True,
                "order": _parse_int(row[3], None) if len(row) > 3 else None,
                "skills": [],
                "certifications": [],
            })
    data["roadmaps"] = roadmaps

    # Promotion Rules
    rules = []
    for row_num, row in sheet_rows("Promotion Rules"):
        if not row or not row[0]:
            continue
        second = row[1] if len(row) > 1 else None
        legacy = False
        try:
            if second is not None and str(second).strip() != "":
                float(str(second))
                legacy = True
        except (TypeError, ValueError):
            legacy = False
        if legacy:
            role_name = _clean_str(row[0])
            department = None
            base = 1
        else:
            department = _clean_str(row[0]) or None
            role_name = _clean_str(row[1]) if len(row) > 1 else ""
            base = 2
        if not role_name:
            continue
        rules.append({
            "_row": row_num,
            "department": department,
            "role_name": role_name,
            "min_experience_months": _parse_int(row[base], 0) if len(row) > base else 0,
            "required_readiness_pct": _parse_int(row[base + 1], 80) if len(row) > base + 1 else 80,
            "manager_approval_required": _parse_bool(row[base + 2]) if len(row) > base + 2 else True,
            "min_skills_completed_pct": _parse_int(row[base + 3], 100) if len(row) > base + 3 else 100,
            "min_certs_completed": _parse_int(row[base + 4], 0) if len(row) > base + 4 else 0,
        })
    data["promotion_rules"] = rules

    return data


async def resolve_roadmap_catalog_refs(organization_id: str, data: dict) -> list[dict]:
    """Fill missing Course IDs from Catalog Index / live catalog using Course Name (+ Provider).

    Mutates ``data["roadmaps"]`` in place. Returns structured error details
    (same shape as validation details) when a row cannot be resolved uniquely.
    """
    details: list[dict[str, Any]] = []
    index = data.get("catalog_index") or []
    index_by_id = {
        (row.get("course_id") or "").strip().lower(): row
        for row in index
        if (row.get("course_id") or "").strip()
    }

    def _index_matches(name: str, provider: str) -> list[dict]:
        needle = name.strip().lower()
        prov = provider.strip().lower()
        hits = []
        for row in index:
            title = (row.get("course_name") or "").strip().lower()
            if title != needle:
                continue
            row_prov = (row.get("provider") or "").strip().lower()
            if prov and row_prov and prov not in row_prov and row_prov not in prov:
                continue
            hits.append(row)
        return hits

    async def _live_matches(name: str, provider: str) -> list[dict]:
        from app.services.dynamic_catalog_service import dynamic_catalog_service

        sources = ("managed_learning", "microsoft_learn", "coursera")
        found: list[dict] = []
        seen: set[str] = set()
        for source in sources:
            try:
                result = await dynamic_catalog_service.search_catalog(
                    source=source,
                    q=name,
                    provider=provider or None,
                    page=1,
                    page_size=25,
                    organization_id=organization_id,
                )
            except Exception:
                continue
            for course in result.get("courses") or []:
                uid = (course.get("uid") or "").strip()
                title = (course.get("title") or "").strip()
                if not uid or not title or uid in seen:
                    continue
                if title.lower() != name.strip().lower():
                    continue
                course_prov = (course.get("provider") or course.get("source") or "").strip().lower()
                if provider and course_prov and provider.strip().lower() not in course_prov:
                    continue
                seen.add(uid)
                found.append(course)
        return found

    for r in data.get("roadmaps") or []:
        course_id = (r.get("course_id") or "").strip()
        course_name = (r.get("course_name") or "").strip()
        provider = (r.get("provider") or "").strip()

        if course_id:
            # Enrich from Catalog Index / live catalog when possible.
            idx = index_by_id.get(course_id.lower())
            if idx:
                r["course_name"] = course_name or idx.get("course_name") or course_id
                if not r.get("catalog_type") and idx.get("catalog_type"):
                    r["catalog_type"] = idx.get("catalog_type")
                if not r.get("provider") and idx.get("provider"):
                    r["provider"] = idx.get("provider")
                continue
            try:
                from app.services import catalog_service

                item = await catalog_service.get_course_by_uid(course_id)
            except Exception:
                item = None
            if item:
                r["course_name"] = course_name or item.get("title") or course_id
                if not r.get("catalog_type") and item.get("type"):
                    r["catalog_type"] = item.get("type")
                if not r.get("provider"):
                    r["provider"] = item.get("provider") or item.get("source")
            continue

        if not course_name:
            details.append({
                "sheet": "Career Roadmaps",
                "row": r.get("_row"),
                "column": "Course Name",
                "reason": "Provide Course Name (or Course ID) for each roadmap row.",
            })
            continue

        hits = _index_matches(course_name, provider)
        if len(hits) == 1:
            hit = hits[0]
            r["course_id"] = hit.get("course_id") or ""
            r["course_name"] = hit.get("course_name") or course_name
            if not r.get("catalog_type") and hit.get("catalog_type"):
                r["catalog_type"] = hit.get("catalog_type")
            if not r.get("provider") and hit.get("provider"):
                r["provider"] = hit.get("provider")
            continue
        if len(hits) > 1:
            details.append({
                "sheet": "Career Roadmaps",
                "row": r.get("_row"),
                "column": "Course Name",
                "reason": (
                    f"Course Name '{course_name}' matches {len(hits)} rows in Catalog Index — "
                    "set Course ID or a more specific Provider."
                ),
            })
            continue

        live = await _live_matches(course_name, provider)
        if len(live) == 1:
            hit = live[0]
            r["course_id"] = hit.get("uid") or ""
            r["course_name"] = hit.get("title") or course_name
            if not r.get("catalog_type") and hit.get("type"):
                r["catalog_type"] = hit.get("type")
            if not r.get("provider"):
                r["provider"] = hit.get("provider") or hit.get("source")
            # Fill skills/certs from catalog metadata when Excel left them blank.
            if not r.get("skills"):
                raw_skills = hit.get("skills") or hit.get("skills_covered") or hit.get("subjects") or []
                if isinstance(raw_skills, str):
                    raw_skills = _split_csv_list(raw_skills)
                r["skills"] = [str(s).strip() for s in raw_skills if str(s).strip()][:12]
            if not r.get("certifications"):
                raw_certs = hit.get("certifications") or []
                if isinstance(raw_certs, str):
                    raw_certs = _split_csv_list(raw_certs)
                r["certifications"] = [str(c).strip() for c in raw_certs if str(c).strip()][:8]
            continue
        if len(live) > 1:
            details.append({
                "sheet": "Career Roadmaps",
                "row": r.get("_row"),
                "column": "Course Name",
                "reason": (
                    f"Course Name '{course_name}' matches {len(live)} catalog items — "
                    "copy the Course ID from Catalog Index or set Provider."
                ),
            })
            continue

        details.append({
            "sheet": "Career Roadmaps",
            "row": r.get("_row"),
            "column": "Course Name",
            "reason": (
                f"Could not resolve '{course_name}' to a Course ID. "
                "Add it to Catalog Index (Course ID + Course Name) or paste a real catalog Course ID."
            ),
        })

    return details


def validate_import_data(data: dict) -> dict:
    """Validate imported data and return a report of issues."""
    errors: list[str] = []
    warnings: list[str] = []
    details: list[dict[str, Any]] = []

    def _err(sheet: str, row: Any, column: str, reason: str) -> None:
        errors.append(f"{sheet} row {row}, column '{column}': {reason}")
        details.append({"sheet": sheet, "row": row, "column": column, "reason": reason})

    def _warn(sheet: str, row: Any, column: str, reason: str) -> None:
        warnings.append(f"{sheet} row {row}, column '{column}': {reason}")

    for issue in data.get("_sheet_issues") or []:
        _err(
            issue.get("sheet", "?"),
            issue.get("row", 1),
            issue.get("column", "-"),
            issue.get("reason", "Invalid workbook structure."),
        )

    # ── Departments ────────────────────────────────────────────────────────────
    dept_lookup: dict[str, str] = {}
    for d in data["departments"]:
        key = d["name"].lower()
        if key in dept_lookup:
            _err(
                "Departments",
                d.get("_row"),
                "Department Name",
                f"Duplicate department '{d['name']}' (already defined on row {dept_lookup[key]}).",
            )
        else:
            dept_lookup[key] = d.get("_row")

    # ── Career Roles ───────────────────────────────────────────────────────────
    role_keys: dict[tuple[str, str], int] = {}
    role_names_by_dept: dict[str, list[str]] = {}
    for r in data["roles"]:
        if not r["name"]:
            _err("Career Roles", r.get("_row"), "Role Name", "Role name is required.")
            continue
        if not r["department"]:
            _err("Career Roles", r.get("_row"), "Department", "Department is required.")
            continue
        key = (r["department"].lower(), r["name"].lower())
        if key in role_keys:
            _err(
                "Career Roles",
                r.get("_row"),
                "Role Name",
                f"Duplicate role '{r['name']}' in department '{r['department']}' "
                f"(already defined on row {role_keys[key]}).",
            )
            continue
        role_keys[key] = r.get("_row")
        role_names_by_dept.setdefault(r["department"].lower(), []).append(r["name"])
        if r["department"].lower() not in dept_lookup:
            _err(
                "Career Roles",
                r.get("_row"),
                "Department",
                f"Role '{r['name']}' references department '{r['department']}' "
                "which is not defined on the Departments sheet.",
            )

    all_role_names = {key[1] for key in role_keys}
    role_dept_keys = {
        f"{(r.get('department') or '').strip().lower()}::{(r.get('name') or '').strip().lower()}"
        for r in data["roles"]
    }

    for r in data["roles"]:
        next_role = (r.get("next_role") or "").strip().lower()
        if not next_role:
            continue
        if next_role == r["name"].lower():
            _err(
                "Career Roles",
                r.get("_row"),
                "Next Role",
                f"Role '{r['name']}' cannot be its own next role.",
            )
            continue
        dept_roles = [rn.lower() for rn in role_names_by_dept.get(r["department"].lower(), [])]
        if next_role not in dept_roles:
            _err(
                "Career Roles",
                r.get("_row"),
                "Next Role",
                f"Next role '{r.get('next_role')}' is not defined for department '{r['department']}'.",
            )
            continue
        visited: set[str] = set()
        node = r["name"].lower()
        dept = r["department"].lower()
        while node:
            if node in visited:
                _err(
                    "Career Roles",
                    r.get("_row"),
                    "Next Role",
                    f"Circular promotion chain detected involving role '{r['name']}'.",
                )
                break
            visited.add(node)
            node = next(
                (
                    (rr.get("next_role") or "").lower()
                    for rr in data["roles"]
                    if rr["department"].lower() == dept and rr["name"].lower() == node
                ),
                "",
            )

    # Legacy Skills / Certifications / Course Catalog (optional)
    for s in data.get("skills") or []:
        if s["role_name"].lower() not in all_role_names:
            _err(
                "Skills",
                s.get("_row"),
                "Role",
                f"Skill references role '{s['role_name']}' which is not defined on Career Roles.",
            )
    for c in data.get("certifications") or []:
        if c["role_name"].lower() not in all_role_names:
            _err(
                "Certifications",
                c.get("_row"),
                "Role",
                f"Certification references role '{c['role_name']}' which is not defined on Career Roles.",
            )

    course_ids = {c["course_id"] for c in data.get("courses") or []}
    roadmap_sheet = "Career Roadmaps" if data.get("_format") != "legacy" else "Learning Roadmap"
    roadmap_keys: set[tuple[str, str, str]] = set()
    for r in data["roadmaps"]:
        if not r["role_name"]:
            _err(roadmap_sheet, r.get("_row"), "Role", "Role is required.")
            continue
        course_id = (r.get("course_id") or "").strip()
        course_name = (r.get("course_name") or "").strip()
        if not course_id and not course_name:
            _err(
                roadmap_sheet,
                r.get("_row"),
                "Course Name",
                "Provide Course Name or Course ID (name can resolve via Catalog Index).",
            )
            continue
        if not course_id:
            _err(
                roadmap_sheet,
                r.get("_row"),
                "Course ID",
                "Course ID is still empty after catalog lookup — add the course to Catalog Index "
                "or paste a Course ID from that sheet.",
            )
            continue
        if not course_name and course_id not in course_ids:
            _warn(
                roadmap_sheet,
                r.get("_row"),
                "Course Name",
                "Course Name is empty; the Course ID will be shown as the label.",
            )
        dept = (r.get("department") or "").strip()
        key = (dept.lower(), r["role_name"].lower(), course_id.lower())
        if key in roadmap_keys:
            _err(
                roadmap_sheet,
                r.get("_row"),
                "Course ID",
                f"Duplicate roadmap entry '{course_id}' for role '{r['role_name']}'.",
            )
            continue
        roadmap_keys.add(key)
        if dept:
            rd_key = f"{dept.lower()}::{r['role_name'].lower()}"
            if rd_key not in role_dept_keys:
                _err(
                    roadmap_sheet,
                    r.get("_row"),
                    "Role",
                    f"Roadmap references role '{r['role_name']}' in department '{dept}' "
                    "which is not defined on the Career Roles sheet.",
                )
        elif r["role_name"].lower() not in all_role_names:
            _err(
                roadmap_sheet,
                r.get("_row"),
                "Role",
                f"Roadmap references role '{r['role_name']}' which is not defined on the Career Roles sheet.",
            )
        # Legacy Learning Roadmap required Course Catalog; new format does not.
        if data.get("_format") == "legacy" and r["course_id"] not in course_ids:
            _err(
                "Learning Roadmap",
                r.get("_row"),
                "Course ID",
                f"Roadmap references course ID '{r['course_id']}' which is not defined on the Course Catalog sheet.",
            )

    # ── Promotion Rules ────────────────────────────────────────────────────────
    rule_keys: set[str] = set()
    for p in data["promotion_rules"]:
        if not p["role_name"]:
            _err("Promotion Rules", p.get("_row"), "Role", "Role is required.")
            continue
        dept = (p.get("department") or "").strip()
        key = f"{dept.lower()}::{p['role_name'].lower()}"
        if key in rule_keys:
            _err(
                "Promotion Rules",
                p.get("_row"),
                "Role",
                f"Duplicate promotion rule for '{p['role_name']}'"
                + (f" in {dept}" if dept else "")
                + ".",
            )
            continue
        rule_keys.add(key)
        if dept:
            if key not in role_dept_keys:
                _err(
                    "Promotion Rules",
                    p.get("_row"),
                    "Role",
                    f"Promotion rule references role '{p['role_name']}' in department '{dept}' "
                    "which is not defined on the Career Roles sheet.",
                )
        elif p["role_name"].lower() not in all_role_names:
            _err(
                "Promotion Rules",
                p.get("_row"),
                "Role",
                f"Promotion rule references role '{p['role_name']}' which is not defined on the Career Roles sheet.",
            )

    if not data["departments"]:
        warnings.append("No departments found — add them on the Departments sheet first.")
    if not data["roles"]:
        warnings.append("No roles found on the Career Roles sheet.")
    if not data["roadmaps"]:
        warnings.append("No career roadmap rows — you can add learning items later in the UI.")

    # Catalog Index is a lookup sheet only — never counted as imported assignments.
    counts = {
        "departments": len(data.get("departments") or []),
        "roles": len(data.get("roles") or []),
        "roadmaps": len(data.get("roadmaps") or []),
        "promotion_rules": len(data.get("promotion_rules") or []),
    }
    if data.get("skills"):
        counts["skills"] = len(data["skills"])
    if data.get("certifications"):
        counts["certifications"] = len(data["certifications"])
    if data.get("courses"):
        counts["courses"] = len(data["courses"])

    return {
        "valid": not errors,
        "errors": errors,
        "warnings": warnings,
        "details": details,
        "counts": counts,
    }


async def apply_import_data(organization_id: str, data: dict) -> dict:
    """Apply validated import data atomically.

    The framework is replaced all-or-nothing: existing documents are snapshotted,
    cleared, and the imported data is written back. If any write fails, the
    snapshot is restored so a failed import never leaves a partially-applied
    framework.
    """
    resolve_errors = await resolve_roadmap_catalog_refs(organization_id, data)
    report = validate_import_data(data)
    if resolve_errors:
        for issue in resolve_errors:
            report["errors"].append(
                f"{issue.get('sheet')} row {issue.get('row')}, column '{issue.get('column')}': "
                f"{issue.get('reason')}"
            )
            report.setdefault("details", []).append(issue)
        report["valid"] = False
    if not report["valid"]:
        raise ValueError("Import contains validation errors: " + "; ".join(report["errors"]))

    oid = _oid_filter(organization_id)
    collections = (
        database.org_framework_departments,
        database.org_framework_roles,
        database.org_framework_skills,
        database.org_framework_certifications,
        database.org_framework_courses,
        database.org_framework_roadmaps,
        database.org_framework_promotion_rules,
    )

    snapshots: dict[str, list[dict]] = {}
    for coll in collections:
        snapshots[coll.name] = await coll.find(oid).to_list(100000)

    async def _restore() -> None:
        for coll in collections:
            await coll.delete_many(oid)
        for coll in collections:
            if snapshots.get(coll.name):
                await coll.insert_many(snapshots[coll.name])

    now = _now()

    def _without_row(record: dict) -> dict:
        return {k: v for k, v in record.items() if k != "_row"}

    try:
        for coll in collections:
            await coll.delete_many(oid)

        departments = [
            {
                "organization_id": organization_id,
                **_without_row(d),
                "created_at": now,
                "updated_at": now,
            }
            for d in data["departments"]
        ]
        if departments:
            await database.org_framework_departments.insert_many(departments)

        roles = []
        for r in data["roles"]:
            role_id = f"{organization_id}:{r['department']}:{r['name']}"
            roles.append({
                "organization_id": organization_id,
                "role_id": role_id,
                "department": r["department"],
                "name": r["name"],
                "next_role": r.get("next_role") or None,
                "level_number": 1,
                "description": r.get("description") or "",
                "created_at": now,
                "updated_at": now,
            })
        if roles:
            await database.org_framework_roles.insert_many(roles)

        # Optional legacy sheets still applied when present.
        skills = [
            {
                "organization_id": organization_id,
                "skill_id": f"{organization_id}:{s['role_name']}:{s['skill_name']}",
                **_without_row(s),
                "created_at": now,
                "updated_at": now,
            }
            for s in data.get("skills") or []
        ]
        if skills:
            await database.org_framework_skills.insert_many(skills)

        certifications = [
            {
                "organization_id": organization_id,
                "cert_id": f"{organization_id}:{c['role_name']}:{c['certification_name']}",
                **_without_row(c),
                "created_at": now,
                "updated_at": now,
            }
            for c in data.get("certifications") or []
        ]
        if certifications:
            await database.org_framework_certifications.insert_many(certifications)

        courses = [
            {
                **_without_row(c),
                "organization_id": organization_id,
                "created_at": now,
                "updated_at": now,
            }
            for c in data.get("courses") or []
        ]
        if courses:
            await database.org_framework_courses.insert_many(courses)

        roadmaps = []
        for idx, r in enumerate(data["roadmaps"]):
            skills_list = r.get("skills") if isinstance(r.get("skills"), list) else _split_csv_list(r.get("skills"))
            certs_list = (
                r.get("certifications")
                if isinstance(r.get("certifications"), list)
                else _split_csv_list(r.get("certifications"))
            )
            roadmaps.append({
                "organization_id": organization_id,
                "roadmap_id": f"{organization_id}:{r['role_name']}:{r['course_id']}:{idx}:{now.microsecond}",
                "role_name": r["role_name"],
                "course_id": r["course_id"],
                "course_name": (r.get("course_name") or "").strip() or r["course_id"],
                "provider": (r.get("provider") or None),
                "catalog_type": (r.get("catalog_type") or None),
                "skills": skills_list[:12],
                "certifications": certs_list[:8],
                "mandatory": r.get("mandatory", True),
                "order": r.get("order") or (idx + 1),
                "prerequisite_course_id": None,
                "created_at": now,
                "updated_at": now,
            })
        if roadmaps:
            await database.org_framework_roadmaps.insert_many(roadmaps)

        roles_by_name: dict[str, list[dict]] = {}
        for role in data["roles"]:
            roles_by_name.setdefault((role.get("name") or "").strip().lower(), []).append(role)
        promotion_rules = []
        for p in data["promotion_rules"]:
            row = {**_without_row(p), "organization_id": organization_id, "updated_at": now}
            dept = (row.get("department") or "").strip()
            if not dept:
                matches = roles_by_name.get((row.get("role_name") or "").strip().lower()) or []
                if len(matches) == 1:
                    dept = (matches[0].get("department") or "").strip()
            if not dept:
                continue
            row["department"] = dept
            promotion_rules.append(row)
        if promotion_rules:
            await database.org_framework_promotion_rules.insert_many(promotion_rules)

        # Derive L1/L2/… from Next Role links after bulk insert.
        dept_names = sorted({(r.get("department") or "").strip() for r in data["roles"] if r.get("department")})
        for dept_name in dept_names:
            await recompute_department_level_numbers(organization_id, dept_name)
    except Exception:
        await _restore()
        raise

    await create_version_snapshot(organization_id, "Excel Import")
    return report["counts"]
