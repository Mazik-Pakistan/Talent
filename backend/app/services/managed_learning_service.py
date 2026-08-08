"""Managed learning roadmap service.

This layer stores organization-managed courses in MongoDB and exposes a
spreadsheet import flow that can understand hierarchical roadmap templates
without assuming a fixed flat row structure.

It is intentionally designed to plug into the existing learning module rather
than replace it: catalog browsing still lives behind the current learning
router, assignments keep using the existing collection, and certificates reuse
that same verification flow.
"""

from __future__ import annotations

import csv
import io
import random
import re
from collections import defaultdict
from datetime import UTC, datetime
from pathlib import Path
from typing import Any, Iterable

from bson import ObjectId
from fastapi import HTTPException, UploadFile, status

from app.core.database import database
from app.core.rbac import CurrentUser
from app.schemas.learning import (
    ManagedLearningCourseCreateRequest,
    ManagedLearningImportPreview,
    ManagedLearningImportRow,
    ManagedLearningCourseUpdateRequest,
)
from app.schemas.provider import LearningProviderCreate
from app.services.dashboard_service import create_notification
from app.services.provider_service import provider_service

MANAGED_SOURCE = "managed_learning"
MAX_IMPORT_BYTES = 8 * 1024 * 1024
MAX_IMPORT_ROWS = 5000

COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "designation": (
        "designation",
        "job title",
        "job_title",
        "role",
        "role title",
        "position",
        "track",
        "path",
        "category/competency/skill",
    ),
    "learning_month": (
        "learning month",
        "month",
        "month no",
        "month #",
        "learning month / phase",
        "phase",
        "module month",
        "month mapping",
    ),
    "category": ("category", "topic", "group", "bucket", "theme", "broad category"),
    "competency": (
        "competency",
        "skill",
        "skill area",
        "capability",
        "sub category",
        "subcategory",
        "skill descriptions",
        "learning outcomes",
        "associated remarks",
    ),
    "title": (
        "course name",
        "course title",
        "title",
        "learning title",
        "course",
        "name",
    ),
    "url": (
        "course url",
        "url",
        "link",
        "course link",
        "resource url",
        "resource link",
    ),
    "duration_minutes": (
        "duration",
        "duration minutes",
        "minutes",
        "time",
        "estimated duration",
        "length",
        "workload",
        "total",
    ),
    "provider": ("provider", "source", "platform", "vendor", "origin"),
    "description": ("description", "summary", "notes", "course description"),
}

PREFERRED_SHEETS = (
    "linkedin",
    "learning",
    "roadmap",
    "road maps",
    "courses",
    "catalog",
    "curriculum",
)


def _now() -> datetime:
    return datetime.now(UTC)


def _iso(value: Any) -> Any:
    return value.isoformat() if hasattr(value, "isoformat") else value


def _clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, datetime):
        value = value.date()
    if hasattr(value, "isoformat"):
        try:
            return str(value.isoformat()).strip()
        except Exception:  # noqa: BLE001
            pass
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _slug(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", (value or "").lower()).strip("-")


def _normalize_provider_name(value: Any) -> str:
    if value is None:
        return "Managed Learning"
    text = " ".join(str(value).split())
    if not text:
        return "Managed Learning"
    aliases = {
        "linkedin learning": "LinkedIn Learning",
        "linkedin": "LinkedIn Learning",
        "microsoft learn": "Microsoft Learn",
        "microsoft": "Microsoft Learn",
        "coursera": "Coursera",
        "udemy": "Udemy",
        "pluralsight": "Pluralsight",
        "skillsoft": "Skillsoft",
        "datacamp": "DataCamp",
        "edx": "edX",
    }
    lower = text.lower()
    return aliases.get(lower, text)


def _provider_slug(value: str) -> str:
    return _slug(_normalize_provider_name(value)) or "managed-learning"


def _normalize(value: Any) -> str:
    return " ".join(_clean(value).lower().split())


def _matches_alias(cell: str, aliases: tuple[str, ...]) -> bool:
    if cell in aliases:
        return True
    if len(cell) >= 4:
        return any(len(alias) >= 4 and alias in cell for alias in aliases)
    return False


def _first_line(value: Any) -> str:
    for line in _clean(value).splitlines():
        line = line.strip()
        if line:
            return line
    return ""


def _compact(values: Iterable[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for raw in values:
        cleaned = " ".join(str(raw or "").split())
        if not cleaned:
            continue
        key = cleaned.lower()
        if key in seen:
            continue
        seen.add(key)
        out.append(cleaned)
    return out


def _parse_duration(raw: Any, unit: str = "minutes") -> int | None:
    text = _clean(raw).lower()
    if not text:
        return None
    if isinstance(raw, (int, float)) and not isinstance(raw, bool):
        value = float(raw)
        total = int(round(value * 60)) if unit == "hours" else int(round(value))
        return total if total > 0 else None

    hours = 0
    minutes = 0
    matched = False
    for number, time_unit in re.findall(r"(\d+(?:\.\d+)?)\s*([hms]|hours?|minutes?|mins?|hr|hrs)", text):
        matched = True
        value = float(number)
        if time_unit.startswith("h"):
            hours += value
        elif time_unit.startswith("m"):
            minutes += value
        elif time_unit.startswith("s"):
            minutes += value / 60.0
    if matched:
        total = int(round(hours * 60 + minutes))
        return total if total > 0 else None

    digits = re.findall(r"\d+(?:\.\d+)?", text)
    if not digits:
        return None
    value = float(digits[0])
    if any(token in text for token in ("hour", "hr", "h")):
        return int(round(value * 60))
    if unit == "hours":
        return int(round(value * 60)) if value > 0 else None
    return int(round(value)) if value > 0 else None


def _split_label(value: str) -> str:
    return _clean(value).replace("_", " ")


def _course_key(*, provider: str, designation: str, learning_month: str, category: str, competency: str, title: str, url: str | None) -> str:
    return "|".join(
        [
            _normalize(provider),
            _normalize(designation),
            _normalize(learning_month),
            _normalize(category),
            _normalize(competency),
            _normalize(title),
            _normalize(url),
        ]
    )


def _hierarchy_key(*, designation: str, learning_month: str, category: str, competency: str) -> str:
    return "|".join([_normalize(designation), _normalize(learning_month), _normalize(category), _normalize(competency)])


def _course_uid(doc: dict) -> str:
    return f"learning_course:{doc['_id']}"


def _public_course(doc: dict) -> dict:
    return {
        "uid": _course_uid(doc),
        "type": "course",
        "source": MANAGED_SOURCE,
        "provider": doc.get("provider") or "Managed Learning",
        "designation": doc.get("designation") or "",
        "learning_month": doc.get("learning_month") or "",
        "category": doc.get("category") or "",
        "competency": doc.get("competency") or "",
        "title": doc.get("title") or "",
        "summary": doc.get("description") or "",
        "url": doc.get("url") or "",
        "duration_minutes": doc.get("duration_minutes"),
        "archived": bool(doc.get("archived")),
        "source_filename": doc.get("source_filename"),
        "source_row": doc.get("source_row"),
        "source_kind": doc.get("source_kind") or "manual",
        "hierarchy_path": doc.get("hierarchy_path") or [],
        "hierarchy_key": doc.get("hierarchy_key"),
        "last_modified": _iso(doc.get("updated_at") or doc.get("created_at")),
    }


def _sortable_date(value):
    if isinstance(value, datetime):
        # Normalize naive datetimes (legacy docs) to UTC-aware so comparisons
        # never raise "can't compare offset-naive and offset-aware datetimes".
        return value if value.tzinfo is not None else value.replace(tzinfo=UTC)
    if isinstance(value, str):
        try:
            dt = datetime.fromisoformat(value.replace("Z", "+00:00"))
            return dt if dt.tzinfo is not None else dt.replace(tzinfo=UTC)
        except ValueError:
            pass
    return datetime.min.replace(tzinfo=UTC)


class ManagedLearningService:
    def _normalize_provider_name(self, value: Any) -> str:
        return _normalize_provider_name(value)

    def _provider_slug(self, value: str) -> str:
        return _provider_slug(value)

    def _collection_query(self, *, include_archived: bool = False) -> dict:
        query: dict[str, Any] = {}
        if not include_archived:
            query["$or"] = [{"archived": {"$exists": False}}, {"archived": False}]
        return query

    def _normalize_request_payload(self, payload: ManagedLearningCourseCreateRequest | ManagedLearningCourseUpdateRequest) -> dict[str, Any]:
        data = payload.model_dump()
        for key in ("title", "designation", "learning_month", "category", "competency"):
            if key in data and data[key] is not None:
                data[key] = " ".join(str(data[key]).split())
        if "provider" in data and data.get("provider") is not None:
            data["provider"] = _normalize_provider_name(data["provider"])
        elif "provider" in data:
            data["provider"] = "Managed Learning"
        if data.get("description") is not None:
            data["description"] = data["description"].strip()
        if data.get("provider") is None:
            data["provider"] = "Managed Learning"
        for key in ("designation", "learning_month", "category", "competency"):
            data[key] = data.get(key) or ""
        return data

    async def _ensure_unique_indexes(self) -> None:
        # Collection-specific indexes are created in app.core.database.
        return None

    async def _find_existing(self, course_key: str) -> dict | None:
        return await database.learning_courses.find_one({"course_key": course_key})

    async def _ensure_provider(self, provider_name: str | None, *, current_user: CurrentUser | None = None) -> dict | None:
        normalized = _normalize_provider_name(provider_name)
        if not normalized:
            return None
        slug = _provider_slug(normalized)
        existing = await database.learning_providers.find_one({"slug": slug})
        if existing:
            return existing
        doc = {
            "name": normalized,
            "slug": slug,
            "active": True,
            "created_at": _now(),
            "created_by_id": current_user.id if current_user else None,
        }
        result = await database.learning_providers.insert_one(doc)
        doc["_id"] = result.inserted_id
        return doc

    async def _save_course_doc(self, doc: dict, organization_id: str | None = None) -> dict:
        doc = dict(doc)
        if organization_id:
            doc["organization_id"] = organization_id
        now = _now()
        doc.setdefault("created_at", now)
        doc["updated_at"] = now
        doc["provider"] = _normalize_provider_name(doc.get("provider")) or "Managed Learning"
        await self._ensure_provider(doc.get("provider"))
        doc["hierarchy_path"] = [
            part
            for part in [doc.get("designation"), doc.get("learning_month"), doc.get("category"), doc.get("competency")]
            if part
        ]
        doc["course_key"] = _course_key(
            provider=doc.get("provider") or "Managed Learning",
            designation=doc.get("designation") or "",
            learning_month=doc.get("learning_month") or "",
            category=doc.get("category") or "",
            competency=doc.get("competency") or "",
            title=doc.get("title") or "",
            url=doc.get("url") or None,
        )
        doc["hierarchy_key"] = _hierarchy_key(
            designation=doc.get("designation") or "",
            learning_month=doc.get("learning_month") or "",
            category=doc.get("category") or "",
            competency=doc.get("competency") or "",
        )
        if not doc.get("provider"):
            doc["provider"] = "Managed Learning"
        if doc.get("_id"):
            await database.learning_courses.update_one({"_id": doc["_id"]}, {"$set": doc})
            return await database.learning_courses.find_one({"_id": doc["_id"]})
        result = await database.learning_courses.insert_one(doc)
        doc["_id"] = result.inserted_id
        return doc

    def _course_changes(self, existing: dict | None, incoming: dict) -> tuple[bool, list[str]]:
        if not existing:
            return True, []
        comparable = ("provider", "designation", "learning_month", "category", "competency", "title", "url", "duration_minutes", "description", "archived")
        diffs = []
        changed = False
        for key in comparable:
            if _normalize(existing.get(key)) != _normalize(incoming.get(key)):
                changed = True
                diffs.append(key)
        return changed, diffs

    def _row_to_preview_row(
        self,
        *,
        row_number: int,
        row: dict[str, Any],
        existing: dict | None,
        seen_keys: set[str],
        import_keys: set[str],
    ) -> ManagedLearningImportRow:
        designation = row.get("designation") or ""
        learning_month = row.get("learning_month") or ""
        category = row.get("category") or ""
        competency = row.get("competency") or ""
        title = row.get("title") or ""
        url = row.get("url") or ""
        provider = _normalize_provider_name(row.get("provider") or "Managed Learning")
        duration_minutes = row.get("duration_minutes")
        description = row.get("description") or ""
        issues: list[str] = []

        if not designation:
            issues.append("designation missing")
        if not title:
            issues.append("course title missing")
        if not learning_month:
            issues.append("learning month missing")
        if not category:
            issues.append("category missing")
        if not competency:
            issues.append("competency missing")
        if not provider:
            issues.append("provider missing")
        if url and not re.match(r"^https?://", url, re.I):
            issues.append("course url is invalid")
        if duration_minutes is not None and duration_minutes <= 0:
            issues.append("duration must be positive")

        course_key = _course_key(
            provider=provider,
            designation=designation,
            learning_month=learning_month,
            category=category,
            competency=competency,
            title=title,
            url=url or None,
        )
        if course_key in seen_keys:
            status_value = "duplicate"
            issues.append("duplicate row in file")
        else:
            seen_keys.add(course_key)
            import_keys.add(course_key)
            if issues:
                status_value = "invalid"
            elif existing:
                changed, diffs = self._course_changes(existing, {
                    "provider": provider,
                    "designation": designation,
                    "learning_month": learning_month,
                    "category": category,
                    "competency": competency,
                    "title": title,
                    "url": url or None,
                    "duration_minutes": duration_minutes,
                    "description": description,
                    "archived": False,
                })
                status_value = "updated" if changed else "duplicate"
                if not changed:
                    issues.append("matches existing course")
                else:
                    issues.extend([f"{diff} changed" for diff in diffs])
            else:
                status_value = "new"

        return ManagedLearningImportRow(
            row=row_number,
            designation=designation or None,
            learning_month=learning_month or None,
            category=category or None,
            competency=competency or None,
            title=title or None,
            url=url or None,
            duration_minutes=duration_minutes,
            provider=provider or None,
            description=description or None,
            status=status_value,
            issues=_compact(issues),
            existing_id=str(existing["_id"]) if existing and existing.get("_id") else None,
        )

    def _sheet_candidates(self, workbook) -> Any:
        by_name = {str(ws.title or "").strip().lower(): ws for ws in workbook.worksheets}
        for name in PREFERRED_SHEETS:
            if name in by_name:
                return by_name[name]
        for ws in workbook.worksheets:
            title = str(ws.title or "").strip().lower()
            if any(token in title for token in PREFERRED_SHEETS):
                return ws
        return workbook.active

    def _sheet_rows(self, sheet) -> list[list[Any]]:
        rows = list(sheet.iter_rows(values_only=False))
        merged_map: dict[tuple[int, int], Any] = {}
        for merged in getattr(sheet, "merged_cells", []).ranges:
            min_col, min_row, max_col, max_row = merged.bounds
            anchor = sheet.cell(min_row, min_col).value
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    merged_map[(row, col)] = anchor
        out: list[list[Any]] = []
        for row_idx, cells in enumerate(rows, start=1):
            values: list[Any] = []
            for col_idx, cell in enumerate(cells, start=1):
                value = cell.value
                if value in (None, ""):
                    value = merged_map.get((row_idx, col_idx), value)
                values.append(value)
            out.append(values)
        return out

    def _find_header_row(self, rows: list[list[Any]]) -> int:
        best_index = 0
        best_score = -1
        for idx, row in enumerate(rows[:50]):
            normalized = [_normalize(cell) for cell in row]
            score = 0
            for aliases in COLUMN_ALIASES.values():
                if any(_matches_alias(cell, aliases) for cell in normalized):
                    score += 1
            score += sum(1 for cell in normalized if cell)
            if score > best_score:
                best_score = score
                best_index = idx
        return best_index

    def _resolve_columns(self, header: list[Any]) -> tuple[dict[str, int | None], str]:
        normalized = [_normalize(cell) for cell in header]
        indexes: dict[str, int | None] = {key: None for key in COLUMN_ALIASES}
        claimed: set[int] = set()

        for key, aliases in COLUMN_ALIASES.items():
            for idx, cell in enumerate(normalized):
                if idx in claimed:
                    continue
                if cell in aliases:
                    indexes[key] = idx
                    claimed.add(idx)
                    break

        for key, aliases in COLUMN_ALIASES.items():
            if indexes[key] is not None:
                continue
            for idx, cell in enumerate(normalized):
                if idx in claimed:
                    continue
                if len(cell) >= 4 and any(len(alias) >= 4 and alias in cell for alias in aliases):
                    indexes[key] = idx
                    claimed.add(idx)
                    break

        duration_unit = "minutes"
        duration_idx = indexes.get("duration_minutes")
        if duration_idx is not None:
            header_cell = normalized[duration_idx]
            if header_cell == "total" or "hour" in header_cell:
                duration_unit = "hours"
        return indexes, duration_unit

    def _row_to_record(
        self,
        row: list[Any],
        indexes: dict[str, int | None],
        *,
        previous_context: dict[str, str],
        row_number: int,
        filename: str | None,
        sheet_name: str,
        duration_unit: str = "minutes",
    ) -> dict[str, Any] | None:
        def get(key: str) -> str:
            idx = indexes.get(key)
            if idx is None or idx >= len(row):
                return ""
            return _clean(row[idx])

        designation = _first_line(get("designation")) or previous_context.get("designation", "")
        learning_month = get("learning_month") or previous_context.get("learning_month", "")
        category = get("category") or previous_context.get("category", "")
        competency = get("competency") or previous_context.get("competency", "")
        title = get("title")
        url = get("url")
        provider = _normalize_provider_name(get("provider") or previous_context.get("provider", "Managed Learning"))
        description = get("description")
        duration_minutes = _parse_duration(row[indexes["duration_minutes"]], duration_unit) if indexes.get("duration_minutes") is not None else None

        if not any([designation, learning_month, category, competency, title, url, description, provider]):
            return None

        previous_context.update(
            {
                "designation": designation or previous_context.get("designation", ""),
                "learning_month": learning_month or previous_context.get("learning_month", ""),
                "category": category or previous_context.get("category", ""),
                "competency": competency or previous_context.get("competency", ""),
                "provider": provider or previous_context.get("provider", "Managed Learning"),
            }
        )

        if not title:
            return None

        return {
            "row": row_number,
            "designation": designation,
            "learning_month": learning_month,
            "category": category,
            "competency": competency,
            "title": title,
            "url": url,
            "provider": provider or "Managed Learning",
            "description": description,
            "duration_minutes": duration_minutes,
            "filename": filename,
            "sheet_name": sheet_name,
            "course_key": _course_key(
                provider=provider or "Managed Learning",
                designation=designation,
                learning_month=learning_month,
                category=category,
                competency=competency,
                title=title,
                url=url or None,
            ),
            "hierarchy_key": _hierarchy_key(
                designation=designation,
                learning_month=learning_month,
                category=category,
                competency=competency,
            ),
        }

    async def _read_upload(self, file: UploadFile) -> tuple[list[list[Any]], str]:
        name = (file.filename or "").lower()
        raw = await file.read()
        if len(raw) > MAX_IMPORT_BYTES:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="File is too large (8MB limit).")
        if name.endswith(".csv"):
            text = raw.decode("utf-8-sig", errors="replace")
            rows = list(csv.reader(io.StringIO(text)))
            if not rows:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="The CSV file is empty.")
            return rows, "CSV"
        if not (name.endswith(".xlsx") or name.endswith(".xlsm")):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Please upload a .xlsx or .csv spreadsheet.")

        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(raw), data_only=True, read_only=False)
        sheet = self._sheet_candidates(workbook)
        rows = self._sheet_rows(sheet)
        if not rows:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="The spreadsheet is empty.")
        return rows, sheet.title or "Sheet1"

    async def preview_import(self, current_user: CurrentUser, file: UploadFile, provider_name: str | None = None) -> dict:
        rows, sheet_name = await self._read_upload(file)
        header_index = self._find_header_row(rows)
        header = rows[header_index] if rows else []
        indexes, duration_unit = self._resolve_columns(header)
        records = rows[header_index + 1 :]

        seen_keys: set[str] = set()
        preview_rows: list[ManagedLearningImportRow] = []
        new_courses = updated_courses = duplicate_courses = invalid_rows = 0
        fallback_provider = _normalize_provider_name(provider_name or None)
        context = {"designation": "", "learning_month": "", "category": "", "competency": "", "provider": fallback_provider}

        existing_docs = await database.learning_courses.find({}).to_list(length=5000)
        existing_by_key = {doc.get("course_key"): doc for doc in existing_docs if doc.get("course_key")}

        for row_number, row in enumerate(records, start=header_index + 2):
            if not any(_clean(cell) for cell in row):
                continue
            record = self._row_to_record(
                row,
                indexes,
                previous_context=context,
                row_number=row_number,
                filename=file.filename,
                sheet_name=sheet_name,
                duration_unit=duration_unit,
            )
            if not record:
                continue
            existing = existing_by_key.get(record["course_key"])
            preview_row = self._row_to_preview_row(
                row_number=row_number,
                row=record,
                existing=existing,
                seen_keys=seen_keys,
                import_keys=set(),
            )
            preview_rows.append(preview_row)
            if preview_row.status == "new":
                new_courses += 1
            elif preview_row.status == "updated":
                updated_courses += 1
            elif preview_row.status == "duplicate":
                duplicate_courses += 1
            else:
                invalid_rows += 1
            if len(preview_rows) >= MAX_IMPORT_ROWS:
                break

        return ManagedLearningImportPreview(
            filename=file.filename,
            total_rows=len(preview_rows),
            new_courses=new_courses,
            updated_courses=updated_courses,
            duplicate_courses=duplicate_courses,
            invalid_rows=invalid_rows,
            rows=preview_rows,
        ).model_dump()

    async def import_file(self, current_user: CurrentUser, file: UploadFile, provider_name: str | None = None) -> dict:
        preview = await self.preview_import(current_user, file, provider_name=provider_name)
        imported = 0
        updated = 0
        skipped = 0
        errors = []

        for row in preview.get("rows") or []:
            if row.get("status") == "invalid":
                skipped += 1
                continue
            payload = {
                # Preserve the provider resolved during preview so each tab stays isolated.
                "provider": _normalize_provider_name(row.get("provider") or provider_name or "Managed Learning"),
                "designation": row.get("designation") or "",
                "learning_month": row.get("learning_month") or "",
                "category": row.get("category") or "",
                "competency": row.get("competency") or "",
                "title": row.get("title") or "",
                "url": row.get("url") or None,
                "description": row.get("description") or None,
                "duration_minutes": row.get("duration_minutes"),
                "archived": False,
                "source_kind": "import",
                "source_filename": file.filename,
                "source_row": row.get("row"),
            }
            existing = await database.learning_courses.find_one({"_id": ObjectId(row["existing_id"])} if row.get("existing_id") and ObjectId.is_valid(row["existing_id"]) else {"course_key": _course_key(
                provider=payload["provider"],
                designation=payload["designation"],
                learning_month=payload["learning_month"],
                category=payload["category"],
                competency=payload["competency"],
                title=payload["title"],
                url=payload["url"],
            )})
            try:
                if existing:
                    payload["_id"] = existing["_id"]
                    await self._save_course_doc({**existing, **payload}, organization_id=current_user.organization_id)
                    updated += 1
                else:
                    await self._save_course_doc({
                        **payload,
                        "created_by_id": current_user.id,
                        "updated_by_id": current_user.id,
                    }, organization_id=current_user.organization_id)
                    imported += 1
            except Exception as exc:  # noqa: BLE001
                errors.append({"row": row.get("row"), "error": str(exc)})

        return {
            "message": f"Imported {imported} new course(s), updated {updated}, skipped {skipped}.",
            "imported": imported,
            "updated": updated,
            "skipped": skipped,
            "errors": errors,
            "preview": preview,
        }

    async def list_courses(
        self,
        *,
        q: str | None = None,
        provider: str | None = None,
        designation: str | None = None,
        learning_month: str | None = None,
        category: str | None = None,
        competency: str | None = None,
        archived: bool | None = None,
        sort_by: str | None = "newest",
        page: int = 1,
        page_size: int = 20,
        organization_id: str | None = None,
    ) -> dict:
        query = self._collection_query(include_archived=archived is True)
        if archived is True:
            query = {}
        if organization_id:
            org_filter = {"$or": [{"organization_id": organization_id}, {"organization_id": {"$exists": False}}]}
            if query:
                query = {"$and": [query, org_filter]}
            else:
                query = org_filter
        if designation:
            query["designation"] = {"$regex": f"^{re.escape(designation.strip())}$", "$options": "i"}
        if learning_month:
            query["learning_month"] = {"$regex": f"^{re.escape(learning_month.strip())}$", "$options": "i"}
        if category:
            query["category"] = {"$regex": f"^{re.escape(category.strip())}$", "$options": "i"}
        if competency:
            query["competency"] = {"$regex": f"^{re.escape(competency.strip())}$", "$options": "i"}
        docs = await database.learning_courses.find(query).to_list(length=5000)
        if provider:
            docs = [d for d in docs if _normalize_provider_name(d.get("provider") or "Managed Learning") == provider]
        if q:
            needle = q.strip().lower()
            docs = [
                doc
                for doc in docs
                if needle in " ".join(
                    [
                        doc.get("title") or "",
                        doc.get("designation") or "",
                        doc.get("learning_month") or "",
                        doc.get("category") or "",
                        doc.get("competency") or "",
                        doc.get("provider") or "",
                        *[str(t) for t in (doc.get("tags") or [])],
                        doc.get("instructor") or "",
                    ]
                ).lower()
            ]

        for doc in docs:
            for field in ("created_at", "updated_at"):
                value = doc.get(field)
                if isinstance(value, str):
                    try:
                        doc[field] = datetime.fromisoformat(value.replace("Z", "+00:00"))
                    except (ValueError, TypeError):
                        doc[field] = None

        def _sort_ts(doc: dict, *fields: str) -> str:
            """Return an ISO string for sorting — always a string, never raises."""
            for field in fields:
                value = doc.get(field)
                if value is None:
                    continue
                if isinstance(value, datetime):
                    try:
                        # Normalize naive → UTC so isoformat is consistent
                        if value.tzinfo is None:
                            value = value.replace(tzinfo=UTC)
                        return value.isoformat()
                    except Exception:  # noqa: BLE001
                        continue
                if isinstance(value, str) and value:
                    return value
            return "0000-00-00T00:00:00+00:00"

        # Phase 3: sortable catalog (newest first by default).
        if sort_by == "oldest":
            docs.sort(key=lambda d: _sort_ts(d, "created_at"))
        elif sort_by == "updated":
            docs.sort(key=lambda d: _sort_ts(d, "updated_at", "created_at"), reverse=True)
        elif sort_by == "title_asc":
            docs.sort(key=lambda d: str(d.get("title") or "").lower())
        elif sort_by == "title_desc":
            docs.sort(key=lambda d: str(d.get("title") or "").lower(), reverse=True)
        elif sort_by == "duration":
            docs.sort(key=lambda d: d.get("duration_minutes") if isinstance(d.get("duration_minutes"), (int, float)) else float("inf"))
        elif sort_by == "provider":
            docs.sort(key=lambda d: str(d.get("provider") or "").lower())
        else:
            # newest (default)
            docs.sort(key=lambda d: _sort_ts(d, "created_at"), reverse=True)

        total = len(docs)
        start = (page - 1) * page_size
        page_items = docs[start : start + page_size]
        hierarchy: dict[str, Any] = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
        for doc in docs:
            hierarchy[doc.get("designation") or "Unassigned"][doc.get("learning_month") or "Unassigned"][doc.get("category") or "Unassigned"].append(_public_course(doc))
        grouped = []
        for designation_key, months in hierarchy.items():
            grouped_months = []
            for month_key, categories in months.items():
                grouped_categories = []
                for category_key, courses in categories.items():
                    competencies: dict[str, list[dict]] = defaultdict(list)
                    for course in courses:
                        competencies[course.get("competency") or "Unassigned"].append(course)
                    grouped_categories.append(
                        {
                            "category": category_key,
                            "competencies": [
                                {"competency": comp, "courses": comp_courses}
                                for comp, comp_courses in sorted(competencies.items(), key=lambda item: item[0].lower())
                            ],
                        }
                    )
                grouped_months.append({"learning_month": month_key, "categories": sorted(grouped_categories, key=lambda item: item["category"].lower())})
            grouped.append({"designation": designation_key, "months": sorted(grouped_months, key=lambda item: item["learning_month"].lower())})
        pages = max(1, (total + page_size - 1) // page_size) if total else 1
        return {
            "courses": [_public_course(doc) for doc in page_items],
            "hierarchy": sorted(grouped, key=lambda item: item["designation"].lower()),
            "total": total,
            "page": page,
            "page_size": page_size,
            "pages": pages,
        }

    async def list_providers(self) -> list[dict]:
        docs = await database.learning_providers.find({"active": {"$ne": False}}).sort("name", 1).to_list(length=200)
        names = [_normalize_provider_name(doc.get("name")) for doc in docs if doc.get("name")]
        course_docs = await database.learning_courses.find({}, {"provider": 1}).to_list(length=5000)
        for doc in course_docs:
            name = _normalize_provider_name(doc.get("provider"))
            if name and name not in names:
                names.append(name)
        managed_names = []
        external_names = {"Microsoft Learn", "Coursera"}
        for name in names:
            if not name or name in external_names:
                continue
            managed_names.append(name)
        if any(name != "Managed Learning" for name in managed_names):
            managed_names = [name for name in managed_names if name != "Managed Learning"]
        ordered = sorted({name for name in managed_names if name}, key=lambda value: value.lower())
        return [{"name": name, "slug": _provider_slug(name)} for name in ordered]

    async def create_provider(self, current_user: CurrentUser, provider_name: str) -> dict:
        normalized = _normalize_provider_name(provider_name)
        if not normalized or not normalized.strip():
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Provider name is required.")
        slug = _provider_slug(normalized)
        existing = await database.learning_providers.find_one({"slug": slug})
        if existing:
            return {"provider": {"name": existing.get("name") or normalized, "slug": existing.get("slug") or slug}}
        try:
            result = await provider_service.create_provider(
                current_user,
                LearningProviderCreate(
                    name=normalized,
                    provider_type="manual",
                    import_method="manual",
                    active=True,
                ),
            )
            return result
        except HTTPException as exc:
            if exc.status_code == status.HTTP_409_CONFLICT:
                existing = await database.learning_providers.find_one({"slug": slug})
                if existing:
                    return {"provider": {"name": existing.get("name") or normalized, "slug": existing.get("slug") or slug}}
            raise

    async def list_facets(self, organization_id: str | None = None) -> dict:
        query = self._collection_query(include_archived=False)
        if organization_id:
            org_filter = {"$or": [{"organization_id": organization_id}, {"organization_id": {"$exists": False}}]}
            if query:
                query = {"$and": [query, org_filter]}
            else:
                query = org_filter
        docs = await database.learning_courses.find(query).to_list(length=5000)

        # Sync any course provider names into the registry first (idempotent).
        # This ensures providers added via import always appear in catalog tabs.
        await provider_service.sync_providers_from_courses()

        # Registry is the single source of truth for which providers appear in tabs.
        # Courses whose provider was deleted from the registry are NOT surfaced —
        # this ensures that deleting a provider immediately removes it from the catalog.
        EXCLUDED = {"Microsoft Learn", "Coursera"}
        registry_docs = await database.learning_providers.find(
            {"active": {"$ne": False}}
        ).sort("name", 1).to_list(length=500)
        registered_names: set[str] = set()
        providers: list[str] = []
        for provider_doc in registry_docs:
            provider_name = _normalize_provider_name(provider_doc.get("name"))
            if not provider_name or provider_name in EXCLUDED:
                continue
            key = provider_name.lower()
            if key in registered_names:
                continue
            registered_names.add(key)
            providers.append(provider_name)

        if any(name != "Managed Learning" for name in providers):
            providers = [name for name in providers if name != "Managed Learning"]
        providers = sorted(providers, key=lambda value: value.lower())
        designations = sorted({(doc.get("designation") or "").strip() for doc in docs if doc.get("designation")})
        months = sorted({(doc.get("learning_month") or "").strip() for doc in docs if doc.get("learning_month")})
        categories = sorted({(doc.get("category") or "").strip() for doc in docs if doc.get("category")})
        competencies = sorted({(doc.get("competency") or "").strip() for doc in docs if doc.get("competency")})
        return {
            "providers": providers,
            "designations": designations,
            "months": months,
            "categories": categories,
            "competencies": competencies,
        }

    async def get_course_by_uid(self, uid: str) -> dict | None:
        if not uid.startswith("learning_course:"):
            return None
        course_id = uid.split(":", 1)[1]
        if not ObjectId.is_valid(course_id):
            return None
        doc = await database.learning_courses.find_one({"_id": ObjectId(course_id)})
        if not doc:
            return None
        return _public_course(doc)

    async def create_course(self, current_user: CurrentUser, payload: ManagedLearningCourseCreateRequest) -> dict:
        data = self._normalize_request_payload(payload)
        doc = {
            **data,
            "archived": bool(data.get("archived")),
            "source_kind": "manual",
            "created_by_id": current_user.id,
            "updated_by_id": current_user.id,
        }
        if current_user.organization_id:
            doc["organization_id"] = current_user.organization_id
        doc["course_key"] = _course_key(
            provider=doc.get("provider") or "Managed Learning",
            designation=doc.get("designation") or "",
            learning_month=doc.get("learning_month") or "",
            category=doc.get("category") or "",
            competency=doc.get("competency") or "",
            title=doc.get("title") or "",
            url=doc.get("url") or None,
        )
        existing = await self._find_existing(doc["course_key"])
        if existing:
            raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Course already exists.")
        saved = await self._save_course_doc(doc, organization_id=current_user.organization_id)
        if current_user.organization_id:
            try:
                from app.services.course_sync_service import sync_to_framework
                await sync_to_framework(current_user.organization_id, saved)
            except Exception:
                pass
        return {"course": _public_course(saved)}

    async def update_course(self, current_user: CurrentUser, course_id: str, payload: ManagedLearningCourseUpdateRequest) -> dict:
        if not ObjectId.is_valid(course_id):
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Course not found.")
        doc = await database.learning_courses.find_one({"_id": ObjectId(course_id)})
        if not doc:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Course not found.")
        updates = self._normalize_request_payload(payload)
        updates["updated_by_id"] = current_user.id
        for key, value in list(updates.items()):
            if value is None:
                updates.pop(key)
        if updates:
            merged = {**doc, **updates}
            merged["_id"] = doc["_id"]
            merged["course_key"] = _course_key(
                provider=merged.get("provider") or "Managed Learning",
                designation=merged.get("designation") or "",
                learning_month=merged.get("learning_month") or "",
                category=merged.get("category") or "",
                competency=merged.get("competency") or "",
                title=merged.get("title") or "",
                url=merged.get("url") or None,
            )
            existing = await self._find_existing(merged["course_key"])
            if existing and str(existing["_id"]) != str(doc["_id"]):
                raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="A different course already uses the same roadmap slot.")
            saved = await self._save_course_doc(merged, organization_id=doc.get("organization_id"))
        else:
            saved = doc
        if doc.get("organization_id"):
            try:
                from app.services.course_sync_service import sync_to_framework
                await sync_to_framework(doc["organization_id"], saved)
            except Exception:
                pass
        return {"course": _public_course(saved)}

    async def archive_course(self, current_user: CurrentUser, course_id: str, *, archived: bool = True) -> dict:
        if not ObjectId.is_valid(course_id):
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Course not found.")
        result = await database.learning_courses.update_one(
            {"_id": ObjectId(course_id)},
            {"$set": {"archived": archived, "updated_at": _now(), "updated_by_id": current_user.id}},
        )
        if result.matched_count == 0:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Course not found.")
        doc = await database.learning_courses.find_one({"_id": ObjectId(course_id)})
        return {"course": _public_course(doc)}

    async def delete_course(self, current_user: CurrentUser, course_id: str) -> dict:
        if not ObjectId.is_valid(course_id):
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Course not found.")

        oid = ObjectId(course_id)
        await database.learning_enrollments.delete_many({"course_uid": f"learning_course:{course_id}"})
        await database.learning_assignments.delete_many({"course_uid": f"learning_course:{course_id}"})
        await database.learning_bookmarks.delete_many({"course_uid": f"learning_course:{course_id}"})

        result = await database.learning_courses.delete_one({"_id": oid})
        if result.deleted_count == 0:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Course not found.")
        if doc.get("organization_id"):
            try:
                from app.services.course_sync_service import sync_delete_from_framework
                await sync_delete_from_framework(doc["organization_id"], course_id)
            except Exception:
                pass
        return {"deleted": True}

    async def bulk_course_action(
        self,
        current_user: CurrentUser,
        course_ids: list[str],
        action: str,
    ) -> dict:
        """Bulk archive / restore / delete courses (Phase 3 management)."""
        valid_ids = [ObjectId(cid) for cid in course_ids if cid and ObjectId.is_valid(cid)]
        if not valid_ids:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST, detail="No valid course IDs provided."
            )
        query = {"_id": {"$in": valid_ids}}
        now = _now()

        if action == "archive":
            result = await database.learning_courses.update_many(
                query, {"$set": {"archived": True, "updated_at": now, "updated_by_id": current_user.id}}
            )
            return {"action": "archive", "affected": result.modified_count}
        if action == "restore":
            result = await database.learning_courses.update_many(
                query, {"$set": {"archived": False, "updated_at": now, "updated_by_id": current_user.id}}
            )
            return {"action": "restore", "affected": result.modified_count}
        if action == "delete":
            result = await database.learning_courses.delete_many(query)
            return {"action": "delete", "affected": result.deleted_count}
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Action must be 'archive', 'restore', or 'delete'.",
        )

    async def recommend_for_employee(self, employee: dict, *, limit: int = 8, organization_id: str | None = None) -> list[dict]:
        designation = (employee.get("job_title") or "").strip()
        if not designation:
            return []
        user_id = employee.get("user_id")
        completed = []
        assigned = []
        if user_id:
            completed_docs = await database.learning_enrollments.find({"user_id": user_id, "status": "completed"}, {"course_uid": 1}).to_list(length=1000)
            assigned_docs = await database.learning_assignments.find({"user_id": user_id}, {"course_uid": 1}).to_list(length=1000)
            completed = [d.get("course_uid") for d in completed_docs if d.get("course_uid")]
            assigned = [d.get("course_uid") for d in assigned_docs if d.get("course_uid")]
        query = self._collection_query()
        if organization_id:
            org_filter = {"$or": [{"organization_id": organization_id}, {"organization_id": {"$exists": False}}]}
            query = {"$and": [query, org_filter]}
        courses = await database.learning_courses.find(query).to_list(length=5000)
        active = [doc for doc in courses if not doc.get("archived")]
        same_designation = [doc for doc in active if _normalize(doc.get("designation")) == _normalize(designation)]
        if not same_designation:
            same_designation = [doc for doc in active if _normalize(designation) in _normalize(doc.get("designation")) or _normalize(doc.get("designation")) in _normalize(designation)]
        if not same_designation:
            same_designation = active

        def month_rank(value: Any) -> int:
            text = _clean(value)
            match = re.search(r"(\d+)", text)
            return int(match.group(1)) if match else 999

        completed_docs = [doc for doc in same_designation if _course_uid(doc) in set(completed)]
        highest_month = min((month_rank(doc.get("learning_month")) for doc in completed_docs), default=1)
        target_month = highest_month + 1 if completed_docs else 1
        target_month_courses = [doc for doc in same_designation if month_rank(doc.get("learning_month")) == target_month]
        if not target_month_courses:
            target_month_courses = same_designation

        def key(doc: dict) -> tuple[int, int, str]:
            uid = _course_uid(doc)
            if uid in completed:
                return (3, month_rank(doc.get("learning_month")), doc.get("title") or "")
            if uid in assigned:
                return (2, month_rank(doc.get("learning_month")), doc.get("title") or "")
            if month_rank(doc.get("learning_month")) == target_month:
                return (0, month_rank(doc.get("learning_month")), doc.get("title") or "")
            return (1, month_rank(doc.get("learning_month")), doc.get("title") or "")

        prioritized = sorted(target_month_courses, key=key)
        unseen = [doc for doc in prioritized if _course_uid(doc) not in set(completed) and _course_uid(doc) not in set(assigned)]
        random.seed(f"{employee.get('employee_id') or user_id or designation}-{target_month}")
        if len(unseen) > 3:
            head = unseen[:3]
            tail = unseen[3:]
            random.shuffle(tail)
            unseen = head + tail
        recommendations = []
        for doc in unseen[:limit]:
            public = _public_course(doc)
            public["reason"] = (
                f"Matches {designation} roadmap, {public.get('learning_month') or 'current month'}, {public.get('category') or 'assigned category'}."
            )
            public["priority"] = "medium"
            recommendations.append(public)
        return recommendations

    async def analytics(self) -> dict:
        courses = await database.learning_courses.find({}).to_list(length=5000)
        assignments = await database.learning_assignments.find({}).to_list(length=5000)
        enrollments = await database.learning_enrollments.find({}).to_list(length=5000)
        certificates = await database.learning_certificates.find({}).to_list(length=5000)

        active_courses = [c for c in courses if not c.get("archived")]
        completed_enrollments = [e for e in enrollments if e.get("status") == "completed"]
        verified_certs = [c for c in certificates if c.get("verification_status") == "verified"]
        completion_rate = round((len(completed_enrollments) / len(enrollments)) * 100, 1) if enrollments else 0.0
        learning_hours = round(sum((c.get("learning_hours") or 0) for c in verified_certs) + sum((e.get("duration_minutes") or 0) for e in completed_enrollments) / 60, 1)

        popular: dict[str, int] = defaultdict(int)
        for e in enrollments:
            if e.get("course_title"):
                popular[e["course_title"]] += 1
        popular_courses = sorted(popular.items(), key=lambda item: (-item[1], item[0].lower()))[:10]

        dept_progress: dict[str, dict[str, int]] = defaultdict(lambda: {"assigned": 0, "completed": 0})
        designation_progress: dict[str, dict[str, int]] = defaultdict(lambda: {"assigned": 0, "completed": 0})
        for assignment in assignments:
            dept = assignment.get("department") or "Unassigned"
            designation = assignment.get("job_title") or "Unassigned"
            dept_progress[dept]["assigned"] += 1
            designation_progress[designation]["assigned"] += 1
            if assignment.get("status") == "completed":
                dept_progress[dept]["completed"] += 1
                designation_progress[designation]["completed"] += 1

        monthly_trend: dict[str, dict[str, int]] = defaultdict(lambda: {"assigned": 0, "completed": 0, "certificates": 0})
        for assignment in assignments:
            key = _clean(assignment.get("created_at"))[:7] if assignment.get("created_at") else "unknown"
            monthly_trend[key]["assigned"] += 1
            if assignment.get("status") == "completed":
                monthly_trend[key]["completed"] += 1
        for cert in certificates:
            key = _clean(cert.get("created_at"))[:7] if cert.get("created_at") else "unknown"
            monthly_trend[key]["certificates"] += 1

        return {
            "total_courses": len(active_courses),
            "archived_courses": len(courses) - len(active_courses),
            "assigned_courses": len(assignments),
            "completed_courses": len(completed_enrollments),
            "pending_certificates": len([c for c in certificates if c.get("verification_status") == "pending"]),
            "learning_hours": learning_hours,
            "completion_rate": completion_rate,
            "most_popular_courses": [{"title": title, "enrollments": count} for title, count in popular_courses],
            "learning_trend": [
                {"month": month, **stats}
                for month, stats in sorted(monthly_trend.items(), key=lambda item: item[0])
            ],
            "department_progress": [
                {"department": department, **stats, "completion_rate": round((stats["completed"] / stats["assigned"]) * 100, 1) if stats["assigned"] else 0}
                for department, stats in sorted(dept_progress.items(), key=lambda item: item[0].lower())
            ],
            "designation_progress": [
                {"designation": designation, **stats, "completion_rate": round((stats["completed"] / stats["assigned"]) * 100, 1) if stats["assigned"] else 0}
                for designation, stats in sorted(designation_progress.items(), key=lambda item: item[0].lower())
            ],
        }

    async def notify_managers_of_import(self, current_user: CurrentUser, preview: dict) -> None:
        # Non-blocking helper for future workflow hooks.
        _ = current_user, preview
        return None


managed_learning_service = ManagedLearningService()
