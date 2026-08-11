"""Universal Provider Import Engine (Phase 2).

One common import engine for every learning provider:

  - Manual providers  -> Excel import (.xlsx / .xls / .csv)
  - API providers     -> API synchronization (same compare/insert/update/archive path)

Design goals:
  * Provider-agnostic — the engine never hardcodes a provider name.
  * Validate before import; never partially import an invalid file.
  * Upsert by identity priority: External Course ID -> Course URL -> Provider + Course Name.
  * Configurable missing-course action: keep / archive / delete.
  * Import history with rollback support.
  * Batch inserts for performance; detailed validation reports.
"""

from __future__ import annotations

import csv
import io
import re
from collections import defaultdict
from datetime import UTC, datetime
from typing import Any

from bson import ObjectId
from fastapi import HTTPException, UploadFile, status
from pymongo import InsertOne, UpdateOne
from pymongo.errors import BulkWriteError, DuplicateKeyError

from app.core.database import database
from app.core.rbac import CurrentUser
from app.schemas.import_engine import (
    IMPORT_MAX_ROWS,
    IMPORT_MAX_SIZE_BYTES,
    FileLevelIssue,
    ImportCourseCreateRequest,
    ImportCourseUpdateRequest,
    ImportIssue,
    ImportPreview,
    ImportPreviewRow,
)
from app.services.managed_learning_service import (
    COLUMN_ALIASES,
    _clean,
    _compact,
    _matches_alias,
    _normalize,
    _normalize_provider_name,
    _parse_duration,
    _slug,
)

# --------------------------------------------------------------------------- #
# Constants
# --------------------------------------------------------------------------- #
DEFAULT_MISSING_ACTION = "keep"

# Columns treated as required for a valid import (at minimum title).
MIN_REQUIRED_COLUMNS: tuple[str, ...] = ("title",)

# Columns that are strongly recommended for roadmap data.
RECOMMENDED_COLUMNS: tuple[str, ...] = (
    "designation",
    "learning_month",
    "category",
    "competency",
)

# Extra columns we can ingest (Phase 3: instructor, tags, difficulty, external id).
EXTRA_COLUMNS: tuple[str, ...] = ("instructor", "tags", "difficulty", "external_id", "level", "skills")

PREFERRED_SHEETS = (
    "linkedin",
    "learning",
    "roadmap",
    "road maps",
    "courses",
    "catalog",
    "curriculum",
    "training",
)

# Safety: block formula injection cells at rest in the database.
_FORMULA_START = ("=", "+", "-", "@", "\t", "\r")


def _now() -> datetime:
    return datetime.now(UTC)


def _iso(value: Any) -> Any:
    return value.isoformat() if hasattr(value, "isoformat") else value


def _looks_like_formula(value: Any) -> bool:
    text = _clean(value)
    if not text:
        return False
    return text.startswith(_FORMULA_START) and any(ch.isalpha() for ch in text[1:4])


def _sanitize_cell(value: Any) -> Any:
    """Strip formula-injection prefixes from text cells."""
    if not isinstance(value, str):
        return value
    text = value.strip()
    if text.startswith(_FORMULA_START):
        # Remove dangerous prefix, keep the rest as plain data.
        return text[1:].strip()
    return value


class ImportEngineService:
    """Provider-agnostic import/sync engine."""

    # ------------------------------------------------------------------ #
    # Provider resolution
    # ------------------------------------------------------------------ #
    async def resolve_provider(
        self, provider_id: str | None, provider_name: str | None, *, allow_inactive: bool = False
    ) -> dict | None:
        """Resolve a provider record by id, slug, or name. None when unknown."""
        provider = None
        if provider_id:
            if ObjectId.is_valid(provider_id):
                provider = await database.learning_providers.find_one({"_id": ObjectId(provider_id)})
            if not provider:
                provider = await database.learning_providers.find_one({"slug": provider_id})

        if not provider and provider_name:
            normalized = _normalize_provider_name(provider_name)
            provider = await database.learning_providers.find_one(
                {"name": {"$regex": f"^{re.escape(normalized)}$", "$options": "i"}}
            )
            if not provider:
                provider = await database.learning_providers.find_one({"slug": _slug(normalized)})

        if provider and not allow_inactive and not provider.get("active", True):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Provider '{provider.get('name')}' is inactive. Activate it before importing courses.",
            )
        return provider

    async def ensure_provider_for_name(self, provider_name: str | None, *, created_by_id: str | None = None) -> dict | None:
        """Get-or-create a provider record from a free-text provider name."""
        normalized = _normalize_provider_name(provider_name)
        if not normalized:
            return None
        slug = _slug(normalized)
        existing = await database.learning_providers.find_one({"slug": slug})
        if existing:
            return existing
        now = _now()
        doc = {
            "name": normalized,
            "slug": slug,
            "provider_type": "manual",
            "import_method": "excel",
            "description": None,
            "logo_url": None,
            "active": True,
            "created_at": now,
            "updated_at": now,
            "created_by_id": created_by_id,
            "created_by_name": None,
        }
        result = await database.learning_providers.insert_one(doc)
        doc["_id"] = result.inserted_id
        return doc

    # ------------------------------------------------------------------ #
    # File reading
    # ------------------------------------------------------------------ #
    async def _read_upload(self, file: UploadFile) -> tuple[list[list[Any]], str]:
        name = (file.filename or "").lower()
        raw = await file.read()

        if len(raw) > IMPORT_MAX_SIZE_BYTES:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"File is too large. Maximum size is {IMPORT_MAX_SIZE_BYTES // (1024 * 1024)} MB.",
            )

        if name.endswith(".csv"):
            try:
                text = raw.decode("utf-8-sig", errors="replace")
            except Exception as exc:  # noqa: BLE001
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Could not read the CSV file (encoding issue).") from exc
            rows = list(csv.reader(io.StringIO(text)))
            if not rows:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="The CSV file is empty.")
            return rows, "CSV"

        if not (name.endswith(".xlsx") or name.endswith(".xls") or name.endswith(".xlsm")):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Unsupported file format. Please upload a .xlsx or .xls spreadsheet (or .csv).",
            )

        try:
            from openpyxl import load_workbook

            workbook = load_workbook(io.BytesIO(raw), data_only=True, read_only=False)
        except Exception as exc:  # noqa: BLE001
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"The workbook appears to be corrupted or password-protected: {_clean(exc)[:200]}",
            ) from exc

        sheet = self._pick_sheet(workbook)
        rows = self._sheet_rows(sheet)
        if not rows:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="The spreadsheet is empty.")
        return rows, sheet.title or "Sheet1"

    def _pick_sheet(self, workbook) -> Any:
        by_name = {str(ws.title or "").strip().lower(): ws for ws in workbook.worksheets}
        for name in PREFERRED_SHEETS:
            if name in by_name:
                return by_name[name]
        for ws in workbook.worksheets:
            title = str(ws.title or "").strip().lower()
            if any(token in title for token in PREFERRED_SHEETS):
                return ws
        return workbook.active

    def _sheet_rows(self, sheet) -> list[list[Any]]:
        rows = list(sheet.iter_rows(values_only=False))
        merged_map: dict[tuple[int, int], Any] = {}
        for merged in getattr(sheet, "merged_cells", []).ranges:
            min_col, min_row, max_col, max_row = merged.bounds
            anchor = sheet.cell(min_row, min_col).value
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    merged_map[(row, col)] = anchor
        out: list[list[Any]] = []
        for row_idx, cells in enumerate(rows, start=1):
            values: list[Any] = []
            for col_idx, cell in enumerate(cells, start=1):
                value = cell.value
                if value in (None, ""):
                    value = merged_map.get((row_idx, col_idx), value)
                values.append(_sanitize_cell(value))
            out.append(values)
        return out

    # ------------------------------------------------------------------ #
    # Header resolution + file-level validation
    # ------------------------------------------------------------------ #
    def _find_header_row(self, rows: list[list[Any]]) -> tuple[int, list[FileLevelIssue]]:
        best_index = 0
        best_score = -1
        for idx, row in enumerate(rows[:60]):
            normalized = [_normalize(cell) for cell in row]
            score = sum(1 for aliases in COLUMN_ALIASES.values() if any(_matches_alias(cell, aliases) for cell in normalized))
            score += sum(1 for cell in normalized if cell)
            if score > best_score:
                best_score = score
                best_index = idx

        issues: list[FileLevelIssue] = []
        if best_score <= 0:
            issues.append(FileLevelIssue(code="no_header", message="No recognizable header row was found. Expected columns such as Course Title, URL, Designation, Learning Month, Category, Competency."))
        return best_index, issues

    def _resolve_columns(self, header: list[Any]) -> tuple[dict[str, int | None], str]:
        normalized = [_normalize(cell) for cell in header]

        # Include extra columns in resolution.
        aliases = dict(COLUMN_ALIASES)
        aliases["instructor"] = ("instructor", "trainer", "teacher", "facilitator", "author")
        aliases["tags"] = ("tags", "tag", "keywords", "skills tags")
        aliases["difficulty"] = ("difficulty", "level", "difficulty level", "experience level")
        aliases["external_id"] = ("course id", "course_id", "external id", "external_id", "id", "course code", "code")
        aliases["skills"] = ("skills", "skill", "skills covered", "learning outcomes")

        indexes: dict[str, int | None] = {key: None for key in aliases}
        claimed: set[int] = set()

        for key, key_aliases in aliases.items():
            for idx, cell in enumerate(normalized):
                if idx in claimed:
                    continue
                if cell in key_aliases:
                    indexes[key] = idx
                    claimed.add(idx)
                    break

        for key, key_aliases in aliases.items():
            if indexes[key] is not None:
                continue
            for idx, cell in enumerate(normalized):
                if idx in claimed:
                    continue
                if len(cell) >= 4 and any(len(a) >= 4 and a in cell for a in key_aliases):
                    indexes[key] = idx
                    claimed.add(idx)
                    break

        duration_unit = "minutes"
        duration_idx = indexes.get("duration_minutes")
        if duration_idx is not None:
            header_cell = normalized[duration_idx]
            if header_cell == "total" or "hour" in header_cell:
                duration_unit = "hours"
        return indexes, duration_unit

    def _validate_headers(
        self,
        header: list[Any],
        indexes: dict[str, int | None],
    ) -> list[FileLevelIssue]:
        issues: list[FileLevelIssue] = []

        # Missing required columns
        missing_required = [col for col in MIN_REQUIRED_COLUMNS if indexes.get(col) is None]
        if missing_required:
            issues.append(
                FileLevelIssue(
                    code="missing_required_columns",
                    message=f"Missing required column(s): {', '.join(missing_required)}.",
                )
            )
        missing_recommended = [col for col in RECOMMENDED_COLUMNS if indexes.get(col) is None]
        if missing_recommended:
            issues.append(
                FileLevelIssue(
                    code="missing_recommended_columns",
                    message=f"Missing recommended column(s): {', '.join(missing_recommended)}. Roadmap grouping will be limited.",
                )
            )

        # Duplicate headers
        normalized = [_normalize(cell) for cell in header]
        seen: dict[str, int] = {}
        dupes: list[str] = []
        for cell in normalized:
            if not cell:
                continue
            seen[cell] = seen.get(cell, 0) + 1
            if seen[cell] == 2:
                dupes.append(cell)
        if dupes:
            issues.append(
                FileLevelIssue(
                    code="duplicate_headers",
                    message=f"Duplicate column header(s): {', '.join(dupes)}. Please make headers unique.",
                )
            )

        return issues

    # ------------------------------------------------------------------ #
    # Row-level validation
    # ------------------------------------------------------------------ #
    def _validate_row(
        self,
        row_number: int,
        row: list[Any],
        indexes: dict[str, int | None],
        duration_unit: str,
        provider_name: str | None,
        seen: dict[str, list[int]],
    ) -> dict[str, Any] | None:
        """Validate a single row, returning a normalized record or None (blank)."""
        issues: list[ImportIssue] = []

        def get(key: str) -> str:
            idx = indexes.get(key)
            if idx is None or idx >= len(row):
                return ""
            return _clean(row[idx])

        title = get("title")
        url = get("url")
        designation = get("designation")
        learning_month = get("learning_month")
        category = get("category")
        competency = get("competency")
        external_id = get("external_id")
        instructor = get("instructor")
        tags_raw = get("tags")
        difficulty = get("difficulty")
        skills_raw = get("skills")
        description = get("description")

        if not any([title, url, designation, learning_month, category, competency, external_id, description]):
            return None  # blank row -> skip silently

        # Required fields — catalog import only needs a title. Roadmap placement
        # (designation / month / category / competency) is optional here and can
        # be filled later in Build roadmap.
        if not title:
            issues.append(ImportIssue(row=row_number, field="title", message="Course title is required."))

        # URL validation
        if url and not re.match(r"^https?://", url, re.I):
            issues.append(ImportIssue(row=row_number, field="url", message="Course URL is invalid (must start with http:// or https://)."))

        # Duration validation
        duration_minutes: int | None = None
        if indexes.get("duration_minutes") is not None and row[indexes["duration_minutes"]] not in (None, ""):
            duration_minutes = _parse_duration(row[indexes["duration_minutes"]], duration_unit)
            if duration_minutes is None:
                issues.append(ImportIssue(row=row_number, field="duration_minutes", message="Duration is invalid."))
            elif duration_minutes <= 0:
                issues.append(ImportIssue(row=row_number, field="duration_minutes", message="Duration must be positive."))

        # Formula injection protection
        for field_name in ("title", "designation", "learning_month", "category", "competency", "description", "instructor"):
            raw = row[indexes[field_name]] if indexes.get(field_name) is not None and indexes[field_name] < len(row) else None
            if raw is not None and _looks_like_formula(raw):
                issues.append(ImportIssue(row=row_number, field=field_name, message="Formula-like content detected and was neutralized."))

        # Duplicate detection within the file
        identity_keys: dict[str, str] = {}
        if external_id:
            identity_keys["external_id"] = f"ext:{external_id.lower()}"
        if url:
            identity_keys["url"] = f"url:{url.lower()}"
        if provider_name and title:
            identity_keys["name"] = f"name:{_normalize(provider_name)}|{_normalize(title)}"

        for key_name, key in identity_keys.items():
            if key in seen:
                seen[key].append(row_number)
                issues.append(ImportIssue(row=row_number, field=key_name, message=f"Duplicate course {key_name.replace('_', ' ')} in the same file (first seen on row {seen[key][0]})."))
            else:
                seen[key] = [row_number]

        tags = [t.strip() for t in tags_raw.split(",") if t and t.strip()] if tags_raw else []
        skills = [s.strip() for s in skills_raw.split(",") if s and s.strip()] if skills_raw else []

        return {
            "row": row_number,
            "external_id": external_id or None,
            "title": title,
            "url": url or None,
            "provider": provider_name,
            "designation": designation or "",
            "learning_month": learning_month or "",
            "category": category or "",
            "competency": competency or "",
            "description": description or None,
            "duration_minutes": duration_minutes,
            "instructor": instructor or None,
            "tags": tags,
            "skills": skills,
            "difficulty": difficulty or None,
            "issues": issues,
        }

    # ------------------------------------------------------------------ #
    # Upsert identity matching
    # ------------------------------------------------------------------ #
    async def _match_existing(
        self, record: dict, provider: dict | None
    ) -> tuple[dict | None, str | None]:
        """Find an existing course by identity priority.

        Priority: External Course ID -> Course URL -> Provider + Course Name
        -> course_key (covers legacy docs without provider_id that still share
        the unique course_key index).
        Returns (doc, match_method).
        """
        provider_id = str(provider["_id"]) if provider else None

        if record.get("external_id"):
            doc = await database.learning_courses.find_one(
                {"external_id": record["external_id"], "provider_id": provider_id}
            )
            if doc:
                return doc, "external_id"

        if record.get("url"):
            url_key = _normalize(record["url"])
            docs = await database.learning_courses.find(
                {"provider_id": provider_id, "url": {"$ne": None}}
            ).to_list(length=200)
            for doc in docs:
                if _normalize(doc.get("url")) == url_key:
                    return doc, "url"

        if provider and record.get("title"):
            title_key = _normalize(record["title"])
            name_key = _normalize(provider.get("name"))
            docs = await database.learning_courses.find(
                {"provider_id": provider_id}
            ).to_list(length=2000)
            for doc in docs:
                if _normalize(doc.get("title")) == title_key and _normalize(doc.get("provider")) == name_key:
                    return doc, "provider_name"

        # Fallback: unique course_key. Legacy Managed Learning rows often have
        # no provider_id, so the scoped matches above miss them and a naive
        # insert then fails on the unique course_key index. Matching here lets
        # Excel/API re-imports merge as "updated" instead.
        from app.services.managed_learning_service import _course_key

        provider_name = (
            (provider.get("name") if provider else None)
            or record.get("provider")
            or "Managed Learning"
        )
        course_key = record.get("course_key") or _course_key(
            provider=provider_name,
            designation=record.get("designation") or "",
            learning_month=record.get("learning_month") or "",
            category=record.get("category") or "",
            competency=record.get("competency") or "",
            title=record.get("title") or "",
            url=record.get("url") or None,
        )
        if course_key:
            doc = await database.learning_courses.find_one({"course_key": course_key})
            if doc:
                return doc, "course_key"

        return None, None

    # ------------------------------------------------------------------ #
    # Preview
    # ------------------------------------------------------------------ #
    async def preview_import(
        self,
        current_user: CurrentUser,
        file: UploadFile,
        *,
        provider_id: str | None = None,
        provider_name: str | None = None,
    ) -> dict:
        """Validate the upload and produce a full preview. Never writes data."""
        provider = await self.resolve_provider(provider_id, provider_name, allow_inactive=True)
        resolved_provider_name = provider.get("name") if provider else _normalize_provider_name(provider_name)

        rows, sheet_name = await self._read_upload(file)
        if len(rows) - 1 > IMPORT_MAX_ROWS:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Too many rows. Maximum is {IMPORT_MAX_ROWS}.",
            )

        header_index, header_issues = self._find_header_row(rows)
        header = rows[header_index] if rows else []
        indexes, duration_unit = self._resolve_columns(header)
        header_issues += self._validate_headers(header, indexes)

        invalid_file = any(issue.code in ("no_header", "missing_required_columns") for issue in header_issues)

        seen: dict[str, list[int]] = {}
        preview_rows: list[ImportPreviewRow] = []
        counts = {"new": 0, "updated": 0, "duplicate": 0, "invalid": 0, "skipped": 0}

        for row_number, row in enumerate(rows[header_index + 1:], start=header_index + 2):
            record = self._validate_row(row_number, row, indexes, duration_unit, resolved_provider_name, seen)
            if record is None:
                continue  # blank row

            errors = [i for i in record["issues"] if i.severity == "error"]
            duplicate_in_file = any("Duplicate course" in i.message for i in errors)

            # Match existing only when the row itself is otherwise valid.
            existing = None
            match_method = None
            if not errors and not invalid_file:
                existing, match_method = await self._match_existing(record, provider)

            if errors and not duplicate_in_file:
                status_value = "invalid"
                counts["invalid"] += 1
            elif duplicate_in_file:
                status_value = "duplicate"
                counts["duplicate"] += 1
            elif existing:
                status_value = "updated"
                counts["updated"] += 1
            else:
                status_value = "new"
                counts["new"] += 1

            preview_rows.append(
                ImportPreviewRow(
                    row=record["row"],
                    status=status_value,
                    external_id=record.get("external_id"),
                    title=record.get("title"),
                    url=record.get("url"),
                    provider=record.get("provider"),
                    designation=record.get("designation"),
                    learning_month=record.get("learning_month"),
                    category=record.get("category"),
                    competency=record.get("competency"),
                    duration_minutes=record.get("duration_minutes"),
                    description=record.get("description"),
                    instructor=record.get("instructor"),
                    tags=record.get("tags") or [],
                    issues=record["issues"],
                    existing_id=str(existing["_id"]) if existing else None,
                    match_method=match_method,
                )
            )

            if len(preview_rows) >= IMPORT_MAX_ROWS:
                break

        valid = not invalid_file and counts["invalid"] == 0
        preview = ImportPreview(
            provider_id=str(provider["_id"]) if provider else None,
            provider_name=resolved_provider_name,
            filename=file.filename,
            sheet_name=sheet_name,
            total_rows=len(preview_rows),
            new_courses=counts["new"],
            updated_courses=counts["updated"],
            duplicate_rows=counts["duplicate"],
            invalid_rows=counts["invalid"],
            skipped_rows=counts["skipped"],
            rows=preview_rows,
            file_issues=header_issues,
            valid=valid,
        )
        return preview.model_dump()

    # ------------------------------------------------------------------ #
    # Commit
    # ------------------------------------------------------------------ #
    async def commit_import(
        self,
        current_user: CurrentUser,
        file: UploadFile,
        *,
        provider_id: str | None = None,
        provider_name: str | None = None,
        missing_action: str = DEFAULT_MISSING_ACTION,
    ) -> dict:
        """Validate + import in one shot (the existing /managed flow), recording history."""
        preview_data = await self.preview_import(current_user, file, provider_id=provider_id, provider_name=provider_name)
        return await self.commit_preview(current_user, preview_data, file=file, missing_action=missing_action)

    async def commit_preview(
        self,
        current_user: CurrentUser,
        preview_data: dict,
        *,
        file: UploadFile | None = None,
        missing_action: str = DEFAULT_MISSING_ACTION,
    ) -> dict:
        """Commit a previously-previewed import. Blocks when the file is invalid."""
        preview = ImportPreview(**preview_data)
        if not preview.valid:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail={
                    "message": "The file has validation errors and was not imported.",
                    "validation": preview_data,
                },
            )

        provider = None
        if preview.provider_id and ObjectId.is_valid(preview.provider_id):
            provider = await database.learning_providers.find_one({"_id": ObjectId(preview.provider_id)})
        if provider is None and preview.provider_name:
            provider = await self.resolve_provider(None, preview.provider_name, allow_inactive=True)
        if provider is not None and not provider.get("active", True):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Provider '{provider.get('name')}' is inactive. Activate it before importing courses.",
            )

        now = _now()
        created_ids: list[str] = []
        updated_before: list[dict] = []
        imported = 0
        updated = 0
        skipped = 0
        failed = 0
        errors: list[dict] = []
        imported_keys: set[str] = set()

        valid_rows = [r for r in preview.rows if r.status in ("new", "updated")]

        for row in valid_rows:
            record = {
                "external_id": row.external_id,
                "title": row.title,
                "url": row.url,
                "provider": provider.get("name") if provider else row.provider or "Managed Learning",
                "designation": row.designation or "",
                "learning_month": row.learning_month or "",
                "category": row.category or "",
                "competency": row.competency or "",
                "description": row.description,
                "duration_minutes": row.duration_minutes,
                "instructor": row.instructor,
                "tags": row.tags or [],
                "provider_id": str(provider["_id"]) if provider else None,
            }
            if record.get("url"):
                imported_keys.add(_normalize(record["url"]))

            try:
                if row.existing_id and ObjectId.is_valid(row.existing_id):
                    # Update path
                    existing = await database.learning_courses.find_one({"_id": ObjectId(row.existing_id)})
                    if not existing:
                        skipped += 1
                        continue
                    snapshot = {k: existing.get(k) for k in (
                        "title", "url", "designation", "learning_month", "category",
                        "competency", "description", "duration_minutes", "provider",
                        "provider_id", "external_id", "archived", "instructor", "tags",
                    )}
                    updated_before.append({"_id": str(existing["_id"]), "snapshot": snapshot})
                    updated_doc = {**existing, **record, "updated_at": now, "updated_by_id": current_user.id}
                    updated_doc = await self._finalize_course_doc(updated_doc)
                    set_doc = {k: v for k, v in updated_doc.items() if k != "_id"}
                    update_ops: dict[str, Any] = {"$set": set_doc}
                    if "external_id" not in set_doc:
                        update_ops["$unset"] = {"external_id": ""}
                    await database.learning_courses.update_one({"_id": existing["_id"]}, update_ops)
                    updated += 1
                else:
                    # Insert path
                    new_doc = {
                        **record,
                        "archived": False,
                        "source_kind": "import",
                        "source_filename": file.filename if file else preview.filename,
                        "source_row": row.row,
                        "created_by_id": current_user.id,
                        "updated_by_id": current_user.id,
                        "created_at": now,
                        "updated_at": now,
                    }
                    new_doc = await self._finalize_course_doc(new_doc)
                    result = await database.learning_courses.insert_one(new_doc)
                    created_ids.append(str(result.inserted_id))
                    imported += 1
            except Exception as exc:  # noqa: BLE001
                failed += 1
                errors.append({"row": row.row, "error": _clean(exc)[:300]})

        # Missing-course handling (courses for this provider not in the file)
        archived_ids: list[str] = []
        deleted_ids: list[dict] = []
        if provider is not None and missing_action in ("archive", "delete"):
            provider_query = {"provider_id": str(provider["_id"])}
            provider_courses = await database.learning_courses.find(provider_query).to_list(length=5000)
            for course in provider_courses:
                if course.get("archived"):
                    continue
                course_url = _normalize(course.get("url")) if course.get("url") else None
                in_file = (course_url and course_url in imported_keys) if imported_keys else False
                # Also match by title when no url
                if not in_file and course.get("title"):
                    in_file = any(
                        _normalize(row.title) == _normalize(course.get("title"))
                        for row in valid_rows if row.title
                    )
                if in_file:
                    continue
                if missing_action == "archive":
                    await database.learning_courses.update_one(
                        {"_id": course["_id"]},
                        {"$set": {"archived": True, "updated_at": now, "updated_by_id": current_user.id}},
                    )
                    archived_ids.append(str(course["_id"]))
                elif missing_action == "delete":
                    await database.learning_courses.delete_one({"_id": course["_id"]})
                    deleted_ids.append({"_id": str(course["_id"]), "doc": course})

        # Record history
        history_id = await self._record_history(
            provider=provider,
            current_user=current_user,
            import_type="excel",
            filename=file.filename if file else preview.filename,
            rows_total=len(preview.rows),
            rows_imported=imported,
            rows_updated=updated,
            rows_failed=failed,
            rows_archived=len(archived_ids),
            rows_deleted=len(deleted_ids),
            validation_summary={
                "new_courses": preview.new_courses,
                "updated_courses": preview.updated_courses,
                "duplicate_rows": preview.duplicate_rows,
                "invalid_rows": preview.invalid_rows,
            },
            message=(
                f"Imported {imported} new course(s), updated {updated}, "
                f"archived {len(archived_ids)}, deleted {len(deleted_ids)}, skipped {skipped}."
            ),
            created_ids=created_ids,
            updated_before=updated_before,
            archived_ids=archived_ids,
            deleted_ids=deleted_ids,
        )

        await self._notify_import_result(
            provider=provider,
            imported=imported,
            updated=updated,
            failed=failed,
            archived=len(archived_ids),
            deleted=len(deleted_ids),
        )

        return {
            "message": f"Imported {imported} new course(s), updated {updated}, archived {len(archived_ids)}, deleted {len(deleted_ids)}, skipped {skipped}.",
            "imported": imported,
            "updated": updated,
            "skipped": skipped,
            "archived": len(archived_ids),
            "deleted": len(deleted_ids),
            "errors": errors,
            "history_id": history_id,
            "preview": preview_data,
        }

    async def _finalize_course_doc(self, doc: dict) -> dict:
        """Compute course_key / hierarchy_path so records remain compatible with the
        managed-learning catalog and the unique course_key index."""
        from app.services.managed_learning_service import _course_key, _hierarchy_key

        provider_name = doc.get("provider") or "Managed Learning"
        doc["provider"] = _normalize_provider_name(provider_name) or "Managed Learning"
        doc["course_key"] = _course_key(
            provider=doc["provider"],
            designation=doc.get("designation") or "",
            learning_month=doc.get("learning_month") or "",
            category=doc.get("category") or "",
            competency=doc.get("competency") or "",
            title=doc.get("title") or "",
            url=doc.get("url") or None,
        )
        doc["hierarchy_key"] = _hierarchy_key(
            designation=doc.get("designation") or "",
            learning_month=doc.get("learning_month") or "",
            category=doc.get("category") or "",
            competency=doc.get("competency") or "",
        )
        doc["hierarchy_path"] = [
            part
            for part in [doc.get("designation"), doc.get("learning_month"), doc.get("category"), doc.get("competency")]
            if part
        ]
        # Never persist null/empty external_id — the unique (external_id, provider_id)
        # index must only cover real ids.
        if not doc.get("external_id"):
            doc.pop("external_id", None)
        return doc

    def _public_course(self, doc: dict) -> dict:
        """Serialize a course for Import Engine single-course responses."""
        return {
            "id": str(doc["_id"]),
            "uid": f"learning_course:{doc['_id']}",
            "provider_id": doc.get("provider_id"),
            "provider": doc.get("provider") or "Managed Learning",
            "external_id": doc.get("external_id"),
            "title": doc.get("title") or "",
            "url": doc.get("url") or "",
            "designation": doc.get("designation") or "",
            "learning_month": doc.get("learning_month") or "",
            "category": doc.get("category") or "",
            "competency": doc.get("competency") or "",
            "description": doc.get("description") or "",
            "duration_minutes": doc.get("duration_minutes"),
            "instructor": doc.get("instructor"),
            "tags": list(doc.get("tags") or []),
            "archived": bool(doc.get("archived")),
            "source_kind": doc.get("source_kind") or "manual",
            "course_key": doc.get("course_key"),
            "hierarchy_key": doc.get("hierarchy_key"),
            "hierarchy_path": doc.get("hierarchy_path") or [],
            "created_at": doc.get("created_at"),
            "updated_at": doc.get("updated_at"),
        }

    async def _course_key_conflict(self, course_key: str, *, exclude_id: ObjectId | None = None) -> dict | None:
        query: dict[str, Any] = {"course_key": course_key}
        if exclude_id is not None:
            query["_id"] = {"$ne": exclude_id}
        return await database.learning_courses.find_one(query)

    async def create_course(self, current_user: CurrentUser, payload: ImportCourseCreateRequest) -> dict:
        """Create one course via the Import Engine domain (always sets provider_id)."""
        provider = await self.resolve_provider(payload.provider_id, None, allow_inactive=False)
        if not provider:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Provider not found.")

        now = _now()
        record = {
            "external_id": payload.external_id,
            "title": payload.title,
            "url": payload.url,
            "provider": provider.get("name"),
            "designation": payload.designation or "",
            "learning_month": payload.learning_month or "",
            "category": payload.category or "",
            "competency": payload.competency or "",
            "description": payload.description,
            "duration_minutes": payload.duration_minutes,
            "instructor": payload.instructor,
            "tags": list(payload.tags or []),
            "provider_id": str(provider["_id"]),
        }

        existing, match_method = await self._match_existing(record, provider)
        if existing:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=f"Course already exists (matched by {match_method}).",
            )

        doc = {
            **record,
            "archived": bool(payload.archived),
            "source_kind": "manual",
            "created_by_id": current_user.id,
            "updated_by_id": current_user.id,
            "created_at": now,
            "updated_at": now,
        }
        if current_user.organization_id:
            doc["organization_id"] = current_user.organization_id

        doc = await self._finalize_course_doc(doc)
        conflict = await self._course_key_conflict(doc["course_key"])
        if conflict:
            raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Course already exists.")

        try:
            result = await database.learning_courses.insert_one(doc)
        except DuplicateKeyError as exc:
            raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Course already exists.") from exc

        doc["_id"] = result.inserted_id

        if current_user.organization_id:
            try:
                from app.services.course_sync_service import sync_to_framework

                await sync_to_framework(current_user.organization_id, doc)
            except Exception:
                pass

        return {"course": self._public_course(doc)}

    async def update_course(
        self, current_user: CurrentUser, course_id: str, payload: ImportCourseUpdateRequest
    ) -> dict:
        """Update one course via the Import Engine domain (keeps/sets provider_id)."""
        if not ObjectId.is_valid(course_id):
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Course not found.")

        existing = await database.learning_courses.find_one({"_id": ObjectId(course_id)})
        if not existing:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Course not found.")

        updates = payload.model_dump(exclude_unset=True)
        provider_id = updates.pop("provider_id", None) or existing.get("provider_id")

        provider = None
        if provider_id:
            provider = await self.resolve_provider(provider_id, None, allow_inactive=False)
            if not provider:
                raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Provider not found.")

        if not provider and existing.get("provider"):
            provider = await self.resolve_provider(None, existing.get("provider"), allow_inactive=False)

        if not provider:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="A valid provider_id is required so the course stays visible to the Import Engine.",
            )

        merged = {**existing, **updates}
        merged["_id"] = existing["_id"]
        merged["provider"] = provider.get("name")
        merged["provider_id"] = str(provider["_id"])
        merged["updated_by_id"] = current_user.id
        merged["updated_at"] = _now()

        merged = await self._finalize_course_doc(merged)

        conflict = await self._course_key_conflict(merged["course_key"], exclude_id=existing["_id"])
        if conflict:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="A different course already uses the same roadmap slot.",
            )

        # Identity collision against another course for this provider.
        match_doc, match_method = await self._match_existing(merged, provider)
        if match_doc and str(match_doc["_id"]) != str(existing["_id"]):
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=f"Course already exists (matched by {match_method}).",
            )

        set_doc = {k: v for k, v in merged.items() if k != "_id"}
        update_ops: dict[str, Any] = {"$set": set_doc}
        if "external_id" not in set_doc:
            update_ops["$unset"] = {"external_id": ""}
        try:
            await database.learning_courses.update_one({"_id": existing["_id"]}, update_ops)
        except DuplicateKeyError as exc:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="A different course already uses the same roadmap slot.",
            ) from exc

        saved = await database.learning_courses.find_one({"_id": existing["_id"]})

        org_id = existing.get("organization_id") or current_user.organization_id
        if org_id:
            try:
                from app.services.course_sync_service import sync_to_framework

                await sync_to_framework(org_id, saved)
            except Exception:
                pass

        return {"course": self._public_course(saved)}

    # ------------------------------------------------------------------ #
    # History + rollback
    # ------------------------------------------------------------------ #
    async def _record_history(
        self,
        *,
        provider: dict | None,
        current_user: CurrentUser,
        import_type: str,
        filename: str | None,
        rows_total: int,
        rows_imported: int,
        rows_updated: int,
        rows_failed: int,
        rows_archived: int,
        rows_deleted: int,
        validation_summary: dict,
        message: str,
        created_ids: list[str],
        updated_before: list[dict],
        archived_ids: list[str],
        deleted_ids: list[dict],
    ) -> str:
        now = _now()
        doc = {
            "provider_id": str(provider["_id"]) if provider else None,
            "provider_name": provider.get("name") if provider else None,
            "imported_by_id": current_user.id,
            "imported_by_name": current_user.full_name or current_user.email,
            "import_type": import_type,
            "status": "completed",
            "filename": filename,
            "rows_total": rows_total,
            "rows_imported": rows_imported,
            "rows_updated": rows_updated,
            "rows_failed": rows_failed,
            "rows_archived": rows_archived,
            "rows_deleted": rows_deleted,
            "validation_summary": validation_summary,
            "message": message,
            "created_ids": created_ids,
            "updated_before": updated_before,
            "archived_ids": archived_ids,
            "deleted_ids": deleted_ids,
            "created_at": now,
            "rollback_at": None,
        }
        result = await database.learning_import_history.insert_one(doc)
        return str(result.inserted_id)

    async def list_import_history(
        self,
        current_user: CurrentUser,
        *,
        provider_id: str | None = None,
        import_type: str | None = None,
        page: int = 1,
        page_size: int = 20,
    ) -> dict:
        query: dict[str, Any] = {}
        if provider_id:
            query["provider_id"] = provider_id
        if import_type:
            query["import_type"] = import_type

        total = await database.learning_import_history.count_documents(query)
        docs = await database.learning_import_history.find(query).sort("created_at", -1).skip((page - 1) * page_size).limit(page_size).to_list(length=page_size)

        entries = []
        for d in docs:
            entries.append({
                "id": str(d["_id"]),
                "provider_id": d.get("provider_id"),
                "provider_name": d.get("provider_name"),
                "imported_by_id": d.get("imported_by_id"),
                "imported_by_name": d.get("imported_by_name"),
                "import_type": d.get("import_type"),
                "status": d.get("status"),
                "filename": d.get("filename"),
                "rows_total": d.get("rows_total", 0),
                "rows_imported": d.get("rows_imported", 0),
                "rows_updated": d.get("rows_updated", 0),
                "rows_failed": d.get("rows_failed", 0),
                "rows_archived": d.get("rows_archived", 0),
                "rows_deleted": d.get("rows_deleted", 0),
                "validation_summary": d.get("validation_summary") or {},
                "message": d.get("message"),
                "created_at": _iso(d.get("created_at")),
                "rollback_at": _iso(d.get("rollback_at")),
            })
        pages = max(1, (total + page_size - 1) // page_size) if total else 1
        return {"history": entries, "total": total, "page": page, "page_size": page_size, "pages": pages}

    async def get_import_history(self, history_id: str) -> dict:
        if not ObjectId.is_valid(history_id):
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Import history not found.")
        doc = await database.learning_import_history.find_one({"_id": ObjectId(history_id)})
        if not doc:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Import history not found.")
        return {
            "id": str(doc["_id"]),
            "provider_id": doc.get("provider_id"),
            "provider_name": doc.get("provider_name"),
            "imported_by_name": doc.get("imported_by_name"),
            "import_type": doc.get("import_type"),
            "status": doc.get("status"),
            "filename": doc.get("filename"),
            "rows_total": doc.get("rows_total", 0),
            "rows_imported": doc.get("rows_imported", 0),
            "rows_updated": doc.get("rows_updated", 0),
            "rows_failed": doc.get("rows_failed", 0),
            "rows_archived": doc.get("rows_archived", 0),
            "rows_deleted": doc.get("rows_deleted", 0),
            "validation_summary": doc.get("validation_summary") or {},
            "message": doc.get("message"),
            "created_at": _iso(doc.get("created_at")),
            "rollback_at": _iso(doc.get("rollback_at")),
        }

    async def import_report_csv(self, history_id: str) -> str:
        """CSV validation report for a completed import."""
        if not ObjectId.is_valid(history_id):
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Import history not found.")
        doc = await database.learning_import_history.find_one({"_id": ObjectId(history_id)})
        if not doc:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Import history not found.")

        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(["Import Report", doc.get("filename") or "", str(doc.get("created_at") or "")])
        writer.writerow(["Provider", doc.get("provider_name") or ""])
        writer.writerow(["Imported By", doc.get("imported_by_name") or ""])
        writer.writerow([])
        writer.writerow(["New Courses", doc.get("rows_imported", 0)])
        writer.writerow(["Updated Courses", doc.get("rows_updated", 0)])
        writer.writerow(["Archived Courses", doc.get("rows_archived", 0)])
        writer.writerow(["Deleted Courses", doc.get("rows_deleted", 0)])
        writer.writerow(["Failed Rows", doc.get("rows_failed", 0)])
        writer.writerow([])
        summary = doc.get("validation_summary") or {}
        writer.writerow(["Validation Summary"])
        for key, value in summary.items():
            writer.writerow([key.replace("_", " ").title(), value])
        return buffer.getvalue()

    async def rollback_import(self, current_user: CurrentUser, history_id: str) -> dict:
        """Roll back a completed import: undo creates, updates, archives, deletes."""
        if not ObjectId.is_valid(history_id):
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Import history not found.")
        doc = await database.learning_import_history.find_one({"_id": ObjectId(history_id)})
        if not doc:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Import history not found.")
        if doc.get("status") != "completed":
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Only completed imports can be rolled back.")
        if doc.get("rollback_at"):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="This import has already been rolled back.")

        now = _now()

        # 1. Delete created courses
        created_ids = [ObjectId(cid) for cid in (doc.get("created_ids") or []) if ObjectId.is_valid(cid)]
        if created_ids:
            await database.learning_courses.delete_many({"_id": {"$in": created_ids}})

        # 2. Restore updated courses from snapshots
        for entry in doc.get("updated_before") or []:
            course_id = entry.get("_id")
            snapshot = entry.get("snapshot") or {}
            if course_id and ObjectId.is_valid(course_id):
                await database.learning_courses.update_one(
                    {"_id": ObjectId(course_id)},
                    {"$set": {**snapshot, "updated_at": now}},
                )

        # 3. Restore archived courses
        archived_ids = [ObjectId(cid) for cid in (doc.get("archived_ids") or []) if ObjectId.is_valid(cid)]
        if archived_ids:
            await database.learning_courses.update_many(
                {"_id": {"$in": archived_ids}},
                {"$set": {"archived": False, "updated_at": now}},
            )

        # 4. Re-insert deleted courses
        restored_deleted = 0
        for entry in doc.get("deleted_ids") or []:
            full_doc = entry.get("doc")
            if not full_doc:
                continue
            try:
                await database.learning_courses.insert_one(full_doc)
                restored_deleted += 1
            except Exception:  # noqa: BLE001
                pass

        await database.learning_import_history.update_one(
            {"_id": doc["_id"]},
            {"$set": {"status": "rolled_back", "rollback_at": now}},
        )

        return {
            "rolled_back": True,
            "courses_deleted": len(created_ids),
            "courses_restored": len(updated_before),
            "courses_unarchived": len(archived_ids),
            "courses_reinserted": restored_deleted,
            "history_id": history_id,
        }

    # ------------------------------------------------------------------ #
    # Notifications
    # ------------------------------------------------------------------ #
    async def _notify_import_result(
        self,
        *,
        provider: dict | None,
        imported: int,
        updated: int,
        failed: int,
        archived: int,
        deleted: int,
    ) -> None:
        try:
            from app.services.dashboard_service import create_notification

            provider_name = provider.get("name") if provider else "Learning Provider"
            summary = f"{imported} new, {updated} updated, {archived} archived, {deleted} deleted."
            if failed:
                summary += f" {failed} failed."

            await create_notification(
                recipient_id=None,
                recipient_role="super_admin",
                notif_type="course_import_completed",
                title="Course import completed",
                message=f"Imported courses for {provider_name}: {summary}",
                link="/dashboard/recruiter/learning?tab=providers",
            )
        except Exception:  # noqa: BLE001
            pass

    # ------------------------------------------------------------------ #
    # API sync (uses the same upsert engine)
    # ------------------------------------------------------------------ #
    async def sync_from_api(
        self,
        current_user: CurrentUser,
        provider_id: str,
        *,
        missing_action: str = DEFAULT_MISSING_ACTION,
    ) -> dict:
        """API-provider sync entry point.

        Fetches courses from either a legacy connector (Coursera / Microsoft
        Learn) or the generic config-driven importer, then runs the same
        compare -> insert -> update path as Excel imports. Writes use batched
        ``bulk_write`` so large catalogs (capped at API_MAX_COURSES) finish
        within a normal request timeout.
        """
        if not ObjectId.is_valid(provider_id):
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Provider not found.")
        provider = await database.learning_providers.find_one({"_id": ObjectId(provider_id)})
        if not provider:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Provider not found.")
        if not provider.get("active", True):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Provider is inactive.")

        from app.services.generic_api_provider_service import API_MAX_COURSES

        courses = await self._fetch_provider_courses(provider)
        capped = len(courses) >= API_MAX_COURSES

        now = _now()
        imported = updated = failed = 0
        created_ids: list[str] = []
        updated_before: list[dict] = []

        provider_id_str = str(provider["_id"])
        existing_docs = await database.learning_courses.find(
            {"$or": [
                {"provider_id": provider_id_str},
                {"provider": {"$regex": f"^{re.escape(provider.get('name') or '')}$", "$options": "i"}},
            ]}
        ).to_list(length=100000)
        by_external: dict[str, dict] = {}
        by_url: dict[str, dict] = {}
        by_title: dict[str, dict] = {}
        for doc in existing_docs:
            ext = doc.get("external_id")
            if ext:
                by_external.setdefault(str(ext), doc)
            url = doc.get("url")
            if url:
                by_url.setdefault(_normalize(url), doc)
            title = doc.get("title")
            if title:
                by_title.setdefault(_normalize(title), doc)

        ops: list[Any] = []
        op_kinds: list[str] = []  # "insert" | "update" aligned with ops
        pending_inserts: list[dict] = []
        max_snapshots = 500
        planned_imported = planned_updated = prepare_failed = 0

        for course in courses:
            record = {
                "external_id": course.get("external_id") or course.get("id"),
                "title": course.get("title"),
                "url": course.get("url"),
                "description": course.get("description"),
                "duration_minutes": course.get("duration_minutes"),
                "category": course.get("category") or "",
                "competency": course.get("competency") or "",
                "designation": course.get("designation") or "",
                "learning_month": course.get("learning_month") or "",
                "instructor": course.get("instructor"),
                "tags": course.get("tags") or [],
                "skills": course.get("skills") or [],
                "difficulty": course.get("difficulty"),
                "provider": provider.get("name"),
                "provider_id": provider_id_str,
            }
            if not record.get("title"):
                continue

            existing = None
            if record.get("external_id"):
                existing = by_external.get(str(record["external_id"]))
            if existing is None and record.get("url"):
                existing = by_url.get(_normalize(record["url"]))
            if existing is None and record.get("title"):
                existing = by_title.get(_normalize(record["title"]))

            try:
                if existing:
                    if len(updated_before) < max_snapshots:
                        snapshot = {k: existing.get(k) for k in (
                            "title", "url", "designation", "learning_month", "category",
                            "competency", "description", "duration_minutes", "provider",
                            "provider_id", "external_id", "archived", "instructor", "tags",
                        )}
                        updated_before.append({"_id": str(existing["_id"]), "snapshot": snapshot})
                    updated_doc = {
                        **existing,
                        **record,
                        "archived": False,
                        "updated_at": now,
                        "updated_by_id": current_user.id,
                    }
                    updated_doc = await self._finalize_course_doc(updated_doc)
                    set_doc = {k: v for k, v in updated_doc.items() if k != "_id"}
                    ops.append(UpdateOne({"_id": existing["_id"]}, {"$set": set_doc}))
                    op_kinds.append("update")
                    planned_updated += 1
                else:
                    new_doc = {
                        **record,
                        "archived": False,
                        "source_kind": "api",
                        "source_filename": None,
                        "source_row": None,
                        "created_by_id": current_user.id,
                        "updated_by_id": current_user.id,
                        "created_at": now,
                        "updated_at": now,
                    }
                    new_doc = await self._finalize_course_doc(new_doc)
                    pending_inserts.append(new_doc)
                    ops.append(InsertOne(new_doc))
                    op_kinds.append("insert")
                    planned_imported += 1
            except Exception:  # noqa: BLE001
                prepare_failed += 1

        imported = updated = write_failed = 0
        batch_size = 500

        def _apply_chunk_success(kinds: list[str], inserted_ids: dict | None = None) -> None:
            nonlocal imported, updated
            insert_pos = 0
            for kind in kinds:
                if kind == "insert":
                    imported += 1
                    if inserted_ids is not None and insert_pos in inserted_ids:
                        created_ids.append(str(inserted_ids[insert_pos]))
                    insert_pos += 1
                else:
                    updated += 1

        for i in range(0, len(ops), batch_size):
            chunk = ops[i : i + batch_size]
            kinds = op_kinds[i : i + batch_size]
            try:
                result = await database.learning_courses.bulk_write(chunk, ordered=False)
                _apply_chunk_success(kinds, getattr(result, "inserted_ids", None) or {})
            except BulkWriteError as exc:
                details = exc.details or {}
                err_indices = {int(err.get("index", -1)) for err in (details.get("writeErrors") or [])}
                # Partial success still writes; count successes vs failures by index.
                insert_pos = 0
                local_inserted = details.get("insertedIds") or {}
                # Normalize insertedIds keys to int
                local_inserted = {int(k): v for k, v in local_inserted.items()}
                for j, kind in enumerate(kinds):
                    if j in err_indices:
                        write_failed += 1
                        if kind == "insert":
                            insert_pos += 1
                        continue
                    if kind == "insert":
                        imported += 1
                        if j in local_inserted:
                            created_ids.append(str(local_inserted[j]))
                        insert_pos += 1
                    else:
                        updated += 1
            except Exception:  # noqa: BLE001
                write_failed += len(chunk)

        if imported and len(created_ids) < imported:
            for doc in pending_inserts:
                if doc.get("_id") is not None:
                    cid = str(doc["_id"])
                    if cid not in created_ids:
                        created_ids.append(cid)

        failed = prepare_failed + write_failed
        # Prefer actual write counts; fall back to planned if writes reported nothing
        # but no failures (shouldn't happen).
        if imported == 0 and updated == 0 and write_failed == 0:
            imported, updated = planned_imported, planned_updated

        cap_note = (
            f" (capped at {API_MAX_COURSES} courses per sync)"
            if capped
            else ""
        )
        message = (
            f"API sync completed: {imported} new course(s), {updated} updated"
            f"{cap_note}."
        )
        history_id = await self._record_history(
            provider=provider,
            current_user=current_user,
            import_type="api",
            filename=None,
            rows_total=len(courses),
            rows_imported=imported,
            rows_updated=updated,
            rows_failed=failed,
            rows_archived=0,
            rows_deleted=0,
            validation_summary={
                "api_sync": True,
                "total_fetched": len(courses),
                "capped_at": API_MAX_COURSES if capped else None,
                "api_connector": provider.get("api_connector") or provider.get("slug"),
            },
            message=message,
            created_ids=created_ids,
            updated_before=updated_before,
            archived_ids=[],
            deleted_ids=[],
        )
        return {
            "message": message,
            "imported": imported,
            "updated": updated,
            "failed": failed,
            "capped": capped,
            "cap": API_MAX_COURSES if capped else None,
            "history_id": history_id,
        }

    async def _fetch_provider_courses(self, provider: dict) -> list[dict]:
        """Fetch courses using a legacy connector or the generic importer.

        Resolution order:
          1. Built-in connector via ``api_connector`` (or legacy slug).
          2. Generic importer driven by the provider's saved api_config.
        """
        from app.services.generic_api_provider_service import API_MAX_COURSES

        adapters = self._api_adapters()
        connector_key = (provider.get("api_connector") or provider.get("slug") or "").strip()
        adapter = adapters.get(connector_key)
        if adapter is not None:
            return await adapter.fetch_courses(max_items=API_MAX_COURSES)

        api_config = provider.get("api_config") or {}
        if not (api_config.get("endpoint") or "").strip():
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=(
                    f"No API configuration is set for '{provider.get('name')}'. "
                    "Open the provider and configure the API endpoint and mapping."
                ),
            )

        from app.services.generic_api_provider_service import ApiImportError, fetch_courses

        try:
            return await fetch_courses(api_config, max_items=API_MAX_COURSES)
        except ApiImportError as exc:
            raise HTTPException(status_code=exc.status_code, detail=exc.message)

    def _api_adapters(self) -> dict[str, Any]:
        """Registry of legacy API provider connectors.

        Only providers whose APIs need special handling (pagination, multi-type
        fetches, URL construction) that the generic config cannot express
        belong here. Keys match ``api_connector`` / slug values stored on the
        provider document. Every new API provider added from the UI is fetched
        by the generic importer instead.
        """
        from app.services.api_connectors import CourseraConnector, MicrosoftLearnConnector

        return {
            "coursera": CourseraConnector(),
            "microsoft-learn": MicrosoftLearnConnector(),
        }


import_engine_service = ImportEngineService()
