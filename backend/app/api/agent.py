"""AI Agent API — conversational hiring & onboarding automation.

Available to recruiters (invitations / offers / joining letters) and to
candidates/employees (self-service onboarding). Everything the agent does
goes through the same permission-checked service methods used by the
regular UI, so the agent can never do more than the signed-in user could
already do by hand.
"""

from __future__ import annotations

import io
from fastapi import APIRouter, File, HTTPException, Query, UploadFile

from app.core.database import database
from app.core.rbac import CurrentUser
from app.core.security import RequireUser
from app.schemas.agent import AgentChatRequest, AgentResetRequest
from app.services import agent_tools
from app.services.agent_service import _load_or_create_session, _now_iso, _save_messages, agent_service

router = APIRouter(prefix="/api/agent", tags=["AI Agent"])

ALLOWED_ROLES = ("candidate", "employee", "recruiter", "super_admin")


def _assert_agent_role(user: CurrentUser) -> None:
    if user.role not in ALLOWED_ROLES:
        raise HTTPException(status_code=403, detail="The AI agent is not available for this account type.")


@router.post("/chat")
async def chat(request: AgentChatRequest, current_user: RequireUser):
    """Send a message to the role-appropriate agent (recruiter or onboarding)."""
    _assert_agent_role(current_user)
    if not request.message and not request.session_id:
        raise HTTPException(status_code=400, detail="A message is required to start a conversation.")
    return await agent_service.chat(current_user, request.message, request.session_id)


@router.get("/sessions")
async def list_sessions(current_user: RequireUser):
    _assert_agent_role(current_user)
    return await agent_service.list_sessions(current_user)


@router.get("/history")
async def get_history(session_id: str, current_user: RequireUser):
    _assert_agent_role(current_user)
    return await agent_service.get_history(current_user, session_id)


@router.post("/reset")
async def reset_session(request: AgentResetRequest, current_user: RequireUser):
    _assert_agent_role(current_user)
    if request.session_id:
        await database.agent_conversations.delete_one({"session_id": request.session_id, "user_id": current_user.id})
    return {"message": "Conversation cleared."}


@router.post("/recruiter/bulk-invite")
async def bulk_invite_from_spreadsheet(
    current_user: RequireUser,
    file: UploadFile = File(...),
    session_id: str | None = Query(None),
):
    """Parse an uploaded .xlsx / .csv roster and invite every valid row.

    Required columns (same as manual Create invitation):
      email, full_name (or name), job_title (or designation/title), department
    Optional: office_location, start_date (YYYY-MM-DD), expires_in_days, phone (ignored)

    If required headers or row values are missing, no invitations are sent — a
    validation report is returned so the recruiter can fix the file first.
    """
    if current_user.role not in ("recruiter", "super_admin"):
        raise HTTPException(status_code=403, detail="Only recruiters can bulk-invite candidates.")

    resolved_session = session_id

    filename = (file.filename or "").lower()
    is_csv = filename.endswith(".csv")
    is_xlsx = filename.endswith(".xlsx") or filename.endswith(".xlsm")
    if not (is_csv or is_xlsx):
        raise HTTPException(status_code=400, detail="Please upload a .xlsx or .csv spreadsheet.")

    raw = await file.read()
    if len(raw) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File is too large (5MB limit).")

    try:
        if is_csv:
            import csv

            text = raw.decode("utf-8-sig", errors="replace")
            reader = csv.reader(io.StringIO(text))
            rows = list(reader)
            if not rows:
                raise HTTPException(status_code=400, detail="The CSV file is empty.")
            header = [str(h).strip().lower() if h else "" for h in rows[0]]
            data_rows = list(rows[1:])
        else:
            from openpyxl import load_workbook

            workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            sheet = workbook.active
            rows_iter = sheet.iter_rows(values_only=True)
            header = [str(h).strip().lower() if h else "" for h in next(rows_iter, [])]
            data_rows = list(rows_iter)
    except HTTPException:
        raise
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=400, detail=f"Could not read the spreadsheet: {exc}") from exc

    def _col(*names: str) -> int | None:
        for name in names:
            if name in header:
                return header.index(name)
        return None

    def _cell(row, idx: int | None) -> str:
        if idx is None or idx >= len(row) or row[idx] is None:
            return ""
        return str(row[idx]).strip()

    idx_email = _col("email", "email address")
    idx_name = _col("full_name", "name", "candidate name", "full name")
    idx_title = _col("job_title", "designation", "title", "job title", "position")
    idx_dept = _col("department", "dept")
    idx_office = _col("office_location", "office", "location")
    idx_start = _col("start_date", "start date", "joining date")
    idx_expires = _col("expires_in_days", "expires", "expiry days")

    missing_headers: list[str] = []
    if idx_email is None:
        missing_headers.append("email")
    if idx_name is None:
        missing_headers.append("full_name (or name)")
    if idx_title is None:
        missing_headers.append("job_title (or designation)")
    if idx_dept is None:
        missing_headers.append("department")

    # Persist a chat-friendly validation report helper
    async def _report(message: str, *, ok: bool, extra: dict | None = None) -> dict:
        payload = {"session_id": None, "message": message, "ok": ok, "sent": [], "failed": [], **(extra or {})}
        convo = await _load_or_create_session(current_user, resolved_session)
        await _save_messages(
            convo["session_id"],
            current_user.id,
            [
                {"role": "user", "content": f"[Uploaded spreadsheet: {file.filename}]", "created_at": _now_iso()},
                {"role": "assistant", "content": message, "created_at": _now_iso(), "meta": {"validation": extra or {}}},
            ],
        )
        payload["session_id"] = convo["session_id"]
        return payload

    if missing_headers:
        found = ", ".join(h for h in header if h) or "(none)"
        message = (
            f"I checked `{file.filename}` before sending any invites — it's missing required columns "
            f"(same as Create invitation):\n"
            f"• Missing: {', '.join(missing_headers)}\n"
            f"• Found headers: {found}\n\n"
            "Required columns: email, full_name (or name), job_title (or designation), department.\n"
            "Optional: office_location, start_date, expires_in_days.\n"
            "(Phone is collected when the candidate registers — it is not used for invitations.)\n\n"
            "Please update the spreadsheet and upload again. No invitations were sent."
        )
        return await _report(
            message,
            ok=False,
            extra={"missing_headers": missing_headers, "found_headers": [h for h in header if h]},
        )

    candidates: list[dict] = []
    row_issues: list[dict] = []
    for i, row in enumerate(data_rows, start=2):  # 1-based sheet row (header is row 1)
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        email = _cell(row, idx_email)
        full_name = _cell(row, idx_name)
        job_title = _cell(row, idx_title)
        department = _cell(row, idx_dept)
        missing_fields = [
            label
            for label, value in (
                ("email", email),
                ("full_name", full_name),
                ("job_title/designation", job_title),
                ("department", department),
            )
            if not value or (label != "email" and len(value) < 2)
        ]
        if missing_fields:
            row_issues.append(
                {
                    "row": i,
                    "email": email or None,
                    "missing_fields": missing_fields,
                }
            )
            continue
        expires_raw = _cell(row, idx_expires)
        candidates.append(
            {
                "email": email,
                "full_name": full_name,
                "job_title": job_title,
                "department": department,
                "office_location": _cell(row, idx_office) or None,
                "start_date": _cell(row, idx_start) or None,
                "expires_in_days": int(expires_raw) if expires_raw.isdigit() else 7,
            }
        )

    if row_issues and not candidates:
        issue_lines = "; ".join(
            f"row {x['row']}" + (f" ({x['email']})" if x.get("email") else "") + f": missing {', '.join(x['missing_fields'])}"
            for x in row_issues[:8]
        )
        more = f" (+{len(row_issues) - 8} more)" if len(row_issues) > 8 else ""
        message = (
            f"I checked `{file.filename}` — every data row is missing required invitation fields "
            f"(designation and department are required, same as Create invitation).\n"
            f"Issues: {issue_lines}{more}\n\n"
            "Please add job_title/designation and department for each candidate, then upload again. "
            "No invitations were sent."
        )
        return await _report(message, ok=False, extra={"row_issues": row_issues, "valid_rows": 0})

    if row_issues:
        # Partial file: block entirely so recruiter fixes the sheet (safer than inviting a subset silently)
        issue_lines = "; ".join(
            f"row {x['row']}" + (f" ({x['email']})" if x.get("email") else "") + f": missing {', '.join(x['missing_fields'])}"
            for x in row_issues[:8]
        )
        more = f" (+{len(row_issues) - 8} more)" if len(row_issues) > 8 else ""
        message = (
            f"I checked `{file.filename}` before inviting anyone.\n"
            f"• Valid rows ready to invite: {len(candidates)}\n"
            f"• Incomplete rows: {len(row_issues)} — {issue_lines}{more}\n\n"
            "Required per row (same as Create invitation): email, full_name, job_title/designation, department.\n"
            "Fix the incomplete rows and upload again. No invitations were sent."
        )
        return await _report(
            message,
            ok=False,
            extra={"row_issues": row_issues, "valid_rows": len(candidates), "blocked": True},
        )

    if not candidates:
        message = (
            f"`{file.filename}` has the right headers but no candidate rows. "
            "Add at least one row with email, full_name, job_title/designation, and department."
        )
        return await _report(message, ok=False, extra={"valid_rows": 0})

    result = await agent_tools.run_tool(current_user, "bulk_invite", {"candidates": candidates})
    if not result.ok:
        raise HTTPException(status_code=400, detail=result.error or "Bulk invite failed.")

    sent = result.data["sent"]
    failed = result.data["failed"]
    summary = (
        f"Checked `{file.filename}` — all {len(candidates)} row(s) had the required fields "
        f"(email, name, designation, department). "
        f"Invited {len(sent)}, failed {len(failed)}."
    )
    if failed:
        summary += " Failures: " + "; ".join(f"{f['email']}: {f['error']}" for f in failed[:5])

    convo = await _load_or_create_session(current_user, resolved_session)
    await _save_messages(
        convo["session_id"],
        current_user.id,
        [
            {"role": "user", "content": f"[Uploaded spreadsheet: {file.filename}]", "created_at": _now_iso()},
            {"role": "assistant", "content": summary, "created_at": _now_iso(), "meta": {"tool_data": result.data}},
        ],
    )

    return {
        "session_id": convo["session_id"],
        "message": summary,
        "ok": True,
        "sent": sent,
        "failed": failed,
    }